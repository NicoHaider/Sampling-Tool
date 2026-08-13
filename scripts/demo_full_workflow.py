"""End-to-End-Demo: Engagement -> Excel-Import -> Sampling -> Export -> PDF.

Aufruf:
    python scripts/demo_full_workflow.py
    python scripts/demo_full_workflow.py --output /tmp/demo

Alle Artefakte landen unter `./demo_output/` (bzw. `--output`):
    - engagement.db          – frische SQLite mit aktuellem Schema
    - source_data.xlsx       – generierte Quelldatei (200 Buchungssätze)
    - DemoSimple_ID001_BDO_sampling_<datum>.xlsx      – SimpleSampler (25 Zeilen)
    - DemoStratified_ID002_BDO_sampling_<datum>.xlsx  – geschichtet nach Land (15)
    - audit_trail.pdf        – PDF-Report aller Audit-Events

Das Skript dient gleichzeitig als manueller Smoke-Test für den
gesamten Sprint-3-Datenpfad und als ausführbare Architektur-Doku.
"""

from __future__ import annotations

import argparse
import shutil
from dataclasses import replace
from datetime import datetime
from pathlib import Path

from _script_io import force_utf8_stdout
from openpyxl import Workbook

from sampling_tool.audit.logger import AuditLogger
from sampling_tool.core.models import (
    Engagement,
    SampleConfig,
    SamplingMethod,
    StratifyMode,
)
from sampling_tool.core.sampling import create_sampler
from sampling_tool.io import AuditTrailPDF, ExcelExporter, ExcelImporter
from sampling_tool.persistence.database import Database
from sampling_tool.persistence.repositories import (
    AuditRepo,
    DatasetRepo,
    EngagementRepo,
    SampleRepo,
)

DEFAULT_DEMO_DIR = Path("demo_output")


def step(n: int, title: str) -> None:
    print(f"\n[{n}] {title}")


def make_source_xlsx(path: Path, rows: int = 200) -> None:
    """Generiert eine plausible Buchungssatz-Datei für die Demo."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Buchungen"
    ws.append(["BuchungsID", "Datum", "Betrag", "Land", "Konto"])
    countries = ["AUT", "DEU", "CHE", "ITA", "FRA"]
    for i in range(1, rows + 1):
        ws.append(
            [
                f"B{i:05d}",
                datetime(2026, 1 + (i % 12), 1 + (i % 27)),
                100.0 + (i * 7.13) % 9000,
                countries[i % len(countries)],
                f"4{(i % 999):03d}",
            ]
        )
    wb.save(path)


def main(demo_dir: Path = DEFAULT_DEMO_DIR) -> None:
    force_utf8_stdout()

    if demo_dir.exists():
        shutil.rmtree(demo_dir)
    demo_dir.mkdir(parents=True)

    step(1, "Frische SQLite-Datenbank anlegen + migrieren")
    db_path = demo_dir / "engagement.db"
    db = Database(db_path)
    db.migrate()
    print(f"    -> {db_path} (Schema-Version {db.schema_version()})")

    step(2, "Engagement erstellen")
    engagement_repo = EngagementRepo(db.connect())
    engagement = engagement_repo.get_or_create(
        Engagement(
            auditor_name="Anna Auditorin",
            auditor_position="Senior Auditor",
            client_name="ACME GmbH",
            audit_type="ISAE 3402 Typ II",
        )
    )
    assert engagement.id is not None
    audit_logger = AuditLogger(
        AuditRepo(db.connect()),
        user_name="anna",
        engagement_id=engagement.id,
    )
    print(f"    -> Engagement #{engagement.id} für {engagement.client_name}")

    step(3, "Quelldatei (Excel) generieren und importieren")
    source_xlsx = demo_dir / "source_data.xlsx"
    make_source_xlsx(source_xlsx, rows=200)

    importer = ExcelImporter()
    result = importer.import_file(source_xlsx)
    # Sprint 11.1: Importer trennt Dataset (Metadaten) von rows.
    dataset = replace(result.dataset, engagement_id=engagement.id)
    # Sprint 11.3: `result.rows` ist ein einmalig konsumierbarer Iterator –
    # `create` zieht ihn durch, danach ist er leer. Alles Weitere liest die
    # Rows deshalb frisch aus dem Repo.
    dataset = DatasetRepo(db.connect()).create(dataset, result.rows)
    audit_logger.log_import(dataset)
    # Sprint 11.5: die Compat-Properties sind weg, Stats liegen unter `.stats`
    # und sind erst nach vollem Verbrauch des Generators aussagekräftig.
    print(
        f"    -> {dataset.row_count} Zeilen, "
        f"{len(dataset.columns)} Spalten, "
        f"{result.stats.skipped_rows} übersprungen"
    )
    assert dataset.id is not None
    dataset_repo = DatasetRepo(db.connect())

    step(4, "Simple-Sampling (25 von 200) ziehen + persistieren")
    simple_cfg = SampleConfig(
        method=SamplingMethod.SIMPLE,
        size=25,
        seed=42,
    )
    simple_result = create_sampler(simple_cfg).sample(
        dataset_repo.iter_rows(dataset.id), population_size=dataset.row_count
    )
    simple_id = SampleRepo(db.connect()).create_from_result(simple_result, dataset.id, "anna")
    audit_logger.log_sampling(simple_result, sample_id=simple_id, dataset_id=dataset.id)
    print(f"    -> Sample #{simple_id}, gezogen: {simple_result.actual_size} Zeilen")

    step(5, "Stratified-Sampling (15 Zeilen, geschichtet nach Land)")
    strat_cfg = SampleConfig(
        method=SamplingMethod.STRATIFIED,
        size=15,
        seed=99,
        stratum_field="Land",
        stratify_mode=StratifyMode.PROPORTIONAL,
    )
    strat_result = create_sampler(strat_cfg).sample(
        dataset_repo.iter_rows(dataset.id), population_size=dataset.row_count
    )
    strat_id = SampleRepo(db.connect()).create_from_result(strat_result, dataset.id, "anna")
    audit_logger.log_sampling(strat_result, sample_id=strat_id, dataset_id=dataset.id)
    print(f"    -> Sample #{strat_id}, gezogen: {strat_result.actual_size} Zeilen")

    step(6, "Beide Samples nach Excel exportieren")
    exporter = ExcelExporter()
    out_simple = exporter.export_sample(
        sample=simple_result,
        dataset=dataset,
        dataset_repo=dataset_repo,
        columns=["BuchungsID", "Datum", "Betrag", "Land"],
        output_dir=demo_dir,
        custom_name="DemoSimple",
        custom_id="001",
        engagement=engagement,
    )
    audit_logger.log_export(simple_id, out_simple, simple_result.actual_size)
    print(f"    -> {out_simple.name}")

    out_strat = exporter.export_sample(
        sample=strat_result,
        dataset=dataset,
        dataset_repo=dataset_repo,
        columns=["BuchungsID", "Land", "Konto", "Betrag"],
        output_dir=demo_dir,
        custom_name="DemoStratified",
        custom_id="002",
        engagement=engagement,
    )
    audit_logger.log_export(strat_id, out_strat, strat_result.actual_size)
    print(f"    -> {out_strat.name}")

    step(7, "AuditTrail-PDF generieren")
    events = AuditRepo(db.connect()).list_for_engagement(engagement.id, limit=200)
    pdf_path = AuditTrailPDF().render(
        engagement=engagement,
        events=events,
        output_path=demo_dir / "audit_trail.pdf",
    )
    print(f"    -> {pdf_path.name} ({len(events)} Events)")

    step(8, "Demo abgeschlossen")
    print(f"    -> Alle Artefakte unter: {demo_dir.resolve()}")
    db.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_DEMO_DIR,
        help="Zielverzeichnis für alle Demo-Artefakte (Default: demo_output)",
    )
    return parser.parse_args()


if __name__ == "__main__":
    # Zweiter Aufruf mit Absicht: anders als in den uebrigen Skripten laeuft
    # `parse_args()` hier VOR `main()`, und `--help` geht nach STDOUT. Ein Guard
    # erst in `main()` kaeme fuer die argparse-Ausgabe zu spaet. Der Aufruf ist
    # idempotent.
    force_utf8_stdout()
    main(parse_args().output)
