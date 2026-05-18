"""Performance-Discovery-Script für große Datasets.

Generiert synthetische Audit-Datensätze in mehreren Größen, misst die
wichtigsten Pipeline-Phasen (Import, DB-Speicherung, Tabelle-Anzeige,
Sampling, Filter, Highlight, Export, PDF) und schreibt einen
Markdown-Bericht.

Aufruf:
    python scripts/perf_probe.py
    python scripts/perf_probe.py --sizes 10000 100000 1000000
    python scripts/perf_probe.py --sizes 100 --quick --audit-events 10

Discovery, keine Optimierungen. Bottlenecks gehen in Sprint 10.2.
"""

from __future__ import annotations

import argparse
import contextlib
import gc
import os
import platform
import shutil
import subprocess
import sys
import tempfile
import time
import tracemalloc
from collections.abc import Iterator
from dataclasses import dataclass, field, replace
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

# Offscreen-Plattform MUSS gesetzt sein, bevor PyQt6 geladen wird.
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import numpy as np
from openpyxl import Workbook

# Repository-Root in sys.path aufnehmen, damit das Script auch direkt
# (ohne `python -m`) aus dem Repo-Root funktioniert.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT / "src") not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT / "src"))

from PyQt6.QtCore import QCoreApplication  # noqa: E402
from PyQt6.QtWidgets import QApplication  # noqa: E402

from sampling_tool.audit.logger import AuditLogger  # noqa: E402
from sampling_tool.core.models import (  # noqa: E402
    AuditEvent,
    Engagement,
    SampleConfig,
    SamplingMethod,
    StratifyMode,
)
from sampling_tool.core.sampling import SimpleSampler, create_sampler  # noqa: E402
from sampling_tool.io import (  # noqa: E402
    AuditTrailPDF,
    ExcelExporter,
    ExcelImporter,
)
from sampling_tool.io.html_report import HtmlReportGenerator  # noqa: E402
from sampling_tool.io.multi_report_exporter import MultiSheetReportExporter  # noqa: E402
from sampling_tool.persistence.database import Database  # noqa: E402
from sampling_tool.persistence.repositories import (  # noqa: E402
    AuditRepo,
    DatasetRepo,
    EngagementRepo,
    SampleRepo,
)
from sampling_tool.ui.widgets.data_table import DataTableView  # noqa: E402

try:  # pragma: no cover - psutil ist optional
    import psutil

    _HAS_PSUTIL = True
except ImportError:  # pragma: no cover - psutil ist optional
    _HAS_PSUTIL = False


# Sprint 12.1 / P-007: Pipeline-Total-Label für Import + DB-Speicherung.
# Seit Sprint 11.3 (Streaming-Import) wandert die Cell-Coercion + JSON-Encode-
# Arbeit aus der Import-Phase in den DB-Insert-Generator. Die historischen
# Einzeltargets (Import < 60 s, DB < 30 s) sind dadurch strukturell verschoben:
# Import misst nur noch den Header-Pass, DB-Speicherung trägt die Coerce-Last.
# Die Bewertung erfolgt deshalb am Pipeline-Total – die Einzelphasen bleiben
# nur in `LEGACY_PRE_STREAMING_TARGETS_1M_SECONDS` für Sprint-10.x-Vergleich.
PIPELINE_TOTAL_LABEL: str = "Import + DB-Speicherung (Pipeline-Total)"
PIPELINE_TOTAL_PHASES: tuple[str, ...] = ("Import", "DB-Speicherung")

# Aktuelle Soft-Targets bei 1M Zeilen (Sprint 11.3+).
SOFT_TARGETS_1M_SECONDS: dict[str, float] = {
    PIPELINE_TOTAL_LABEL: 90.0,
    "Tabelle-Anzeige": 5.0,
    "Sampling Simple": 10.0,
    "Sampling Cluster": 15.0,
    "Sampling Stratified": 15.0,
    "Filter-Toggle (an)": 2.0,
    "Filter-Toggle (aus)": 2.0,
    "Highlight": 2.0,
    "Excel-Export (Sample)": 60.0,
    "Excel-Report (Multi-Sheet)": 60.0,
    "HTML-Report": 30.0,
    "AuditTrail-PDF": 30.0,
}

# Pre-Streaming-Einzeltargets – nur für historische Sprint-10.x-Vergleichbarkeit.
# Werden NICHT in `detect_violations` ausgewertet.
LEGACY_PRE_STREAMING_TARGETS_1M_SECONDS: dict[str, float] = {
    "Import": 60.0,
    "DB-Speicherung": 30.0,
}

# 15 Spalten gemischt – orientiert sich an typischen Buchungssatz-Daten.
COLUMN_NAMES: tuple[str, ...] = (
    "buchung_id",
    "konto",
    "gegenkonto",
    "betrag",
    "soll_haben",
    "buchungsdatum",
    "valuta",
    "beleg_datum",
    "buchungstext",
    "belegnummer",
    "kostenstelle",
    "auftrag",
    "steuersatz",
    "waehrung",
    "status",
)

# Werte mit kleiner Kardinalität – sorgt dafür, dass Cluster/Stratified
# realistisch ziehen können.
_STATUSES: tuple[str, ...] = ("offen", "bezahlt", "verbucht", "storniert", "geprüft")
_KOSTENSTELLEN: tuple[str, ...] = tuple(f"KS{i:02d}" for i in range(1, 11))  # 10 distinct
_WAEHRUNGEN: tuple[str, ...] = ("EUR", "USD", "CHF", "GBP")
_SOLL_HABEN: tuple[str, ...] = ("S", "H")


# ---------------------------------------------------------------------------
# Measurement-Helper
# ---------------------------------------------------------------------------


@dataclass
class Measurement:
    """Eine einzelne Messung (Phase + Zeit + RAM)."""

    label: str
    elapsed_s: float = 0.0
    peak_tracemalloc_mb: float = 0.0
    rss_delta_mb: float | None = None
    note: str = ""


@dataclass
class SizeResult:
    """Ergebnisse aller Phasen für eine Dataset-Größe."""

    size: int
    measurements: list[Measurement] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@contextlib.contextmanager
def measured(label: str) -> Iterator[Measurement]:
    """Misst Zeit (perf_counter) und Peak-RAM (tracemalloc + RSS-Delta).

    `tracemalloc` deckt nur den Python-Heap ab – Numpy-Arrays werden
    nicht voll erfasst. `psutil.Process().memory_info().rss` liefert
    das RSS-Delta als Cross-Check (falls psutil installiert ist).
    """
    measurement = Measurement(label=label)
    proc = psutil.Process() if _HAS_PSUTIL else None
    rss_before = proc.memory_info().rss if proc is not None else 0
    gc.collect()
    tracemalloc.start()
    t0 = time.perf_counter()
    try:
        yield measurement
    finally:
        elapsed = time.perf_counter() - t0
        _, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        rss_after = proc.memory_info().rss if proc is not None else 0
        measurement.elapsed_s = elapsed
        measurement.peak_tracemalloc_mb = peak / 1024 / 1024
        if proc is not None:
            measurement.rss_delta_mb = (rss_after - rss_before) / 1024 / 1024


# ---------------------------------------------------------------------------
# Datengenerierung
# ---------------------------------------------------------------------------


def generate_synthetic_xlsx(n_rows: int, target: Path, *, seed: int = 42) -> None:
    """Schreibt eine .xlsx mit `n_rows` synthetischen Buchungssätzen.

    15 Spalten gemischt: int (5), date (3), float (3), string (4).
    `openpyxl` im write_only-Modus hält den RAM-Footprint flach –
    Werte werden Zeile für Zeile geflusht statt das ganze Sheet zu puffern.
    """
    rng = np.random.default_rng(seed)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Buchungen")
    ws.append(list(COLUMN_NAMES))

    # Numpy generiert alle Spalten en bloc – speichersparsam genug für
    # die Größenordnungen, die wir hier messen wollen (bis ~5M Zeilen).
    base_date = date(2025, 1, 1)
    konten = rng.integers(4000, 8000, size=n_rows)
    gegenkonten = rng.integers(4000, 8000, size=n_rows)
    betraege = rng.uniform(10.0, 9999.99, size=n_rows).round(2)
    soll_haben = rng.choice(_SOLL_HABEN, size=n_rows)
    bu_offsets = rng.integers(0, 365, size=n_rows)
    valuta_offsets = bu_offsets + rng.integers(0, 7, size=n_rows)
    beleg_offsets = bu_offsets - rng.integers(0, 5, size=n_rows)
    belegnummern = rng.integers(100000, 999999, size=n_rows)
    kostenstellen = rng.choice(_KOSTENSTELLEN, size=n_rows)
    auftrag_ids = rng.integers(1000, 9999, size=n_rows)
    steuersaetze = rng.choice([0.0, 7.0, 10.0, 13.0, 20.0], size=n_rows)
    waehrungen = rng.choice(_WAEHRUNGEN, size=n_rows)
    status = rng.choice(_STATUSES, size=n_rows)

    for i in range(n_rows):
        ws.append(
            [
                int(i + 1),
                int(konten[i]),
                int(gegenkonten[i]),
                float(betraege[i]),
                str(soll_haben[i]),
                base_date + timedelta(days=int(bu_offsets[i])),
                base_date + timedelta(days=int(valuta_offsets[i])),
                base_date + timedelta(days=int(beleg_offsets[i])),
                f"Buchung Nr. {i + 1}",
                f"BN{int(belegnummern[i])}",
                str(kostenstellen[i]),
                f"AUF{int(auftrag_ids[i])}",
                float(steuersaetze[i]),
                str(waehrungen[i]),
                str(status[i]),
            ]
        )

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)


# ---------------------------------------------------------------------------
# Phase-Runner
# ---------------------------------------------------------------------------


def build_synthetic_events(
    engagement_id: int,
    sample_ids: list[int],
    *,
    count: int,
    user_name: str = "perf",
) -> list[AuditEvent]:
    """Baut künstlich `count` AuditEvents im Speicher (ohne DB-Insert).

    Mix aus sampling/filter/reset – damit der PDF-Renderer realistisch
    viele Event-Typen formatieren muss.
    """
    events: list[AuditEvent] = []
    types = ("sampling", "import", "export", "reset", "undo", "redo")
    base = datetime.now(UTC)
    for i in range(count):
        evt_type = types[i % len(types)]
        sample_id = sample_ids[i % len(sample_ids)] if sample_ids else None
        events.append(
            AuditEvent(
                id=i + 1,
                event_type=evt_type,
                engagement_id=engagement_id,
                user_name=user_name,
                timestamp=base - timedelta(seconds=count - i),
                sample_id=sample_id,
                sample_size=42,
                sample_percent=4.2,
                total_count=1000,
                seed=i,
                details={"synthetic": True, "index": i},
            )
        )
    return events


def _process_qt_events(times: int = 3) -> None:
    """`QCoreApplication.processEvents` ein paar Mal aufrufen.

    Reicht aus, damit Qt-Modell-Updates + Paint-Events in den
    QTableView-Pfad fließen, ohne dass ein echter Event-Loop läuft.
    """
    app = QCoreApplication.instance()
    if app is None:
        return
    for _ in range(times):
        app.processEvents()


def run_probe_for_size(
    size: int,
    work_dir: Path,
    *,
    quick: bool,
    audit_events: int,
) -> SizeResult:
    """Führt alle Mess-Phasen für eine Dataset-Größe aus."""
    result = SizeResult(size=size)
    size_dir = work_dir / f"size_{size}"
    size_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = size_dir / "input.xlsx"
    db_path = size_dir / "engagement.db"
    export_dir = size_dir / "exports"

    # ---- Phase 0: Synthetic-Generator ---------------------------------
    with measured("Setup (xlsx generieren)") as m:
        generate_synthetic_xlsx(size, xlsx_path)
    m.note = f"{xlsx_path.stat().st_size / 1024 / 1024:.1f} MB"
    result.measurements.append(m)

    # ---- Phase 1: Import ----------------------------------------------
    # Sprint 11.3: `import_result.rows` ist ein einmalig konsumierbarer
    # Generator – `len()` knallt. Die Stats (skipped_rows, processed_count)
    # füllen sich erst beim Verbrauch in Phase 2 (DatasetRepo.create), also
    # wird die Note dort gesetzt.
    with measured("Import") as m:
        importer = ExcelImporter()
        import_result = importer.import_file(xlsx_path)
    dataset = import_result.dataset
    rows = import_result.rows
    m.note = "Streaming – Zeilen-Anzahl steht nach Phase 2 fest"
    result.measurements.append(m)

    # ---- Phase 2: DB-Speicherung --------------------------------------
    with measured("DB-Speicherung") as m:
        db = Database(db_path)
        db.migrate()
        engagement_repo = EngagementRepo(db.connect())
        engagement = engagement_repo.get_or_create(
            Engagement(
                auditor_name="Perf Tester",
                auditor_position="Senior",
                client_name="Synthetic Corp",
                audit_type="ISAE 3402",
            )
        )
        assert engagement.id is not None
        dataset = replace(dataset, engagement_id=engagement.id)
        # Konsumiert den Import-Generator; `dataset.row_count` wird auf die
        # tatsächlich persistierte Anzahl korrigiert (Sprint 11.3).
        dataset = DatasetRepo(db.connect()).create(dataset, rows)
    m.note = f"{dataset.row_count:,} rows, {import_result.stats.skipped_rows} skipped"
    result.measurements.append(m)

    audit_logger = AuditLogger(
        AuditRepo(db.connect()),
        user_name="perf",
        engagement_id=engagement.id,
    )
    audit_logger.log_import(dataset)

    # ---- Phase 3: Tabelle-Anzeige (UI) --------------------------------
    with measured("Tabelle-Anzeige") as m:
        table = DataTableView()
        # Sprint 11.2: View liest on-demand via Repo (FIFO-Cache im
        # Model). `rows` werden nicht mehr ans Model übergeben – wir
        # nutzen das frisch befüllte DatasetRepo der Test-DB.
        table.set_dataset(dataset, DatasetRepo(db.connect()))
        _process_qt_events()
    result.measurements.append(m)

    # ---- Phase 4: Sampling --------------------------------------------
    # Bei sehr kleinen Datasets (Smoke-Test mit 100 Zeilen) muss die
    # Sample-Größe runter, sonst greift `SamplingError` ("Größe > Pop").
    sample_size = min(500, max(1, size // 4))
    # ClusterSampler erwartet `size` = ANZAHL der zu ziehenden Cluster,
    # nicht der Rows. Bei _KOSTENSTELLEN gibt es nur 10 Cluster –
    # wir ziehen die Hälfte.
    cluster_count = max(1, len(_KOSTENSTELLEN) // 2)
    sample_specs: list[tuple[str, SampleConfig]] = [
        (
            "Sampling Simple",
            SampleConfig(method=SamplingMethod.SIMPLE, size=sample_size, seed=42),
        ),
    ]
    if not quick:
        sample_specs.append(
            (
                "Sampling Cluster",
                SampleConfig(
                    method=SamplingMethod.CLUSTER,
                    size=cluster_count,
                    seed=43,
                    cluster_field="kostenstelle",
                ),
            )
        )
        sample_specs.append(
            (
                "Sampling Stratified",
                SampleConfig(
                    method=SamplingMethod.STRATIFIED,
                    size=sample_size,
                    seed=44,
                    stratum_field="status",
                    stratify_mode=StratifyMode.PROPORTIONAL,
                ),
            )
        )

    sample_repo = SampleRepo(db.connect())
    sample_ids: list[int] = []
    simple_result = None
    assert dataset.id is not None
    for label, cfg in sample_specs:
        # Sprint 11.4: Sampler arbeitet auf Generator aus dem Repo
        # (kein materialisiertes Row-Tupel mehr). Pro Sampler-Lauf ein
        # frischer Stream – derselbe Pfad, den auch der MainController
        # nutzt. population_size dokumentiert die Universumsgröße.
        sampling_repo = DatasetRepo(db.connect())
        sampler = create_sampler(cfg)
        # Sprint 12.1 / P-002: SimpleSampler ohne Filter konsumiert nur
        # row_ids (kein DatasetRow-Materialize). Spiegelt den Controller-
        # Pfad (`handle_new_sampling`) – sonst würde der Probe-Lauf den
        # RAM-Fix nicht messen.
        with measured(label) as m:
            if isinstance(sampler, SimpleSampler) and cfg.filter_field is None:
                sampled = sampler.sample_ids(
                    sampling_repo.iter_row_ids(dataset.id),
                    population_size=dataset.row_count,
                )
            else:
                sampled = sampler.sample(
                    sampling_repo.iter_rows(dataset.id),
                    population_size=dataset.row_count,
                )
        sid = sample_repo.create_from_result(sampled, dataset.id, "perf")
        sample_ids.append(sid)
        if cfg.method == SamplingMethod.SIMPLE:
            simple_result = sampled
        m.note = f"{sampled.actual_size} rows"
        result.measurements.append(m)

    if simple_result is None:
        # Defensiv – sample_specs hat immer mindestens SimpleSampler.
        raise RuntimeError("SimpleSampler wurde nicht ausgeführt")

    # ---- Phase 5: Filter-Toggle (UI) ----------------------------------
    with measured("Filter-Toggle (an)") as m:
        table.filter_to_rows(simple_result.selected_row_ids)
        _process_qt_events()
    result.measurements.append(m)

    with measured("Filter-Toggle (aus)") as m:
        table.clear_filter()
        _process_qt_events()
    result.measurements.append(m)

    # ---- Phase 6: Highlight (UI) --------------------------------------
    with measured("Highlight") as m:
        table.highlight_rows(simple_result.selected_row_ids)
        _process_qt_events()
    result.measurements.append(m)

    with measured("Clear-Highlight") as m:
        table.clear_highlight()
        _process_qt_events()
    result.measurements.append(m)

    # ---- Phase 7: Export ----------------------------------------------
    # Sprint 11.4: ExcelExporter holt sich die Sample-Rows on-demand
    # via dataset_repo.get_rows_by_ids – kein voll-materialisiertes
    # Row-Tupel mehr im Argument. Bei großen Datasets ist das der
    # entscheidende RAM-Win (nur Sample-Größe statt N).
    export_dir.mkdir(parents=True, exist_ok=True)
    with measured("Excel-Export (Sample)") as m:
        ExcelExporter().export_sample(
            sample=simple_result,
            dataset=dataset,
            dataset_repo=DatasetRepo(db.connect()),
            columns=list(dataset.columns[:8]),
            output_dir=export_dir,
            custom_name="PerfSample",
            custom_id="001",
            engagement=engagement,
        )
    result.measurements.append(m)

    # Für die Multi-Sheet- und HTML-Reports brauchen wir die echten
    # SampleResults; rekonstruieren wäre teuer, also recyceln wir das
    # SimpleResult als Repräsentant.
    samples_for_report = [simple_result]
    with measured("Excel-Report (Multi-Sheet)") as m:
        MultiSheetReportExporter().export(
            engagement=engagement,
            datasets=[dataset],
            samples=samples_for_report,
            audit_events=[],
            output_path=export_dir / "multi_report.xlsx",
        )
    result.measurements.append(m)

    with measured("HTML-Report") as m:
        HtmlReportGenerator().render(
            engagement=engagement,
            datasets=[dataset],
            samples=samples_for_report,
            audit_events=[],
            output_path=export_dir / "report.html",
        )
    result.measurements.append(m)

    # ---- Phase 8: AuditTrail-PDF --------------------------------------
    synthetic_events = build_synthetic_events(engagement.id, sample_ids, count=audit_events)
    with measured("AuditTrail-PDF") as m:
        AuditTrailPDF().render(
            engagement=engagement,
            events=synthetic_events,
            output_path=export_dir / "audit_trail.pdf",
        )
    pdf_size_mb = (export_dir / "audit_trail.pdf").stat().st_size / 1024 / 1024
    m.note = f"{len(synthetic_events)} events, {pdf_size_mb:.1f} MB"
    result.measurements.append(m)

    # ---- Aufräumen ----------------------------------------------------
    db.close()
    del table  # Qt-Widget freigeben, bevor wir das Verzeichnis löschen.
    gc.collect()
    shutil.rmtree(size_dir, ignore_errors=True)

    return result


# ---------------------------------------------------------------------------
# Bericht-Generator
# ---------------------------------------------------------------------------


def _format_seconds(seconds: float) -> str:
    if seconds < 0.01:
        return f"{seconds * 1000:.1f} ms"
    if seconds < 60:
        return f"{seconds:.2f} s"
    return f"{seconds / 60:.2f} min"


def _format_mb(mb: float) -> str:
    if mb < 1.0:
        return f"{mb * 1024:.0f} KB"
    return f"{mb:.1f} MB"


def detect_violations(results: list[SizeResult]) -> list[tuple[int, str, float, float]]:
    """Findet Phasen, die ihr Soft-Target überschreiten.

    Soft-Targets sind für 1M Zeilen definiert. Bei kleineren Größen
    skalieren wir linear (z. B. 30s/M = 3s/100k) – simple Heuristik,
    aber gibt einen ersten Hinweis auf Skalierungsprobleme.

    Sprint 12.1 / P-007: Import + DB-Speicherung werden zu einem
    Pipeline-Total aggregiert (siehe `PIPELINE_TOTAL_LABEL`), weil
    Sprint-11.3-Streaming die Cell-Coercion zwischen den Phasen
    verschoben hat. Einzelphasen-Verfehlungen für Import/DB werden
    NICHT mehr gemeldet – nur der Total.
    """
    violations: list[tuple[int, str, float, float]] = []
    pipeline_skip = set(PIPELINE_TOTAL_PHASES)
    for r in results:
        scale = r.size / 1_000_000
        by_label = {m.label: m.elapsed_s for m in r.measurements}
        for m in r.measurements:
            if m.label in pipeline_skip:
                continue  # Einzelphase: nicht eigenständig bewertet (P-007)
            target = SOFT_TARGETS_1M_SECONDS.get(m.label)
            if target is None:
                continue
            scaled_target = target * max(scale, 0.1)  # Mindest-Toleranz für Kleinst-Datasets
            if m.elapsed_s > scaled_target:
                violations.append((r.size, m.label, m.elapsed_s, scaled_target))

        # Pipeline-Total: nur prüfen, wenn beide Einzelphasen gemessen wurden.
        pipeline_target = SOFT_TARGETS_1M_SECONDS.get(PIPELINE_TOTAL_LABEL)
        if pipeline_target is not None and all(p in by_label for p in PIPELINE_TOTAL_PHASES):
            pipeline_elapsed = sum(by_label[p] for p in PIPELINE_TOTAL_PHASES)
            scaled_pipeline = pipeline_target * max(scale, 0.1)
            if pipeline_elapsed > scaled_pipeline:
                violations.append((r.size, PIPELINE_TOTAL_LABEL, pipeline_elapsed, scaled_pipeline))
    return violations


def _git_short_rev() -> str:
    try:
        rev = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            check=True,
            timeout=5,
        )
        return rev.stdout.strip()
    except (subprocess.SubprocessError, FileNotFoundError):
        return "unknown"


def write_report(path: Path, results: list[SizeResult]) -> None:
    """Schreibt den Markdown-Bericht nach `path`."""
    lines: list[str] = []
    lines.append("# Performance-Probe")
    lines.append("")
    lines.append(f"Datum: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(
        f"Maschine: {platform.system()} {platform.release()} ({platform.machine()}), "
        f"Python {sys.version.split()[0]}"
    )
    lines.append(f"Toolversion: {_git_short_rev()}")
    lines.append(f"psutil RSS-Cross-Check: {'an' if _HAS_PSUTIL else 'aus'}")
    lines.append("")

    lines.append("## Soft-Targets (1M Zeilen)")
    lines.append("")
    lines.append("| Phase | Target |")
    lines.append("|-------|-------:|")
    for label, target in SOFT_TARGETS_1M_SECONDS.items():
        lines.append(f"| {label} | < {target:.0f} s |")
    lines.append("")
    lines.append(
        "Bei kleineren Größen werden Targets linear skaliert "
        "(z. B. 30 s/M → 3 s/100k); reine Heuristik."
    )
    lines.append("")
    lines.append(
        "**Sprint 12.1 / P-007 – Phasen-Verlagerung:** seit Sprint 11.3 (Streaming-"
        "Import) gehört die Cell-Coercion + JSON-Encode-Arbeit zum DB-Insert-"
        "Generator, nicht mehr zur Import-Phase. Die historischen Einzeltargets "
        "(`Import < 60 s`, `DB-Speicherung < 30 s`) wurden deshalb zu einem "
        "Pipeline-Total `< 90 s` konsolidiert. Die Einzelphasen-Zeiten bleiben "
        "in den Mess-Tabellen sichtbar, werden aber NICHT mehr in der "
        "Verfehlungsübersicht bewertet."
    )
    lines.append("")
    lines.append("Historische Pre-Streaming-Targets (nur Sprint-10.x-Vergleichbarkeit):")
    lines.append("")
    lines.append("| Phase | Legacy-Target |")
    lines.append("|-------|--------------:|")
    for label, target in LEGACY_PRE_STREAMING_TARGETS_1M_SECONDS.items():
        lines.append(f"| {label} | < {target:.0f} s |")
    lines.append("")

    for r in results:
        lines.append(f"## Messung {r.size:,} Zeilen")
        lines.append("")
        if r.errors:
            for err in r.errors:
                lines.append(f"- ⚠ {err}")
            lines.append("")
        lines.append("| Phase | Zeit | Peak (tracemalloc) | RSS-Delta | Anmerkung |")
        lines.append("|-------|-----:|-------------------:|----------:|-----------|")
        for m in r.measurements:
            rss = _format_mb(m.rss_delta_mb) if m.rss_delta_mb is not None else "—"
            lines.append(
                f"| {m.label} | {_format_seconds(m.elapsed_s)} | "
                f"{_format_mb(m.peak_tracemalloc_mb)} | {rss} | {m.note} |"
            )
        lines.append("")

    violations = detect_violations(results)
    if violations:
        lines.append("## Soft-Target-Verfehlungen (Sprint-10.2-Kandidaten)")
        lines.append("")
        lines.append("| Größe | Phase | Gemessen | Skaliertes Target | Überschreitung |")
        lines.append("|------:|-------|---------:|------------------:|---------------:|")
        for size, label, measured_s, target_s in violations:
            over = measured_s - target_s
            lines.append(
                f"| {size:,} | {label} | {_format_seconds(measured_s)} | "
                f"{_format_seconds(target_s)} | +{_format_seconds(over)} |"
            )
        lines.append("")
    else:
        lines.append("## Soft-Target-Verfehlungen")
        lines.append("")
        lines.append("Keine – alle gemessenen Phasen liegen im Soft-Target.")
        lines.append("")

    lines.append("## Auffälligkeiten")
    lines.append("")
    lines.append(
        "Werden manuell ergänzt, nachdem die Tabellen oben gelesen wurden. "
        "Erwartete Bottleneck-Hypothesen (siehe Sprint-10.1-Brief):"
    )
    lines.append("")
    lines.append("- DatasetRepo.create – `executemany`-Bulk-Insert, sollte skalieren")
    lines.append(
        "- values_json-Encoding pro Row – ein json.dumps-Aufruf je Zeile, "
        "potenziell sichtbar bei 1M+"
    )
    lines.append("- DataTableView.highlight_rows – Set-Lookup im BackgroundRole")
    lines.append("- AuditTrail-PDF – reportlab.platypus mit vielen Flowables")
    lines.append("- Stratified mit vielen Strata – largest-remainder-Schleifen")
    lines.append("")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Performance-Discovery-Probe für große Datasets.",
    )
    parser.add_argument(
        "--sizes",
        nargs="+",
        type=int,
        default=[10_000, 100_000, 1_000_000],
        help="Dataset-Größen (Zeilen), die getestet werden sollen.",
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Nur SimpleSampler statt Simple+Cluster+Stratified.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("PERFORMANCE.md"),
        help="Pfad für den Markdown-Bericht (Default: PERFORMANCE.md im Repo-Root).",
    )
    parser.add_argument(
        "--audit-events",
        type=int,
        default=5000,
        help="Anzahl synthetischer Audit-Events für die PDF-Phase.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=None,
        help="Arbeitsverzeichnis für Zwischen-Dateien (Default: tmp/perf).",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)

    # Single QApplication für alle Größen – wir brauchen sie nur, damit
    # QTableView-Widgets konstruiert werden können.
    app = QApplication.instance() or QApplication(sys.argv[:1])
    assert app is not None

    work_dir = args.work_dir or (_REPO_ROOT / "tmp" / "perf")
    work_dir.mkdir(parents=True, exist_ok=True)

    results: list[SizeResult] = []
    try:
        for size in args.sizes:
            print(f"[probe] Lauf für {size:,} Zeilen ...")
            try:
                result = run_probe_for_size(
                    size,
                    work_dir,
                    quick=args.quick,
                    audit_events=args.audit_events,
                )
            except Exception as exc:  # Discovery: alles auffangen, im Bericht dokumentieren
                print(f"[probe] FEHLER bei {size:,}: {exc!r}")
                result = SizeResult(size=size, errors=[f"{type(exc).__name__}: {exc}"])
            results.append(result)
            for m in result.measurements:
                print(
                    f"  {m.label:<28s} {_format_seconds(m.elapsed_s):>10s}  "
                    f"peak={_format_mb(m.peak_tracemalloc_mb):>10s}  {m.note}"
                )
    finally:
        # tmp/perf darf zwischen Läufen vorgehalten werden – beim
        # nächsten Start werden Unterordner ohnehin neu angelegt.
        if work_dir.exists() and not any(work_dir.iterdir()):
            with contextlib.suppress(OSError):
                work_dir.rmdir()

    write_report(args.output, results)
    print(f"[probe] Bericht geschrieben: {args.output}")
    return 0


if __name__ == "__main__":
    # Damit `tempfile.tempdir` nicht zufällig auf macOS-/private-Pfade
    # verweist, die in CI gemounted sind. Defensiv, kein harter Need.
    tempfile.gettempdir()
    raise SystemExit(main())
