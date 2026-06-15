"""Import-Performance-Benchmark (Sprint 26) – Profiling-first.

Misst den Import-Pfad (.xlsx via python-calamine, .csv via stdlib-csv)
aufgeschlüsselt in drei Phasen, damit der dominante Kostenposten gezielt
optimiert werden kann (statt blind „überall ein bisschen"):

1. **Parse/Read**   – Datei → rohe Zeilen/Zellen
   (xlsx: ``CalamineSheet.iter_rows``; csv: ``_read_csv_text`` + ``_parse_csv``).
2. **Encode/Transform** – Coercion (``_coerce_value``) + Tagged-JSON-Encoding
   der Werte (``_values_to_json``, inkl. datetime/date/time-Tagging).
3. **Write**        – SQLite-``executemany`` der vorcodierten JSON-Strings
   + Commit (Index ``idx_dataset_rows_dataset`` ist via Migration aktiv).

Pro Phase wird der Median aus ``--runs`` Läufen (Default 3) gemessen; dazu
ein End-to-End-Lauf über die echten Produktions-Eintrittspunkte
(``ExcelImporter.import_file`` + ``DatasetRepo.create``) als ehrlicher
Total-Wert und ein optionaler ``cProfile``-Lauf zur Gegenprobe der Hotspots.

Das Skript ruft bewusst die **echten** Importer-/Persistenz-Funktionen auf –
es misst Produktionscode, keine Nachbildung. Die ~1-Mio-Zeilen-Fixtures
werden seeded zur Laufzeit erzeugt (gemischte Spalten: Text inkl. Nicht-ASCII,
Ganzzahl, Dezimal, datetime/date/time, Nullwerte) und nach dem Lauf gelöscht –
nichts davon landet im Repo.

Aufruf:
    python scripts/bench_import.py                       # 1M Zeilen, beide Formate, 3 Läufe
    python scripts/bench_import.py --rows 1000000 --runs 5
    python scripts/bench_import.py --format csv --profile
    python scripts/bench_import.py --quick               # 1000 Zeilen, 1 Lauf (Smoke)

# ---------------------------------------------------------------------------
# MESSERGEBNIS / ENTSCHEIDUNGS-GATE (Sprint 26, Apple Silicon, Python 3.13.13,
# 1M Zeilen, Median aus 3 Läufen, Git-Rev e31ef72 = Baseline)
# ---------------------------------------------------------------------------
# Dominanter Posten je Format: **Encode/Transform** (xlsx 5512 ms ≈ 43 %,
# csv 6957 ms ≈ 64 % der Phasen-Summe). Parse und Write sind klar günstiger.
#
# Encode zerfällt in zwei Teile:
#   - Coercion (_coerce_value/_coerce_string): xlsx 3261 ms, csv 5756 ms –
#     der GRÖSSERE Teil, aber er DEFINIERT die importierten Werte/Typen
#     (float→int-Normalisierung, deutsche Komma-Dezimalzahl, date→datetime).
#     Das ist die „byte-für-byte identisch"-Rote-Linie (Hard Constraint §9) –
#     bewusst NICHT angefasst.
#   - Tagged-JSON-Encoding (_values_to_json): xlsx 2118 ms, csv 1226 ms –
#     der vom Sprint-Prompt freigegebene, risikoarme Hebel.
#
# Phase 2 adressiert daher ausschließlich den Tagged-JSON-Encoder:
# orjson-Fast-Path via OPT_PASSTHROUGH_DATETIME (siehe persistence/_json.py).
# Vorher/Nachher-Zahlen: PERFORMANCE.md, Abschnitt „Sprint 26".
# ---------------------------------------------------------------------------
"""

from __future__ import annotations

import argparse
import cProfile
import csv
import io
import pstats
import shutil
import statistics
import subprocess
import sys
import tempfile
import time
from collections.abc import Callable, Iterator
from dataclasses import replace
from datetime import UTC, date, datetime, timedelta
from datetime import time as dt_time
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook

# Repository-Root in sys.path aufnehmen, damit das Script auch direkt
# (ohne `python -m`) aus dem Repo-Root läuft – wie scripts/perf_probe.py.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT / "src") not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT / "src"))

from python_calamine import CalamineWorkbook  # noqa: E402

from sampling_tool.core.models import Dataset, Engagement  # noqa: E402
from sampling_tool.io.importer import (  # noqa: E402
    ExcelImporter,
    _coerce_value,
    _detect_header,
    _excel_header_pass,
    _is_blank,
    _read_csv_text,
    _select_sheet,
)
from sampling_tool.io.importer import _parse_csv as _parse_csv_rows  # noqa: E402
from sampling_tool.persistence._json import _json_dumps, _values_to_json  # noqa: E402
from sampling_tool.persistence.database import Database  # noqa: E402
from sampling_tool.persistence.dataset_repo import DatasetRepo  # noqa: E402
from sampling_tool.persistence.engagement_repo import EngagementRepo  # noqa: E402

# Gemischte Spalten – belasten Coercion und Tagged-Encoder gleichermaßen.
COLUMN_NAMES: tuple[str, ...] = (
    "buchung_id",  # Ganzzahl
    "betrag",  # Dezimal
    "bezeichnung",  # Text inkl. Nicht-ASCII
    "erfasst_am",  # datetime
    "buchungsdatum",  # date  → wird zu datetime normalisiert (calamine)
    "uhrzeit",  # time
    "notiz",  # nullable Text
    "menge",  # nullable Ganzzahl
)

_BASE_DT = datetime(2020, 1, 1, 8, 0, 0)


def _git_short_rev() -> str:
    try:
        out = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=_REPO_ROOT,
            capture_output=True,
            text=True,
            check=True,
        )
        return out.stdout.strip()
    except (subprocess.SubprocessError, OSError):
        return "unbekannt"


# ---------------------------------------------------------------------------
# Fixture-Generierung (seeded, deterministisch)
# ---------------------------------------------------------------------------


def _generate_rows(n_rows: int, seed: int) -> Iterator[list[Any]]:
    """Yieldet ``n_rows`` Datenzeilen mit gemischten, seeded Werten.

    Jede 10. Zeile hat Nullwerte in den nullable-Spalten (``notiz``/``menge``);
    Nicht-ASCII-Text in ``bezeichnung`` belastet den UTF-8-Pfad.
    """
    rng = np.random.default_rng(seed)
    betraege = np.round(rng.uniform(-50_000.0, 50_000.0, size=n_rows), 2)
    mengen = rng.integers(1, 1000, size=n_rows)
    sekunden = rng.integers(0, 5 * 365 * 24 * 3600, size=n_rows)
    staedte = ("Zürich", "München", "Köln", "Wien", "Düsseldorf", "Genève")
    for i in range(n_rows):
        ts = _BASE_DT + timedelta(seconds=int(sekunden[i]))
        is_null = i % 10 == 0
        stadt = staedte[i % len(staedte)]
        yield [
            i + 1,
            float(betraege[i]),
            f"Straße {i + 1} – {stadt} €{i % 100}",
            ts,
            ts.date(),
            dt_time(ts.hour, ts.minute, ts.second),
            None if is_null else f"Beleg №{i + 1} – geprüft",
            None if is_null else int(mengen[i]),
        ]


def _write_xlsx(path: Path, n_rows: int, seed: int) -> None:
    """Schreibt die Fixture als .xlsx (write_only-Streaming, geringer RAM)."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Daten")
    ws.append(list(COLUMN_NAMES))
    for row in _generate_rows(n_rows, seed):
        ws.append(row)
    wb.save(str(path))


def _csv_cell(value: Any) -> str:
    """Stringifiziert einen Zellwert für die CSV-Fixture (None → leer)."""
    if value is None:
        return ""
    if isinstance(value, datetime | date | dt_time):
        return value.isoformat()
    return str(value)


def _write_csv(path: Path, n_rows: int, seed: int) -> None:
    """Schreibt die Fixture als .csv (utf-8, Komma-getrennt)."""
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(COLUMN_NAMES)
        for row in _generate_rows(n_rows, seed):
            writer.writerow([_csv_cell(c) for c in row])


# ---------------------------------------------------------------------------
# Phasen-Bausteine – rufen die echten Produktionsfunktionen auf
# ---------------------------------------------------------------------------


def _parse_xlsx(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Phase Parse (xlsx): Datei → Spalten + rohe Datenzeilen (calamine)."""
    wb = CalamineWorkbook.from_path(str(path))
    sheet = _select_sheet(wb, None)
    columns, _skipped, _warnings, _est = _excel_header_pass(sheet)
    rows_iter = iter(sheet.iter_rows())
    _detect_header(rows_iter)  # Header-Zeile konsumieren (wie der Generator)
    raw_rows = [list(raw) for raw in rows_iter if not _is_blank(raw)]
    return columns, raw_rows


def _parse_csv_phase(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Phase Parse (csv): read_bytes + decode + Sniff + csv.reader-Materialize."""
    text, _enc = _read_csv_text(path)
    columns, data_rows, _skipped, _warnings = _parse_csv_rows(text)
    return columns, data_rows


def _coerce_only(columns: list[str], raw_rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Encode-Teil 1: nur Coercion (rohe Zellen → native Werte-Dicts)."""
    return [
        {col: _coerce_value(raw[i] if i < len(raw) else None) for i, col in enumerate(columns)}
        for raw in raw_rows
    ]


def _json_only(value_dicts: list[dict[str, Any]]) -> list[str]:
    """Encode-Teil 2: nur Tagged-JSON-Encoding der vorcoercten Werte-Dicts."""
    return [_values_to_json(v) for v in value_dicts]


def _encode(columns: list[str], raw_rows: list[list[Any]]) -> list[str]:
    """Phase Encode: Coercion + Tagged-JSON pro Zeile (wie der Insert-Generator)."""
    return _json_only(_coerce_only(columns, raw_rows))


def _write(columns: list[str], json_strings: list[str], db_path: Path) -> None:
    """Phase Write: executemany der vorcodierten JSON-Strings + Commit.

    Spiegelt den INSERT aus ``DatasetRepo.create`` exakt (gleiche Tabelle,
    gleiche Spalten, eine Transaktion). Index ``idx_dataset_rows_dataset``
    ist via Migration aktiv – wie in Produktion.
    """
    db = Database(db_path)
    db.migrate()
    eng = EngagementRepo(db.connect()).get_or_create(
        Engagement(
            auditor_name="Bench",
            auditor_position="Senior",
            client_name="ACME",
            audit_type="ISAE 3402",
        )
    )
    with db.session() as conn:
        cur = conn.execute(
            "INSERT INTO datasets "
            "(engagement_id, name, source_file, imported_at, row_count, columns_json) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                eng.id,
                "bench",
                "bench",
                datetime.now(UTC),
                len(json_strings),
                _json_dumps(list(columns)),
            ),
        )
        dataset_id = cur.lastrowid
        conn.executemany(
            "INSERT INTO dataset_rows (dataset_id, row_index, values_json) VALUES (?, ?, ?)",
            ((dataset_id, i, js) for i, js in enumerate(json_strings, start=1)),
        )
    db.close()


def _end_to_end(path: Path, db_path: Path) -> None:
    """End-to-End über die echten Eintrittspunkte (Importer → DatasetRepo)."""
    db = Database(db_path)
    db.migrate()
    eng = EngagementRepo(db.connect()).get_or_create(
        Engagement(
            auditor_name="Bench",
            auditor_position="Senior",
            client_name="ACME",
            audit_type="ISAE 3402",
        )
    )
    result = ExcelImporter().import_file(path)
    dataset: Dataset = replace(result.dataset, engagement_id=eng.id)
    with db.session() as conn:
        DatasetRepo(conn).create(dataset, result.rows)
    db.close()


# ---------------------------------------------------------------------------
# Mess-Mechanik
# ---------------------------------------------------------------------------


def _median_ms(fn: Callable[[], Any], runs: int) -> float:
    """Median der Laufzeit (ms) über ``runs`` Aufrufe von ``fn``."""
    samples: list[float] = []
    for _ in range(runs):
        t0 = time.perf_counter()
        fn()
        samples.append((time.perf_counter() - t0) * 1000.0)
    return statistics.median(samples)


def _bench_format(
    fmt: str,
    path: Path,
    parse_fn: Callable[[Path], tuple[list[str], list[list[Any]]]],
    runs: int,
    work_dir: Path,
) -> dict[str, float]:
    """Misst Parse/Encode/Write (Median) + End-to-End für ein Format."""
    # Parse: jeder Lauf liest die Datei frisch.
    parse_ms = _median_ms(lambda: parse_fn(path), runs)

    # Für Encode/Write brauchen wir die geparsten Rohzeilen einmal materialisiert.
    columns, raw_rows = parse_fn(path)
    encode_ms = _median_ms(lambda: _encode(columns, raw_rows), runs)

    # Encode-Sub-Split: Coercion vs. Tagged-JSON-Encoding (der optimierbare Teil).
    coerce_ms = _median_ms(lambda: _coerce_only(columns, raw_rows), runs)
    value_dicts = _coerce_only(columns, raw_rows)
    json_ms = _median_ms(lambda: _json_only(value_dicts), runs)

    json_strings = _json_only(value_dicts)
    write_counter = {"n": 0}

    def _one_write() -> None:
        write_counter["n"] += 1
        db_path = work_dir / f"write_{fmt}_{write_counter['n']}.db"
        _write(columns, json_strings, db_path)
        db_path.unlink(missing_ok=True)
        Path(f"{db_path}-wal").unlink(missing_ok=True)
        Path(f"{db_path}-shm").unlink(missing_ok=True)

    write_ms = _median_ms(_one_write, runs)

    e2e_counter = {"n": 0}

    def _one_e2e() -> None:
        e2e_counter["n"] += 1
        db_path = work_dir / f"e2e_{fmt}_{e2e_counter['n']}.db"
        _end_to_end(path, db_path)
        db_path.unlink(missing_ok=True)
        Path(f"{db_path}-wal").unlink(missing_ok=True)
        Path(f"{db_path}-shm").unlink(missing_ok=True)

    e2e_ms = _median_ms(_one_e2e, runs)

    return {
        "rows": float(len(raw_rows)),
        "parse": parse_ms,
        "encode": encode_ms,
        "coerce": coerce_ms,
        "json": json_ms,
        "write": write_ms,
        "e2e": e2e_ms,
    }


def _print_format_report(fmt: str, n_rows: int, runs: int, result: dict[str, float]) -> None:
    parse, encode, write = result["parse"], result["encode"], result["write"]
    phases_total = parse + encode + write
    dominant_name, dominant_ms = max(
        (("Parse", parse), ("Encode", encode), ("Write", write)),
        key=lambda kv: kv[1],
    )
    dominant_pct = (dominant_ms / phases_total * 100.0) if phases_total else 0.0

    print(f"\n=== {fmt.upper()} ({n_rows:,} Zeilen, Median aus {runs} Läufen) ===")
    print(f"  Parse  : {parse:10.1f} ms")
    print(f"  Encode : {encode:10.1f} ms")
    print(f"      ├─ Coercion    : {result['coerce']:10.1f} ms  (semantik-definierend)")
    print(f"      └─ Tagged-JSON : {result['json']:10.1f} ms  (optimierbarer Hebel)")
    print(f"  Write  : {write:10.1f} ms")
    print(f"  {'-' * 22}")
    print(f"  Σ Phasen     : {phases_total:10.1f} ms")
    print(f"  End-to-End   : {result['e2e']:10.1f} ms (Streaming-Pipeline)")
    print(f"  → Dominant   : {dominant_name} ({dominant_pct:.0f} % der Phasen-Summe)")


# ---------------------------------------------------------------------------
# cProfile-Gegenprobe
# ---------------------------------------------------------------------------


def _profile_format(fmt: str, path: Path, work_dir: Path) -> None:
    """Profiliert einen End-to-End-Import und druckt die Top-Hotspots."""
    db_path = work_dir / f"profile_{fmt}.db"
    profiler = cProfile.Profile()
    profiler.enable()
    _end_to_end(path, db_path)
    profiler.disable()
    db_path.unlink(missing_ok=True)
    Path(f"{db_path}-wal").unlink(missing_ok=True)
    Path(f"{db_path}-shm").unlink(missing_ok=True)

    print(f"\n=== cProfile {fmt.upper()} (Top 20 nach kumulativer Zeit) ===")
    stream = io.StringIO()
    stats = pstats.Stats(profiler, stream=stream).sort_stats("cumulative")
    stats.print_stats(20)
    print(stream.getvalue())


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import-Performance-Benchmark (Sprint 26).")
    parser.add_argument("--rows", type=int, default=1_000_000, help="Zeilenzahl der Fixture.")
    parser.add_argument("--runs", type=int, default=3, help="Mess-Läufe pro Phase (Median).")
    parser.add_argument(
        "--format",
        choices=("both", "xlsx", "csv"),
        default="both",
        help="Welches Format gemessen wird.",
    )
    parser.add_argument(
        "--profile",
        action="store_true",
        help="Zusätzlich einen cProfile-Lauf je Format zur Hotspot-Gegenprobe.",
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Schneller Smoke: überschreibt rows=1000, runs=1.",
    )
    parser.add_argument("--seed", type=int, default=20_260_612, help="Seed für die Fixture.")
    parser.add_argument(
        "--keep",
        action="store_true",
        help="Temp-Dateien nach dem Lauf nicht löschen (Debug).",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    rows = 1000 if args.quick else args.rows
    runs = 1 if args.quick else args.runs
    formats = ("xlsx", "csv") if args.format == "both" else (args.format,)

    print("# Import-Performance-Benchmark (Sprint 26)")
    print(f"Git-Rev: {_git_short_rev()} | Zeilen: {rows:,} | Läufe: {runs} | Seed: {args.seed}")

    work_dir = Path(tempfile.mkdtemp(prefix="bench_import_"))
    try:
        if "xlsx" in formats:
            xlsx_path = work_dir / "bench.xlsx"
            print(f"\nGeneriere {rows:,}-Zeilen-xlsx … (einmalig, nicht gemessen)")
            _write_xlsx(xlsx_path, rows, args.seed)
            result = _bench_format("xlsx", xlsx_path, _parse_xlsx, runs, work_dir)
            _print_format_report("xlsx", rows, runs, result)
            if args.profile:
                _profile_format("xlsx", xlsx_path, work_dir)

        if "csv" in formats:
            csv_path = work_dir / "bench.csv"
            print(f"\nGeneriere {rows:,}-Zeilen-csv … (einmalig, nicht gemessen)")
            _write_csv(csv_path, rows, args.seed)
            result = _bench_format("csv", csv_path, _parse_csv_phase, runs, work_dir)
            _print_format_report("csv", rows, runs, result)
            if args.profile:
                _profile_format("csv", csv_path, work_dir)
    finally:
        if not args.keep:
            shutil.rmtree(work_dir, ignore_errors=True)
        else:
            print(f"\nTemp-Dateien behalten unter: {work_dir}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
