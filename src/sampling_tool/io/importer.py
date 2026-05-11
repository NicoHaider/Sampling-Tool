"""Excel-/CSV-Importer mit Streaming-Read und Header-Detection.

Die Klasse `ExcelImporter` ist der einzige Eintrittspunkt für den
Import-Pfad. Sie produziert ein `Dataset` (frozen Dataclass aus
`core.models`) – der Aufrufer setzt anschließend `engagement_id` und
übergibt das Dataset an `DatasetRepo.create()`.

Architektur-Anker (siehe Sprint-3-Brief):
- **Streaming**: `openpyxl.load_workbook(read_only=True)` für große Listen.
- **Header-Detection**: erste „dichte" Zeile (überwiegend Strings) gilt als
  Header, Inhalts-Zeilen folgen. Fallback: erste Zeile.
- **Encoding-Detection** für CSV: utf-8 → utf-8-sig → latin-1 → cp1252.
- **Native Python-Typen** im Output – kein numpy/pandas-Typ verlässt diese
  Datei.
- **Progress-Callback**: `progress(current, total)` wird in regelmäßigen
  Abständen während des Reads aufgerufen.
"""

from __future__ import annotations

import csv
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from sampling_tool.config import SUPPORTED_CSV_SUFFIXES, SUPPORTED_EXCEL_SUFFIXES
from sampling_tool.core.models import Dataset, DatasetRow

ProgressCallback = Callable[[int, int], None]

# Alle Encodings, die wir bei CSV der Reihe nach probieren.
_CSV_ENCODINGS: Final[tuple[str, ...]] = ("utf-8", "utf-8-sig", "latin-1", "cp1252")

# Schwellwert für Header-Detection: ≥ Anteil String-Zellen einer Zeile.
_HEADER_STRING_RATIO: Final[float] = 0.5


# ---------------------------------------------------------------------------
# Result-Container
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class ImportResult:
    """Rückgabe-Wert von `ExcelImporter.import_file`."""

    dataset: Dataset
    skipped_rows: int
    warnings: tuple[str, ...]


# ---------------------------------------------------------------------------
# Fehler
# ---------------------------------------------------------------------------


# `ImportError` ist Builtin und darf nicht verschattet werden – daher das
# Domain-Präfix.
class DataImportError(ValueError):
    """Fachlicher Importfehler (deutsche Endnutzer-Message)."""


# ---------------------------------------------------------------------------
# Importer
# ---------------------------------------------------------------------------


class ExcelImporter:
    """Liest .xlsx/.xlsm/.csv und gibt ein `Dataset` zurück.

    Stateless im fachlichen Sinn – die Instanz hält nur den Progress-Callback.
    Die gleiche Instanz darf für mehrere Imports wiederverwendet werden.
    """

    def __init__(self, progress: ProgressCallback | None = None) -> None:
        self.progress = progress

    # ---- Public API -----------------------------------------------------

    def import_file(self, path: Path, sheet_name: str | None = None) -> ImportResult:
        """Importiert die angegebene Datei und liefert ein `ImportResult`."""
        if not path.exists():
            raise DataImportError(f"Datei nicht gefunden: {path}")

        suffix = path.suffix.lower()
        if suffix in SUPPORTED_CSV_SUFFIXES:
            return self._import_csv(path)
        if suffix in SUPPORTED_EXCEL_SUFFIXES:
            return self._import_excel(path, sheet_name)
        raise DataImportError(
            f"Dateityp '{suffix}' wird nicht unterstützt. "
            f"Erlaubt: {', '.join(SUPPORTED_EXCEL_SUFFIXES + SUPPORTED_CSV_SUFFIXES)}"
        )

    def detect_sheets(self, path: Path) -> list[str]:
        """Listet die Sheet-Namen einer Excel-Datei (für UI-Multi-Sheet-Auswahl)."""
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_EXCEL_SUFFIXES:
            raise DataImportError(
                f"Sheet-Liste nur für Excel-Dateien verfügbar (Datei: {path.name})."
            )
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def preview(
        self,
        path: Path,
        sheet_name: str | None = None,
        n_rows: int = 10,
    ) -> tuple[list[str], list[dict[str, Any]]]:
        """Liefert (Spalten, erste n Zeilen) für eine UI-Vorschau – kein vollständiger Import."""
        if n_rows < 0:
            raise DataImportError("preview(): n_rows muss >= 0 sein.")

        suffix = path.suffix.lower()
        if suffix in SUPPORTED_CSV_SUFFIXES:
            text, _enc = _read_csv_text(path)
            columns, rows_iter, _skipped, _warns = _parse_csv(text)
            preview_rows = [dict(zip(columns, r, strict=False)) for r in rows_iter[:n_rows]]
            return list(columns), preview_rows

        if suffix in SUPPORTED_EXCEL_SUFFIXES:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = _select_sheet(wb, sheet_name)
                columns, data_rows, _skipped, _warns = _parse_excel_sheet(ws, limit=n_rows)
            finally:
                wb.close()
            preview_rows = [dict(zip(columns, r, strict=False)) for r in data_rows]
            return list(columns), preview_rows

        raise DataImportError(f"Vorschau für Dateityp '{suffix}' nicht unterstützt.")

    # ---- Excel ----------------------------------------------------------

    def _import_excel(self, path: Path, sheet_name: str | None) -> ImportResult:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = _select_sheet(wb, sheet_name)
            columns, data_rows, skipped, warnings = _parse_excel_sheet(ws, limit=None)
        finally:
            wb.close()

        if not columns:
            raise DataImportError(
                f"Keine Spaltenüberschriften gefunden in '{path.name}' "
                f"(Sheet: '{sheet_name or 'Standard'}')."
            )

        rows = self._materialize_rows(columns, data_rows)
        dataset = Dataset(
            name=path.stem,
            columns=tuple(columns),
            rows=rows,
            source_file=str(path),
        )
        return ImportResult(dataset=dataset, skipped_rows=skipped, warnings=tuple(warnings))

    # ---- CSV ------------------------------------------------------------

    def _import_csv(self, path: Path) -> ImportResult:
        text, encoding = _read_csv_text(path)
        columns, data_rows, skipped, warnings = _parse_csv(text)

        if not columns:
            raise DataImportError(f"CSV-Datei '{path.name}' enthält keine Daten.")

        if encoding != "utf-8":
            warnings = [*warnings, f"CSV-Encoding erkannt als '{encoding}'."]

        rows = self._materialize_rows(columns, data_rows)
        dataset = Dataset(
            name=path.stem,
            columns=tuple(columns),
            rows=rows,
            source_file=str(path),
        )
        return ImportResult(dataset=dataset, skipped_rows=skipped, warnings=tuple(warnings))

    # ---- Gemeinsam ------------------------------------------------------

    def _materialize_rows(
        self,
        columns: list[str],
        data_rows: list[list[Any]],
    ) -> tuple[DatasetRow, ...]:
        total = len(data_rows)
        rows: list[DatasetRow] = []
        for idx, raw in enumerate(data_rows, start=1):
            values = {
                col: _coerce_value(raw[i] if i < len(raw) else None)
                for i, col in enumerate(columns)
            }
            rows.append(DatasetRow(row_id=idx, values=values))
            self._tick(idx, total)
        # Falls das Dataset leer ist, mindestens einen 0/0-Tick senden, damit
        # UIs einen "fertig"-Status zeichnen können.
        if total == 0:
            self._tick(0, 0)
        return tuple(rows)

    def _tick(self, current: int, total: int) -> None:
        if self.progress is not None:
            self.progress(current, total)


# ---------------------------------------------------------------------------
# Hilfen – Excel
# ---------------------------------------------------------------------------


def _select_sheet(wb: Workbook, sheet_name: str | None) -> Any:
    """Liefert das gewünschte Sheet oder das aktive Default-Sheet."""
    if sheet_name is None:
        ws = wb.active
        if ws is None:
            raise DataImportError("Workbook ist leer (kein aktives Arbeitsblatt).")
        return ws
    if sheet_name not in wb.sheetnames:
        raise DataImportError(
            f"Sheet '{sheet_name}' existiert nicht. Verfügbar: {', '.join(wb.sheetnames)}."
        )
    return wb[sheet_name]


def _parse_excel_sheet(
    ws: Any, limit: int | None
) -> tuple[list[str], list[list[Any]], int, list[str]]:
    """Parst ein openpyxl-Worksheet und liefert (Spalten, Datenzeilen, skipped, warnings)."""
    rows_iter: Iterator[tuple[Any, ...]] = ws.iter_rows(values_only=True)
    header_row, leading_blanks = _detect_header(rows_iter)
    if header_row is None:
        return [], [], leading_blanks, []

    columns, header_warnings = _normalize_columns(header_row)

    data_rows: list[list[Any]] = []
    skipped = leading_blanks
    for raw in rows_iter:
        if _is_blank(raw):
            skipped += 1
            continue
        data_rows.append(list(raw))
        if limit is not None and len(data_rows) >= limit:
            break

    return columns, data_rows, skipped, header_warnings


def _detect_header(
    rows_iter: Iterator[tuple[Any, ...]],
) -> tuple[tuple[Any, ...] | None, int]:
    """Erste Zeile mit überwiegend Strings = Header. Leere davor zählen als skipped."""
    leading_blanks = 0
    for raw in rows_iter:
        if _is_blank(raw):
            leading_blanks += 1
            continue
        if _looks_like_header(raw):
            return raw, leading_blanks
        # Erste nicht-leere Zeile, aber nicht headerlike → trotzdem als Header
        # nehmen (Fallback). Ohne Header geht hier nichts weiter.
        return raw, leading_blanks
    return None, leading_blanks


def _looks_like_header(row: tuple[Any, ...]) -> bool:
    non_empty = [c for c in row if c is not None and str(c).strip() != ""]
    if not non_empty:
        return False
    string_like = sum(1 for c in non_empty if isinstance(c, str))
    return (string_like / len(non_empty)) >= _HEADER_STRING_RATIO


def _is_blank(row: tuple[Any, ...]) -> bool:
    return all(c is None or (isinstance(c, str) and c.strip() == "") for c in row)


def _normalize_columns(header_row: tuple[Any, ...]) -> tuple[list[str], list[str]]:
    """Stringifiziert + trimmt Spaltennamen, vergibt Suffixe bei Duplikaten."""
    raw_names: list[str] = []
    for idx, cell in enumerate(header_row, start=1):
        text = "" if cell is None else str(cell).strip()
        raw_names.append(text or f"Spalte_{idx}")

    seen: dict[str, int] = {}
    final: list[str] = []
    warnings: list[str] = []
    for name in raw_names:
        if name not in seen:
            seen[name] = 1
            final.append(name)
        else:
            seen[name] += 1
            new_name = f"{name}_{seen[name]}"
            warnings.append(f"Doppelter Spaltenname '{name}' → umbenannt zu '{new_name}'.")
            final.append(new_name)
    return final, warnings


# ---------------------------------------------------------------------------
# Hilfen – CSV
# ---------------------------------------------------------------------------


def _read_csv_text(path: Path) -> tuple[str, str]:
    """Probiert die Encoding-Liste durch und gibt (Text, gewähltes Encoding) zurück.

    UTF-8-BOM wird vorab erkannt, weil sonst der utf-8-Decode zwar erfolgreich
    durchläuft, aber das BOM-Zeichen `﻿` als unsichtbares Präfix in der
    ersten Spalte hängenbleibt.
    """
    raw = path.read_bytes()
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig"), "utf-8-sig"

    last_error: Exception | None = None
    for encoding in _CSV_ENCODINGS:
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError as e:
            last_error = e
    # Sehr unwahrscheinlich – latin-1 nimmt jedes Byte. Aber defensiv:
    raise DataImportError(
        f"CSV '{path.name}' konnte mit keinem unterstützten Encoding gelesen werden "
        f"({', '.join(_CSV_ENCODINGS)}). Letzter Fehler: {last_error}"
    )


def _parse_csv(text: str) -> tuple[list[str], list[list[Any]], int, list[str]]:
    """Splittet CSV-Text in Header + Datenzeilen. Delimiter wird geschnüffelt."""
    sample = text[:8192] or text
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # Default: Komma

    reader = csv.reader(text.splitlines(), dialect=dialect)
    all_rows = [row for row in reader]

    # Leere Zeilen am Anfang strippen, davon zählen wir die ersten als
    # "leading blanks" für die skipped-Bilanz.
    leading = 0
    while all_rows and _is_blank(tuple(all_rows[0])):
        all_rows.pop(0)
        leading += 1

    # Trailing-Blanks ebenfalls strippen (zählen aber nicht als skipped).
    while all_rows and _is_blank(tuple(all_rows[-1])):
        all_rows.pop()

    if not all_rows:
        return [], [], leading, []

    header_row = tuple(all_rows[0])
    columns, warnings = _normalize_columns(header_row)

    data_rows: list[list[Any]] = []
    skipped = leading
    for raw in all_rows[1:]:
        if _is_blank(tuple(raw)):
            skipped += 1
            continue
        data_rows.append(list(raw))

    return columns, data_rows, skipped, warnings


# ---------------------------------------------------------------------------
# Typ-Konvertierung
# ---------------------------------------------------------------------------


def _coerce_value(value: Any) -> Any:
    """Mappt openpyxl-/CSV-Zellwerte auf native Python-Typen.

    Reihenfolge: None → datetime/date/time durchreichen → bool als bool →
    int/float bleiben → numerische Strings konvertieren → sonst getrimmter
    String. Numpy/Pandas-Typen werden bewusst NICHT erzeugt – das Dataset
    soll JSON-roundtrippable bleiben.
    """
    if value is None:
        return None
    if isinstance(value, datetime | date | time):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, int | float):
        return value
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        as_int = _try_int(text)
        if as_int is not None:
            return as_int
        as_float = _try_float(text)
        if as_float is not None:
            return as_float
        return text
    # Letztes Mittel: stringifizieren, damit JSON-Persistierung funktioniert.
    return str(value)


def _try_int(text: str) -> int | None:
    try:
        # int("1.0") wirft – das ist Absicht, das wäre ein Float.
        return int(text)
    except ValueError:
        return None


def _try_float(text: str) -> float | None:
    # Deutsche Komma-Dezimalzahl tolerieren ("1,5" → 1.5), aber nur wenn
    # eindeutig (kein zusätzlicher Punkt im String).
    candidate = text.replace(",", ".") if "." not in text and text.count(",") == 1 else text
    try:
        return float(candidate)
    except ValueError:
        return None
