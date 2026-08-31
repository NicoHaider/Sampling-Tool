"""Excel-Exporter für gezogene Stichproben.

Erzeugt eine .xlsx mit zwei Sheets:

- **Sample**: gewählte Spalten + Header-Zeile, BDO-rote Header-Füllung,
  weißer Bold-Text, automatische Spaltenbreiten.
- **Metadaten**: Engagement-Info, Sampling-Methode, Seed, Population,
  Sample-Size, Datum.

Schreibt **atomar** über `sampling_tool.io._atomic.atomic_output` (S2.5 /
A-004): Output geht zuerst in eine exklusiv erzeugte Tempdatei im Zielordner,
danach `os.replace()` auf den Ziel-Pfad. Damit bleibt bei einem Crash mitten
im Schreiben kein halbes File zurück.
"""

from __future__ import annotations

from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sampling_tool import __version__
from sampling_tool.config import (
    BDO_RED,
    EXPORT_FILENAME_PATTERN,
    EXPORT_SUFFIX_SAMPLING,
    EXPORT_TYPE_SAMPLING,
    excel_argb,
    export_date_token,
    local_export_now,
    sanitize_export_filename_token,
)
from sampling_tool.core.models import Dataset, DatasetRow, Engagement, SampleResult
from sampling_tool.core.provenance import SamplingProvenance
from sampling_tool.io._atomic import AtomicReplaceError, atomic_output
from sampling_tool.io._xlsx_safe import safe_row
from sampling_tool.io._xlsx_util import (
    autosize_columns as autosize_columns,
)
from sampling_tool.io._xlsx_util import (
    display_string as display_string,
)

if TYPE_CHECKING:
    from sampling_tool.persistence.repositories import DatasetRepo

ProgressCallback = Callable[[int, int], None]

_SHEET_DATA: Final[str] = "Sample"
_SHEET_META: Final[str] = "Metadaten"
_MAX_COLUMN_WIDTH: Final[int] = 50


class ExportError(ValueError):
    """Fachlicher Exportfehler (deutsche Endnutzer-Message)."""


class ExcelExporter:
    """Schreibt ein `SampleResult` (+ Spalten-Auswahl) als .xlsx auf Platte."""

    def __init__(self, progress: ProgressCallback | None = None) -> None:
        self.progress = progress

    def export_sample(
        self,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        columns: list[str],
        output_dir: Path,
        custom_name: str,
        custom_id: str,
        engagement: Engagement | None = None,
        now: datetime | None = None,
    ) -> Path:
        """Exportiert die gezogenen Zeilen.

        Liefert den vollen Pfad zur erzeugten Datei zurück. Der Dateiname
        folgt dem VBA-Schema aus `config.EXPORT_FILENAME_PATTERN` (mit
        `type`-Token `config.EXPORT_TYPE_SAMPLING`).

        `now` ist der Zeitpunkt für den `{date}`-Token (Sprint 74 / §2.2).
        Der Export-Dialog reicht genau den Zeitpunkt durch, aus dem er seine
        Vorschau gebaut hat – damit heißt die geschriebene Datei per
        Konstruktion wie die Vorschau und nicht nur dann, wenn zwischen
        Dialog-Öffnen und OK-Klick kein Tageswechsel liegt. `None` bedeutet
        „selbst lesen" und ist für Aufrufer ohne Dialog gedacht.

        Sprint-11.4-Streaming: statt einer voll materialisierten Row-Liste
        nimmt der Exporter das `DatasetRepo` entgegen und lädt sich
        ausschließlich die `sample.selected_row_ids` über
        `get_rows_by_ids`. Bei 1M-Dataset und 1k-Sample werden so nur
        1k Rows aus der DB geholt statt 1M (was bisher der Common Case
        war, weil `handle_export_sample` `get_all_rows` aufgerufen hat).

        `engagement` ist optional – wird im Metadaten-Sheet ausgewertet,
        wenn gesetzt. Damit kann der Exporter auch standalone genutzt
        werden.
        """
        self._validate(columns, dataset)

        sample_rows = (
            dataset_repo.get_rows_by_ids(dataset.id, list(sample.selected_row_ids))
            if dataset.id is not None
            else []
        )

        filename = self._build_filename(custom_name, custom_id, now)
        target = output_dir / filename

        wb = Workbook()
        try:
            self._write_sample_sheet(wb, sample_rows, columns)
            self._write_metadata_sheet(wb, sample, dataset, engagement)
            try:
                with atomic_output(target) as tmp:
                    wb.save(tmp)
            except AtomicReplaceError as exc:
                raise ExportError(
                    f"Die Zieldatei „{target.name}“ konnte nicht geschrieben werden: "
                    f"{exc}. Sie ist möglicherweise in Excel geöffnet oder es fehlen "
                    "Schreibrechte."
                ) from exc
        finally:
            wb.close()

        return target

    # ---- Helpers --------------------------------------------------------

    @staticmethod
    def _build_filename(custom_name: str, custom_id: str, now: datetime | None = None) -> str:
        """Baut den Dateinamen aus dem gemeinsamen Pattern (`config.py`).

        `now=None` liest die Uhr selbst – für Aufrufer ohne Dialog. Der
        `or "sample"`-Fallback greift NACH dem Sanitizer und weicht damit vom
        `or "export"`-Fallback der Vorschau ab (`_export_base.py`). Diese
        Asymmetrie ist Bestandsverhalten und bleibt in Sprint 74 bewusst
        unangetastet: sie ist nur bei leerem/whitespace-only Namen sichtbar,
        und den blockiert der Dialog über `HINT_MISSING_NAME`, bevor der
        Writer ihn je zu sehen bekommt.
        """
        safe_name = sanitize_export_filename_token(custom_name) or "sample"
        safe_id = sanitize_export_filename_token(custom_id) or "0"
        body = EXPORT_FILENAME_PATTERN.format(
            name=safe_name,
            id=safe_id,
            type=EXPORT_TYPE_SAMPLING,
            date=export_date_token(now if now is not None else local_export_now()),
        )
        return f"{body}{EXPORT_SUFFIX_SAMPLING}"

    @staticmethod
    def _validate(columns: list[str], dataset: Dataset) -> None:
        if not columns:
            raise ExportError("Mindestens eine Exportspalte muss ausgewählt sein.")
        unknown = [c for c in columns if c not in dataset.columns]
        if unknown:
            raise ExportError(
                f"Folgende Spalten existieren nicht im Dataset: {', '.join(unknown)}. "
                f"Verfügbar: {', '.join(dataset.columns)}."
            )

    def _write_sample_sheet(
        self,
        wb: Workbook,
        rows: list[DatasetRow],
        columns: list[str],
    ) -> None:
        ws = wb.active
        assert ws is not None
        ws.title = _SHEET_DATA

        header_fill = PatternFill(
            start_color=excel_argb(BDO_RED),
            end_color=excel_argb(BDO_RED),
            fill_type="solid",
        )
        header_font = Font(bold=True, color="FFFFFFFF")
        header_align = Alignment(vertical="center", horizontal="left")

        safe_row(ws, columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Stabile Reihenfolge: nach row_id (matcht selected_row_ids-Sortierung).
        # Caller via get_rows_by_ids garantiert das schon, hier nur defensiv.
        rows_to_write = sorted(rows, key=lambda r: r.row_id)
        total = len(rows_to_write)
        max_widths: list[int] = [len(c) for c in columns]
        for idx, row in enumerate(rows_to_write, start=1):
            raw_values = [row.values.get(col) for col in columns]
            values = safe_row(ws, raw_values)
            for i, val in enumerate(values):
                length = len(display_string(val))
                if length > max_widths[i]:
                    max_widths[i] = min(length, _MAX_COLUMN_WIDTH)
            self._tick(idx, total)
        if total == 0:
            self._tick(0, 0)

        for i, width in enumerate(max_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(8, width + 2)

        ws.freeze_panes = "A2"

    def _write_metadata_sheet(
        self,
        wb: Workbook,
        sample: SampleResult,
        dataset: Dataset,
        engagement: Engagement | None,
    ) -> None:
        ws = wb.create_sheet(_SHEET_META)

        provenance = SamplingProvenance.from_sample_result(
            sample, dataset_id=dataset.id, app_version=__version__
        )

        rows: list[tuple[str, Any]] = [
            ("Erstellt am", datetime.now()),
            ("Dataset", dataset.name),
            ("Quelldatei", dataset.source_file),
        ]
        rows.extend(provenance.to_ordered_fields())
        if engagement is not None:
            rows.extend(
                [
                    ("Auditor", engagement.auditor_name),
                    ("Auditor-Position", engagement.auditor_position or "—"),
                    ("Mandant", engagement.client_name),
                    ("Prüfungstyp", engagement.audit_type or "—"),
                ]
            )

        safe_row(ws, ["Feld", "Wert"])
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        for label, value in rows:
            safe_row(ws, [label, value])

        autosize_columns(ws, columns=2, min_width=12)

    def _tick(self, current: int, total: int) -> None:
        if self.progress is not None:
            self.progress(current, total)


# ---------------------------------------------------------------------------
# Hilfen
# ---------------------------------------------------------------------------
