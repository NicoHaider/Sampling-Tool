"""Gemeinsame Spaltenbreiten-Helfer für die xlsx-Exporter.

`exporter.py` (Sample-Export) und `multi_report_exporter.py` (Multi-Sheet-
Report) hatten je eine fast identische `_autosize` + Display-String-
Implementierung (Q-003). Diese Datei ist die eine Quelle dafür; `min_width`
bleibt Parameter, damit jeder Aufrufer seinen bisherigen Wert behält – rein
kosmetische Konsolidierung, keine Spaltenbreiten-Änderung.
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Any, Final

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_MAX_COLUMN_WIDTH: Final[int] = 50


def display_string(value: Any) -> str:
    """Date-aware String-Repräsentation für die Spaltenbreiten-Berechnung."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def autosize_columns(ws: Worksheet, columns: int, *, min_width: int) -> None:
    """Setzt Spaltenbreiten anhand der längsten Zellinhalte (erste `columns` Spalten)."""
    widths = [0] * columns
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row[:columns]):
            length = len(display_string(val))
            if length > widths[i]:
                widths[i] = min(length, _MAX_COLUMN_WIDTH)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(min_width, width + 2)
