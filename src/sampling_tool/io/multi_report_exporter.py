"""Multi-Sheet Excel-Report: Komplett-Bericht eines Engagements.

`MultiSheetReportExporter` schreibt eine .xlsx mit vier Sheets:

1. **Übersicht** – Engagement-Metadaten + Anzahlen (Datasets / Samples /
   Audit-Events / letzte Aktivität).
2. **AuditTrail** – alle Events chronologisch (älteste oben), inkl. einer
   `Details`-Spalte (`AuditEvent.details` als kompakte Ein-Zeilen-Darstellung).
3. **Samples** – jede Stichprobe mit voller Reproduktions-Provenienz
   (`SamplingProvenance`): Methode, angeforderte/tatsächliche Größe, Seed,
   Filter (Feld/Operator/Wert), Cluster-/Stratum-Feld, Parent-Sample,
   Algorithmus-Version, Dataset-ID, Datum.
4. **Statistiken** – Methoden-Verteilung als Tabelle + eingebettetes
   Bar-Chart-Bild via `io.charts`.

Schreibt **atomar** über `sampling_tool.io._atomic.atomic_output` (S2.5 /
A-004): Output geht zuerst in eine exklusiv erzeugte Tempdatei im Zielordner,
danach `os.replace()` auf den Ziel-Pfad. Damit bleibt bei einem Crash mitten
im Schreiben kein halbes File zurück – das gleiche Muster wie in `exporter.py`.
"""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from sampling_tool import __version__
from sampling_tool.config import BDO_RED, excel_argb
from sampling_tool.core.formatting import format_audit_details, format_optional_timestamp
from sampling_tool.core.models import (
    AuditEvent,
    Dataset,
    Engagement,
    SampleResult,
)
from sampling_tool.core.provenance import SamplingProvenance
from sampling_tool.io._atomic import AtomicReplaceError, atomic_output
from sampling_tool.io._xlsx_safe import safe_row
from sampling_tool.io._xlsx_util import autosize_columns as autosize_columns
from sampling_tool.io.charts import render_bar_chart_bytes
from sampling_tool.io.exporter import ExportError

_HEADER_FILL: Final = PatternFill(
    start_color=excel_argb(BDO_RED), end_color=excel_argb(BDO_RED), fill_type="solid"
)
_HEADER_FONT: Final = Font(bold=True, color="FFFFFFFF")
_HEADER_ALIGN: Final = Alignment(vertical="center", horizontal="left")

SHEET_UEBERSICHT: Final[str] = "Übersicht"
SHEET_AUDIT_TRAIL: Final[str] = "AuditTrail"
SHEET_SAMPLES: Final[str] = "Samples"
SHEET_STATISTIKEN: Final[str] = "Statistiken"

ALL_SHEETS: Final[frozenset[str]] = frozenset(
    {SHEET_UEBERSICHT, SHEET_AUDIT_TRAIL, SHEET_SAMPLES, SHEET_STATISTIKEN}
)


class MultiSheetReportExporter:
    """Erzeugt einen Engagement-Komplett-Bericht als Multi-Sheet-Excel."""

    def export(
        self,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
        output_path: Path,
        sheets: set[str] | None = None,
        dataset_ids_by_sample: dict[int, int] | None = None,
    ) -> Path:
        """Schreibt die .xlsx atomar nach `output_path` und gibt den Pfad zurück.

        `sheets` filtert die geschriebenen Sheets nach Namen aus
        `ALL_SHEETS`. `None` oder leeres Set ⇒ alle Sheets. `dataset_ids_by_
        sample` (Sprint 43 / A-001) bildet `sample.id -> dataset.id` ab, weil
        die flache `samples`-Liste diese Zuordnung sonst nicht kennt.
        """
        target = (
            output_path
            if output_path.suffix.lower() == ".xlsx"
            else output_path.with_suffix(".xlsx")
        )

        active = ALL_SHEETS if not sheets else (sheets & ALL_SHEETS)
        if not active:
            active = ALL_SHEETS

        wb = Workbook()
        try:
            if SHEET_UEBERSICHT in active:
                self._write_uebersicht(wb, engagement, datasets, samples, audit_events)
            if SHEET_AUDIT_TRAIL in active:
                self._write_audit_trail(wb, audit_events)
            if SHEET_SAMPLES in active:
                self._write_samples(wb, samples, dataset_ids_by_sample or {})
            if SHEET_STATISTIKEN in active:
                self._write_statistiken(wb, samples, audit_events)

            # Das Default-Sheet "Sheet" von openpyxl bleibt übrig, wenn das
            # erste tatsächlich geschriebene Sheet via `create_sheet` angelegt
            # wurde – wir entfernen es. Wurde nur „Übersicht" geschrieben,
            # hat es bereits den Titel überschrieben.
            for ws in list(wb.worksheets):
                if ws.title == "Sheet":
                    wb.remove(ws)

            try:
                with atomic_output(target) as tmp:
                    wb.save(tmp)
            except AtomicReplaceError as exc:
                raise ExportError(
                    f"Die Zieldatei „{target.name}“ konnte nicht geschrieben werden: "
                    f"{exc}. Sie ist möglicherweise in Excel geöffnet oder es fehlen "
                    "Schreibrechte."
                ) from exc
        finally:
            wb.close()

        return target

    # ---- Sheets ---------------------------------------------------------

    def _write_uebersicht(
        self,
        wb: Workbook,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        ws = wb.active
        assert ws is not None
        ws.title = "1. Übersicht"
        ws.append(["BDO Audit Sampling Tool – Projekt-Bericht"])
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color=excel_argb(BDO_RED))
        ws.append([])

        meta_rows: list[tuple[str, Any]] = [
            ("Mandant", engagement.client_name),
            ("Prüfungstyp", engagement.audit_type or "—"),
            ("Auditor", engagement.auditor_name),
            ("Position", engagement.auditor_position or "—"),
            ("Projekt-ID", engagement.id if engagement.id is not None else "—"),
            ("Bericht erstellt", _format_now()),
        ]
        for label, value in meta_rows:
            safe_row(ws, [label, value])

        ws.append([])
        ws.append(["Statistiken"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color=excel_argb(BDO_RED))
        ws.append(["Datasets", len(datasets)])
        ws.append(["Samples", len(samples)])
        ws.append(["Audit-Events", len(audit_events)])
        if audit_events:
            latest = max(audit_events, key=lambda e: e.timestamp)
            ws.append(["Letzte Aktivität", format_optional_timestamp(latest.timestamp)])
        else:
            ws.append(["Letzte Aktivität", "—"])

        autosize_columns(ws, 2, min_width=12)
        # Erstes Spalten-Label fett.
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value not in ("", None):
                cell.font = Font(bold=True)

    def _write_audit_trail(self, wb: Workbook, events: list[AuditEvent]) -> None:
        ws = wb.create_sheet("2. AuditTrail")
        header = [
            "Zeitstempel",
            "Aktion",
            "User",
            "Sample",
            "Größe",
            "%",
            "Seed",
            "Datei",
            "Korrektur",
            "Details",
        ]
        ws.append(header)
        _style_header_row(ws, len(header))

        chronological = sorted(events, key=lambda e: (e.timestamp, e.id or 0))
        for evt in chronological:
            safe_row(
                ws,
                [
                    format_optional_timestamp(evt.timestamp),
                    evt.event_type,
                    evt.user_name,
                    evt.sample_id if evt.sample_id is not None else "—",
                    evt.sample_size if evt.sample_size is not None else "—",
                    (f"{evt.sample_percent:.2f}" if evt.sample_percent is not None else "—"),
                    evt.seed if evt.seed is not None else "—",
                    Path(evt.export_file or evt.import_file or "").name or "—",
                    f"#{evt.corrects_event_id}" if evt.corrects_event_id is not None else "—",
                    format_audit_details(evt.details),
                ],
            )
        autosize_columns(ws, len(header), min_width=12)
        ws.freeze_panes = "A2"

    def _write_samples(
        self,
        wb: Workbook,
        samples: list[SampleResult],
        dataset_ids_by_sample: dict[int, int],
    ) -> None:
        ws = wb.create_sheet("3. Samples")
        header = [
            "ID",
            "Methode",
            "Angeforderte Größe",
            "Tatsächliche Größe",
            "Population",
            "Anteil %",
            "Seed",
            "Filter-Feld",
            "Filter-Operator",
            "Filter-Wert",
            "Cluster-Feld",
            "Stratum-Feld",
            "Stratify-Modus",
            "Parent-Sample-ID",
            "Algorithmus-Version",
            "Dataset-ID",
            "Erstellt am",
            "Erstellt von",
        ]
        ws.append(header)
        _style_header_row(ws, len(header))

        ordered = sorted(samples, key=lambda s: s.drawn_at)
        for sample in ordered:
            dataset_id = dataset_ids_by_sample.get(sample.id) if sample.id is not None else None
            provenance = SamplingProvenance.from_sample_result(
                sample, dataset_id=dataset_id, app_version=__version__
            )
            percent = (
                sample.actual_size / sample.population_size * 100.0
                if sample.population_size
                else 0.0
            )
            safe_row(
                ws,
                [
                    sample.id if sample.id is not None else "—",
                    provenance.method,
                    provenance.size_requested,
                    provenance.size_actual,
                    provenance.population_size,
                    round(percent, 2),
                    provenance.seed,
                    provenance.filter_field or "—",
                    provenance.filter_operator_symbol,
                    str(provenance.filter_value) if provenance.filter_value is not None else "—",
                    provenance.cluster_field or "—",
                    provenance.stratum_field or "—",
                    provenance.stratify_mode,
                    provenance.parent_sample_id if provenance.parent_sample_id is not None else "—",
                    provenance.algorithm_version,
                    provenance.dataset_id if provenance.dataset_id is not None else "—",
                    format_optional_timestamp(sample.drawn_at),
                    provenance.created_by,
                ],
            )
        autosize_columns(ws, len(header), min_width=12)
        ws.freeze_panes = "A2"

    def _write_statistiken(
        self,
        wb: Workbook,
        samples: list[SampleResult],
        events: list[AuditEvent],
    ) -> None:
        ws = wb.create_sheet("4. Statistiken")

        # Methoden-Verteilung als Tabelle.
        ws.append(["Methode", "Anzahl"])
        _style_header_row(ws, 2)
        method_counts: Counter[str] = Counter(s.config.method.value for s in samples)
        for method, count in method_counts.most_common():
            ws.append([method, count])

        # Event-Typen-Verteilung als zweite Tabelle.
        start = ws.max_row + 2
        ws.cell(row=start, column=1, value="Eventtyp")
        ws.cell(row=start, column=2, value="Anzahl")
        for col in (1, 2):
            cell = ws.cell(row=start, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        type_counts: Counter[str] = Counter(e.event_type for e in events)
        for event_type, count in type_counts.most_common():
            ws.append([event_type, count])

        autosize_columns(ws, 2, min_width=12)

        # Chart als eingebettetes PNG.
        if method_counts:
            labels = list(method_counts.keys())
            values = [float(method_counts[k]) for k in labels]
            png_bytes = render_bar_chart_bytes(
                labels, values, title="Sampling-Methoden", width=480, height=240
            )
            image = XlImage(BytesIO(png_bytes))
            image.anchor = "E2"
            ws.add_image(image)


# ---------------------------------------------------------------------------
# Hilfen
# ---------------------------------------------------------------------------


def _style_header_row(ws: Worksheet, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _format_now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
