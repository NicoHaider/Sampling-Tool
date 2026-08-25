"""Globale pytest-Fixtures.

Excel-/CSV-Test-Fixtures werden hier programmatisch erzeugt (kein Binär-Blob im
Repo). Scope `session`, damit jede Datei genau einmal angelegt wird.
"""

from __future__ import annotations

import os
from collections.abc import Iterator, Sequence
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

# Qt-Headless-Default. Wenn weder DISPLAY noch ein expliziter Plattform-Wert
# gesetzt ist, läuft Qt im offscreen-Modus – damit UI-Tests in CI funktionieren
# und auf Dev-Maschinen kein lästiges Fenster aufpoppt.
if "QT_QPA_PLATFORM" not in os.environ and "DISPLAY" not in os.environ:
    os.environ["QT_QPA_PLATFORM"] = "offscreen"

from sampling_tool.core.models import (
    Dataset,
    DatasetRow,
    Engagement,
    SampleConfig,
    SampleResult,
    SamplingMethod,
)
from sampling_tool.persistence.database import Database
from sampling_tool.persistence.repositories import (
    DatasetRepo,
    EngagementRepo,
    SampleRepo,
)
from tests._test_floor import (
    ENFORCE_TEST_FLOOR_ENV,
    EXECUTED_FLOOR,
    TEST_FLOOR,
    check_executed_floor,
    check_test_floor,
)

# ---------------------------------------------------------------------------
# Testmengen-Wächter (Sprint 77 / Befund #8) + Ausführungs-Wächter (Sprint 79)
# ---------------------------------------------------------------------------

#: nodeids der übersprungenen Tests dieser Session.
#:
#: Ein SET von nodeids, nicht ein Zähler über Reports: ein `skipif` meldet in der
#: Setup-Phase, ein `pytest.skip()` im Testkörper in der Call-Phase (und liefert
#: zusätzlich einen Teardown-Report). Ohne Dedup zählte dieselbe Ursache je nach
#: Herkunft unterschiedlich, und die Kennzahl hinge daran, WIE ein Test
#: übersprungen wird statt DASS er es wird.
#:
#: Modul-global statt `config.stash`, weil pro Prozess genau eine Session läuft –
#: das Projekt nutzt kein pytest-xdist (geprüft: nur pytest-cov und pytest-qt).
_skipped_nodeids: set[str] = set()


def pytest_runtest_logreport(report: pytest.TestReport) -> None:
    """Sammelt die übersprungenen Tests für den Ausführungs-Wächter.

    `wasxfail` schließt `xfail`-Ergebnisse aus: ein erwartet fehlschlagender Test
    liefert ebenfalls `report.skipped`, ist aber ausgeführt worden – ihn
    mitzuzählen würde den Ausführungsstand kleinrechnen.
    """
    if report.skipped and not hasattr(report, "wasxfail"):
        _skipped_nodeids.add(report.nodeid)


def pytest_sessionfinish(session: pytest.Session, exitstatus: int) -> None:
    """Schlägt an, wenn die volle Suite zu wenig sammelt ODER zu wenig ausführt.

    Läuft NUR bei gesetzter `SAMPLING_TOOL_ENFORCE_TEST_FLOOR=1`, also
    ausschließlich in den beiden Workflows – lokal bleiben beide Wächter ein
    No-Op, damit Teil-Läufe nicht fehlschlagen.

    `session.testscollected` ist bewusst die geprüfte Zahl: pytest setzt sie
    NACH `pytest_collection_modifyitems` (`_pytest/main.py`), sie berücksichtigt
    also auch Abwahl per `-k`/`-m`/`--deselect`.
    """
    if os.environ.get(ENFORCE_TEST_FLOOR_ENV) != "1":
        return

    collected = session.testscollected
    skipped = len(_skipped_nodeids)
    reporter = session.config.pluginmanager.get_plugin("terminalreporter")

    # Die Messung steht IMMER im Log, nicht nur im Verstoß-Fall: der Floor wird
    # aus echten CI-Läufen abgelesen (Sprint 79 / §2.5), und dafür muss die Zahl
    # je Plattform ohne Zusatzwerkzeug im Protokoll stehen.
    if reporter is not None:
        reporter.write_line("")
        reporter.write_line(
            f"Testmengen-Wächter: {collected} gesammelt, {skipped} übersprungen, "
            f"{collected - skipped} ausgeführt "
            f"(Floors: {TEST_FLOOR} gesammelt / {EXECUTED_FLOOR} ausgeführt)."
        )

    problems = [
        check_test_floor(collected, TEST_FLOOR),
        check_executed_floor(collected, skipped, EXECUTED_FLOOR),
    ]
    failures = [problem for problem in problems if problem is not None]
    if not failures:
        return

    if reporter is not None:
        for problem in failures:
            reporter.write_line(problem, red=True, bold=True)
    # Einen bereits gesetzten, schwerwiegenderen Status (Abbruch, interner
    # Fehler, Usage-Error) nicht überschreiben – nur den grünen Lauf kippen.
    if session.exitstatus == pytest.ExitCode.OK:
        session.exitstatus = pytest.ExitCode.TESTS_FAILED


@pytest.fixture
def db() -> Iterator[Database]:
    """Frische, migrierte In-Memory-Datenbank pro Test."""
    database = Database(Path(":memory:"))
    database.migrate()
    yield database
    database.close()


@pytest.fixture
def engagement_id(db: Database) -> int:
    """Legt ein Default-Engagement an und gibt dessen DB-id zurück."""
    repo = EngagementRepo(db.connect())
    eng = repo.get_or_create(
        Engagement(
            auditor_name="Anna Auditorin",
            client_name="ACME GmbH",
            auditor_position="Senior Auditor",
            audit_type="ISAE 3402 Typ II",
        )
    )
    assert eng.id is not None
    return eng.id


@pytest.fixture
def dataset_id(db: Database, engagement_id: int) -> int:
    """DB-id des Dummy-Datasets, das `sample_id` zugrunde liegt (Sprint 43:
    `log_sampling` braucht jetzt eine Dataset-Zuordnung)."""
    ds = DatasetRepo(db.connect()).create(
        Dataset(name="dummy", columns=("a",), engagement_id=engagement_id),
        (DatasetRow(row_id=1, values={"a": 1}),),
    )
    assert ds.id is not None
    return ds.id


@pytest.fixture
def sample_id(db: Database, dataset_id: int) -> int:
    """Persistiert eine Dummy-Stichprobe auf `dataset_id` und liefert die
    `samples.id`. Wird von Audit- und Undo-Tests genutzt, deren FKs auf
    `samples(id)` greifen."""
    cfg = SampleConfig(method=SamplingMethod.SIMPLE, size=1, seed=1)
    result = SampleResult(config=cfg, selected_row_ids=(1,), population_size=1)
    return SampleRepo(db.connect()).create_from_result(result, dataset_id, "test")


# ---------------------------------------------------------------------------
# Sprint-11.1-Helper: kleine in-Memory Datasets bauen
# ---------------------------------------------------------------------------


def make_test_dataset(
    rows: Sequence[DatasetRow],
    *,
    name: str = "test",
    columns: tuple[str, ...] | None = None,
    engagement_id: int | None = None,
    id: int | None = None,
) -> Dataset:
    """Baut ein `Dataset` (Metadaten only) mit korrektem `row_count`.

    Wenn `columns` nicht übergeben wird, werden sie aus den Keys der
    ersten Row abgeleitet. Wenn rows leer ist: leere columns tuple.
    """
    if columns is None:
        columns = tuple(rows[0].values.keys()) if rows else ()
    return Dataset(
        name=name,
        columns=columns,
        row_count=len(rows),
        engagement_id=engagement_id,
        id=id,
    )


# ---------------------------------------------------------------------------
# Excel-/CSV-Fixtures – zur Laufzeit erzeugt, session-scoped (einmal pro Run).
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def fixtures_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Gemeinsames Ablage-Verzeichnis für generierte I/O-Test-Fixtures."""
    return tmp_path_factory.mktemp("io_fixtures")


@pytest.fixture(scope="session")
def simple_xlsx(fixtures_dir: Path) -> Path:
    """Einfache .xlsx mit 4 Spalten (str, int, float, datetime) und 10 Zeilen."""
    path = fixtures_dir / "simple_data.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Daten"
    ws.append(["Name", "Betrag", "Quote", "Buchungsdatum"])
    for i in range(1, 11):
        ws.append(
            [
                f"Posten {i}",
                100 + i,
                round(0.1 * i, 4),
                datetime(2026, 1, i, 9, 0, 0),
            ]
        )
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def multi_sheet_xlsx(fixtures_dir: Path) -> Path:
    """xlsx mit zwei Sheets unterschiedlicher Spalten."""
    path = fixtures_dir / "multi_sheet.xlsx"
    wb = Workbook()
    first = wb.active
    assert first is not None
    first.title = "Buchungen"
    first.append(["BuchungsID", "Betrag"])
    for i in range(1, 6):
        first.append([f"B{i:03d}", 10 * i])

    second = wb.create_sheet("Stammdaten")
    second.append(["KundenID", "Land"])
    for i, country in enumerate(["AUT", "DEU", "CHE"], start=1):
        second.append([f"K{i:03d}", country])
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def xlsm_macro(fixtures_dir: Path) -> Path:
    """xlsm-Datei (Makros werden ignoriert, Daten gelesen)."""
    path = fixtures_dir / "with_macro.xlsm"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.append([3, 4])
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def empty_xlsx(fixtures_dir: Path) -> Path:
    """xlsx ohne Inhalt – muss vom Importer mit klarer Fehlermeldung quittiert werden."""
    path = fixtures_dir / "empty.xlsx"
    wb = Workbook()
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def leading_blank_xlsx(fixtures_dir: Path) -> Path:
    """xlsx mit drei Leerzeilen vor dem Header – Header-Detection muss greifen."""
    path = fixtures_dir / "leading_blank.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([None, None, None])
    ws.append([None, None, None])
    ws.append([None, None, None])
    ws.append(["Konto", "Bezeichnung", "Saldo"])
    ws.append(["1000", "Kasse", 500.50])
    ws.append(["2000", "Bank", 1234.00])
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def duplicate_columns_xlsx(fixtures_dir: Path) -> Path:
    """xlsx mit doppelten Spalten-Namen – Importer muss Suffixe vergeben."""
    path = fixtures_dir / "dup_columns.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Betrag", "Betrag", "Betrag"])
    ws.append([10, 20, 30])
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def utf8_csv(fixtures_dir: Path) -> Path:
    """CSV in UTF-8 mit Umlauten."""
    path = fixtures_dir / "utf8.csv"
    path.write_text(
        "Name,Stadt\nMüller,Wien\nKöck,Salzburg\n",
        encoding="utf-8",
    )
    return path


@pytest.fixture(scope="session")
def utf8_bom_csv(fixtures_dir: Path) -> Path:
    """CSV in UTF-8 mit BOM (typisch für Excel-Export)."""
    path = fixtures_dir / "utf8_bom.csv"
    path.write_text(
        "Name,Stadt\nMüller,Wien\nKöck,Salzburg\n",
        encoding="utf-8-sig",
    )
    return path


@pytest.fixture(scope="session")
def cp1252_csv(fixtures_dir: Path) -> Path:
    """CSV in cp1252 (Windows-Default) mit Umlauten."""
    path = fixtures_dir / "cp1252.csv"
    path.write_bytes("Name,Stadt\nMüller,Wien\nKöck,Salzburg\n".encode("cp1252"))
    return path


@pytest.fixture(scope="session")
def umlaut_csv(utf8_csv: Path) -> Path:
    """Alias für `utf8_csv` – häufig genutzte „mit Umlauten"-Datei."""
    return utf8_csv
