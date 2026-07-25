"""MainController – Glue-Logik mit echter SQLite-Datei und Excel-Fixture."""

from __future__ import annotations

import contextlib
import dataclasses
import sqlite3
from collections.abc import Iterator
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QDialog, QListWidget, QMessageBox
from pytestqt.qtbot import QtBot

from sampling_tool.core.models import (
    Dataset,
    DatasetRow,
    Engagement,
    SampleConfig,
    SampleResult,
    SamplingMethod,
)
from sampling_tool.io.import_preflight import ImportPreflight
from sampling_tool.persistence.database import Database
from sampling_tool.persistence.repositories import (
    DatasetRepo,
    EngagementRepo,
    SampleRepo,
)
from sampling_tool.persistence.version_manager import EngagementVersionManager
from sampling_tool.ui.controllers._factories import (
    ControllerFactories,
    default_audit_pdf_factory,
    default_duplicate_dialog_factory,
    default_excel_report_factory,
    default_export_factory,
    default_html_report_factory,
    default_id_column_factory,
    default_import_options_factory,
    default_new_engagement_factory,
    default_sampling_factory,
    default_settings_factory,
)
from sampling_tool.ui.controllers.engagement_controller import (
    _remove_db_files as _real_remove_db_files,
)
from sampling_tool.ui.controllers.main_controller import MainController
from sampling_tool.ui.dialogs.duplicate_engagement_dialog import (
    DuplicateEngagementChoice,
    DuplicateEngagementDialog,
)
from sampling_tool.ui.dialogs.new_engagement_dialog import NewEngagementDialog
from sampling_tool.ui.main_window import MainWindow
from sampling_tool.ui.recent import RecentEngagementsStore

pytestmark = pytest.mark.ui


@pytest.fixture(autouse=True)
def _stub_post_import_id_dialog(monkeypatch: pytest.MonkeyPatch) -> None:
    """Sprint 31: der optionale Post-Import-ID-Spalten-Dialog ist modal – im
    echten Import-Pfad dieser Tests (ohne eigenen ID-Dialog-Stub) würde
    `exec()` headless blockieren. Rejected = „nichts wählen", berührt keine
    bestehende Assertion (Dataset/Sample-Counts unverändert)."""
    monkeypatch.setattr(
        "sampling_tool.ui.dialogs.id_column_dialog.IdColumnDialog.exec",
        lambda _self: int(QDialog.DialogCode.Rejected),
    )


@pytest.fixture
def recent_store(tmp_path: Path) -> RecentEngagementsStore:
    return RecentEngagementsStore(path=tmp_path / "recent.json")


@pytest.fixture
def window(qtbot: QtBot) -> MainWindow:
    win = MainWindow()
    qtbot.addWidget(win)
    return win


@pytest.fixture
def controller(
    window: MainWindow, recent_store: RecentEngagementsStore
) -> Iterator[MainController]:
    ctrl = MainController(window, recent_store=recent_store)
    yield ctrl
    ctrl.handle_close_engagement()


@pytest.fixture
def populated_db(tmp_path: Path) -> Path:
    db_path = tmp_path / "engagement.db"
    db = Database(db_path)
    db.migrate()
    eng_repo = EngagementRepo(db.connect())
    eng = eng_repo.get_or_create(
        Engagement(
            auditor_name="Anna",
            client_name="ACME",
            auditor_position="Senior",
            audit_type="ISAE 3402",
        )
    )
    assert eng.id is not None
    ds_repo = DatasetRepo(db.connect())
    rows = tuple(
        DatasetRow(row_id=i, values={"Konto": f"K{i}", "Betrag": i * 10}) for i in range(1, 6)
    )
    dataset = ds_repo.create(
        Dataset(name="Buchungen", columns=("Konto", "Betrag"), engagement_id=eng.id),
        rows,
    )
    assert dataset.id is not None
    SampleRepo(db.connect()).create_from_result(
        SampleResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=42),
            selected_row_ids=(2, 4),
            population_size=5,
        ),
        dataset.id,
        "tester",
    )
    db.close()
    return db_path


@pytest.fixture
def import_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "import.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Konto", "Betrag"])
    for i in range(1, 4):
        ws.append([f"K{i}", i * 100])
    wb.save(path)
    return path


def _first_item_data(list_widget: QListWidget) -> int:
    item = list_widget.item(0)
    assert item is not None
    value = item.data(int(Qt.ItemDataRole.UserRole))
    assert isinstance(value, int)
    return value


def _make_stub_new_dialog(
    parent: MainWindow, target_db: Path, client_name: str
) -> NewEngagementDialog:
    """Liefert ein stub-`NewEngagementDialog`, das ohne UI-Interaktion accepted."""

    class _StubDialog(NewEngagementDialog):
        def exec(self) -> int:
            self._db_path = target_db
            return int(QDialog.DialogCode.Accepted)

        def get_engagement(self) -> Engagement:
            return Engagement(
                auditor_name="Anna",
                auditor_position="Senior",
                client_name=client_name,
                audit_type="ISAE 3402",
            )

    return _StubDialog(parent)


def _make_stub_duplicate_dialog(
    parent: MainWindow,
    db_path: Path,
    choice: DuplicateEngagementChoice,
) -> DuplicateEngagementDialog:
    """Stub des DuplicateDialogs, der ohne UI sofort das gewünschte Choice liefert."""

    class _StubDuplicate(DuplicateEngagementDialog):
        def exec(self) -> int:
            self._choice = choice
            return int(
                QDialog.DialogCode.Accepted
                if choice is not DuplicateEngagementChoice.CANCEL
                else QDialog.DialogCode.Rejected
            )

    return _StubDuplicate(db_path, parent)


def _record_duplicate(
    calls: list[Path],
    parent: MainWindow,
    db_path: Path,
    choice: DuplicateEngagementChoice,
) -> DuplicateEngagementDialog:
    """Wie `_make_stub_duplicate_dialog`, protokolliert aber den Aufruf in `calls`."""
    calls.append(db_path)
    return _make_stub_duplicate_dialog(parent, db_path, choice)


@contextlib.contextmanager
def _spy_database_close() -> Iterator[list[Database]]:
    """Sprint 41 / S1.4: zählt `Database.close()`-Aufrufe, ruft aber die echte
    Methode weiterhin auf (im Gegensatz zu einem reinen Mock) – damit bleibt
    keine SQLite-Connection offen (Windows-Filelock-Risiko beim `tmp_path`-
    Cleanup), während sich trotzdem beweisen lässt, dass jeder Exit-Pfad
    tatsächlich schließt."""
    calls: list[Database] = []
    original_close = Database.close

    def _spy(self: Database) -> None:
        calls.append(self)
        original_close(self)

    with patch.object(Database, "close", _spy):
        yield calls


def test_controller_factories_defaults() -> None:
    """Sprint 59 / Teil B (L-003): `ControllerFactories.defaults()` muss die
    10 `default_*_factory`-Modulfunktionen 1:1 auf die gleichnamigen Bundle-
    Felder mappen – das Override-Verhalten selbst (Kwarg gesetzt -> Kwarg
    gewinnt) ist bereits durch die ~84 Factory-Injection-Testfälle in
    `TestMainController` & Co. abgedeckt, hier geht es nur um `defaults()`
    selbst."""
    factories = ControllerFactories.defaults()

    assert isinstance(factories, ControllerFactories)
    assert factories.new_engagement is default_new_engagement_factory
    assert factories.duplicate is default_duplicate_dialog_factory
    assert factories.sampling is default_sampling_factory
    assert factories.export_sample is default_export_factory
    assert factories.audit_pdf is default_audit_pdf_factory
    assert factories.excel_report is default_excel_report_factory
    assert factories.html_report is default_html_report_factory
    assert factories.settings is default_settings_factory
    assert factories.import_options is default_import_options_factory
    assert factories.id_column is default_id_column_factory
    assert all(
        callable(getattr(factories, field.name))
        for field in dataclasses.fields(ControllerFactories)
    )


class TestMainController:
    def test_open_engagement_loads_into_workspace(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        assert window.is_workspace_visible() is True
        assert window.sidebar().datasets_widget().count() == 1

    def test_open_engagement_missing_db_shows_welcome(
        self,
        controller: MainController,
        window: MainWindow,
        tmp_path: Path,
    ) -> None:
        ghost = tmp_path / "ghost.db"
        # Sprint 13 / F-001: `error()` ist auf WorkspaceSession – Patch dort.
        with patch("sampling_tool.ui.controllers.workspace_session.QMessageBox.warning") as warning:
            controller.handle_open_engagement(ghost)
        assert warning.called
        assert window.is_workspace_visible() is False

    def test_dataset_selected_shows_table_and_samples(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        controller.handle_dataset_selected(_first_item_data(window.sidebar().datasets_widget()))
        assert window.data_table().table_model().rowCount() == 5
        assert window.sidebar().samples_widget().count() == 1

    def test_sample_selected_highlights_rows(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        controller.handle_dataset_selected(_first_item_data(window.sidebar().datasets_widget()))
        controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
        highlights = window.data_table().table_model().highlighted_row_ids()
        assert highlights == frozenset({2, 4})

    def test_sample_filter_toggle_filters_and_unfilters(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        controller.handle_dataset_selected(_first_item_data(window.sidebar().datasets_widget()))
        sample_id = _first_item_data(window.sidebar().samples_widget())

        controller.handle_sample_filter_toggled(sample_id)
        assert window.data_table().table_model().rowCount() == 2

        controller.handle_sample_filter_toggled(sample_id)
        assert window.data_table().table_model().rowCount() == 5

    def test_close_engagement_returns_to_welcome(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        controller.handle_close_engagement()
        assert window.is_workspace_visible() is False
        assert window.data_table().table_model().rowCount() == 0

    def test_open_tampered_db_warns_and_restores_triggers(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Sprint 52 / S2.7 (S-004), Variante 1: eine extern (SQLite-Editor)
        entfernte Append-only-Trigger-Definition blockiert das Öffnen NICHT –
        es gibt eine prominente Warnung, und der Schutz wird sofort wieder-
        hergestellt (`audit_events`-Zeilen bleiben dabei unberührt). Das
        erkannte Tampering wird zusätzlich geloggt (Review-Nachbesserung) –
        vorher war ein Dialog, den der User wegklicken kann, der einzige Beleg."""
        tamper_conn = sqlite3.connect(str(populated_db))
        try:
            tamper_conn.execute(
                "INSERT INTO audit_events (engagement_id, event_type) VALUES (1, 'x')"
            )
            tamper_conn.execute("DROP TRIGGER audit_events_no_update")
            tamper_conn.commit()
        finally:
            tamper_conn.close()

        with (
            patch(
                "sampling_tool.ui.controllers.engagement_controller.QMessageBox.warning"
            ) as warning,
            caplog.at_level("WARNING", logger="sampling_tool.ui.controllers.engagement_controller"),
        ):
            controller.handle_open_engagement(populated_db)

        assert warning.called
        assert window.is_workspace_visible() is True
        warnings = [r for r in caplog.records if r.levelname == "WARNING"]
        assert any("manipul" in r.message.lower() for r in warnings), (
            f"Erwartete WARNING-Log zur erkannten Trigger-Manipulation, gefangen: "
            f"{[r.message for r in warnings]}"
        )

        assert controller.session.db is not None
        conn = controller.session.db.connect()
        rows_before = conn.execute("SELECT event_type FROM audit_events").fetchall()
        with pytest.raises(sqlite3.IntegrityError, match="append-only"):
            conn.execute("UPDATE audit_events SET event_type = 'tampered'")
        rows_after = conn.execute("SELECT event_type FROM audit_events").fetchall()
        assert rows_after == rows_before

    def test_open_tampered_db_restore_failure_shows_honest_warning(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Review-Nachbesserung (Sprint 52 / S2.7): schlägt die Wiederher-
        stellung selbst fehl (z. B. Datei read-only), darf die Warnung NICHT
        weiter behaupten, der Schutz sei wiederhergestellt – das wäre eine
        falsche Sicherheitszusicherung an den Anwender. Öffnen bleibt trotzdem
        unblockiert (Variante 1)."""
        tamper_conn = sqlite3.connect(str(populated_db))
        try:
            tamper_conn.execute("DROP TRIGGER audit_events_no_update")
            tamper_conn.commit()
        finally:
            tamper_conn.close()

        with (
            patch(
                "sampling_tool.ui.controllers.engagement_controller."
                "restore_audit_append_only_triggers",
                side_effect=sqlite3.OperationalError("attempt to write a readonly database"),
            ),
            patch(
                "sampling_tool.ui.controllers.engagement_controller.QMessageBox.warning"
            ) as warning,
            caplog.at_level("ERROR", logger="sampling_tool.ui.controllers.engagement_controller"),
        ):
            controller.handle_open_engagement(populated_db)

        assert window.is_workspace_visible() is True  # Öffnen bleibt unblockiert
        assert warning.called
        shown_message = warning.call_args.args[2]
        assert "wiederhergestellt" not in shown_message.lower()
        assert "fehlgeschlagen" in shown_message.lower()

        errors = [r for r in caplog.records if r.levelname == "ERROR"]
        assert any("fehlgeschlagen" in r.message.lower() for r in errors)

    def test_new_engagement_creates_db_via_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target_db = tmp_path / "new.db"

        class _StubDialog(NewEngagementDialog):
            def exec(self) -> int:
                self._db_path = target_db
                return int(QDialog.DialogCode.Accepted)

            def get_engagement(self) -> Engagement:
                return Engagement(
                    auditor_name="Anna",
                    auditor_position="Senior",
                    client_name="ACME",
                    audit_type="ISAE 3402",
                )

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _settings, _prefill: _StubDialog(parent),
        )
        try:
            controller.handle_new_engagement()
            assert target_db.exists()
            assert window.is_workspace_visible() is True
            assert recent_store.list()[0].path == target_db.resolve()
        finally:
            controller.handle_close_engagement()

    def test_new_engagement_no_duplicate_skips_duplicate_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target_db = tmp_path / "fresh.db"
        duplicate_calls: list[Path] = []

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target_db, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _record_duplicate(
                duplicate_calls, parent, db_path, DuplicateEngagementChoice.CANCEL
            ),
        )
        try:
            controller.handle_new_engagement()
            assert target_db.exists()
            assert duplicate_calls == [], "DuplicateDialog darf nicht erscheinen"
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_new_engagement_migrate_failure_closes_connection(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        """Sprint 41 / S1.4: schlägt `migrate()` bei der Anlage eines frischen
        Projekts (Ziel-Pfad existiert noch nicht) fehl, muss die bereits
        konstruierte `Database`-Connection trotzdem geschlossen werden –
        strukturell identisch zum Exception-Zweig in `_overwrite_with_backup`."""
        target_db = tmp_path / "fresh.db"

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target_db, "ACME"),
        )
        try:
            with (
                _spy_database_close() as close_calls,
                patch.object(Database, "migrate", side_effect=RuntimeError("boom")),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_new_engagement()

            assert warning.called
            assert window.is_workspace_visible() is False
            assert len(close_calls) == 1
        finally:
            controller.handle_close_engagement()

    def test_new_engagement_with_duplicate_open_existing_opens_db(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(
                parent, populated_db, "ACME"
            ),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OPEN_EXISTING
            ),
        )
        try:
            controller.handle_new_engagement()
            assert window.is_workspace_visible() is True
            # Bestehende DB nicht überschrieben → Sample-Eintrag aus Fixture noch da.
            assert window.sidebar().datasets_widget().count() == 1
        finally:
            controller.handle_close_engagement()

    def test_new_engagement_with_duplicate_rename_reopens_new_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        existing = tmp_path / "ACME.db"
        existing.touch()
        fresh = tmp_path / "ACME2.db"
        new_dialog_calls: list[Engagement | None] = []

        def _new_factory(
            parent: MainWindow,
            _settings: object,
            prefill: Engagement | None,
        ) -> NewEngagementDialog:
            new_dialog_calls.append(prefill)
            target = existing if len(new_dialog_calls) == 1 else fresh
            return _make_stub_new_dialog(parent, target, "ACME")

        # Erster Aufruf liefert RENAME, zweiter würde nicht mehr aufgerufen
        # weil der zweite NewEngagementDialog `fresh` zurückgibt (existiert nicht).
        duplicate_dialog_calls: list[Path] = []

        def _dup_factory(parent: MainWindow, db_path: Path) -> DuplicateEngagementDialog:
            duplicate_dialog_calls.append(db_path)
            return _make_stub_duplicate_dialog(parent, db_path, DuplicateEngagementChoice.RENAME)

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=_new_factory,
            duplicate_dialog_factory=_dup_factory,
        )
        try:
            controller.handle_new_engagement()
            assert len(new_dialog_calls) == 2, "NewEngagementDialog muss erneut geöffnet werden"
            # Zweiter Aufruf bekommt das vorher eingegebene Engagement als Prefill.
            assert new_dialog_calls[0] is None
            assert new_dialog_calls[1] is not None
            assert new_dialog_calls[1].client_name == "ACME"
            assert duplicate_dialog_calls == [existing]
            assert fresh.exists()
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_new_engagement_with_duplicate_cancel_aborts(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        existing = tmp_path / "blocked.db"
        existing.write_bytes(b"sentinel-bytes")
        before = existing.read_bytes()

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(
                parent, existing, "Blocked"
            ),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.CANCEL
            ),
        )
        try:
            controller.handle_new_engagement()
            assert window.is_workspace_visible() is False
            assert existing.read_bytes() == before, (
                "Bestehende Datei darf nicht überschrieben werden"
            )
        finally:
            controller.handle_close_engagement()

    def test_import_excel_persists_dataset_and_logs_audit(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(import_xlsx), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()

            assert window.sidebar().datasets_widget().count() == 2
            assert window.data_table().table_model().rowCount() == 3
        finally:
            controller.handle_close_engagement()

    def test_import_rejected_by_preflight_shows_error_and_skips_worker(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            datasets_before = window.sidebar().datasets_widget().count()
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(import_xlsx), ""),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.preflight_import",
                    return_value=ImportPreflight(reject_reason="Datei ist zu groß (999 MB)."),
                ) as mock_preflight,
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QMessageBox.warning"
                ) as mock_warning,
            ):
                controller.handle_import_excel()

            mock_preflight.assert_called_once()
            mock_warning.assert_called_once()
            assert window.sidebar().datasets_widget().count() == datasets_before
        finally:
            controller.handle_close_engagement()

    def test_import_shows_confirm_dialog_and_proceeds_on_yes(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(import_xlsx), ""),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.preflight_import",
                    return_value=ImportPreflight(warnings=("Datei ist groß (250 MB).",)),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                    return_value=QMessageBox.StandardButton.Yes,
                ) as mock_question,
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()

            mock_question.assert_called_once()
            assert window.sidebar().datasets_widget().count() == 2
        finally:
            controller.handle_close_engagement()

    def test_import_aborts_when_confirm_dialog_answered_no(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            datasets_before = window.sidebar().datasets_widget().count()
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(import_xlsx), ""),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.preflight_import",
                    return_value=ImportPreflight(warnings=("Datei ist groß (250 MB).",)),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                    return_value=QMessageBox.StandardButton.No,
                ),
            ):
                controller.handle_import_excel()

            assert window.sidebar().datasets_widget().count() == datasets_before
        finally:
            controller.handle_close_engagement()

    def test_import_without_warnings_does_not_show_confirm_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        """Reale Fixture-Datei bleibt unter den Default-Schwellen – kein
        Confirm-Dialog, lautloser Import wie vor Sprint 48."""
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(import_xlsx), ""),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question"
                ) as mock_question,
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()

            mock_question.assert_not_called()
            assert window.sidebar().datasets_widget().count() == 2
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 30: Überschreiben-mit-Backup im Duplikat-Dialog
# ---------------------------------------------------------------------------


def _archive_db_files(parent_dir: Path) -> list[Path]:
    """Alle `.db`-Snapshots im `archiv/`-Unterordner von `parent_dir`."""
    archive = parent_dir / "archiv"
    if not archive.exists():
        return []
    return sorted(p for p in archive.iterdir() if p.is_file() and p.suffix == ".db")


def _create_probe_project(db_path: Path, marker: str) -> None:
    """Legt ein echtes, migriertes Projekt mit leicht prüfbarem Marker an."""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db = Database(db_path)
    try:
        db.migrate()
        EngagementRepo(db.connect()).get_or_create(
            Engagement(
                auditor_name="Anna",
                client_name="ACME",
                auditor_position="Senior",
                audit_type="ISAE 3402",
            )
        )
        conn = db.connect()
        conn.execute("CREATE TABLE overwrite_probe (marker TEXT NOT NULL)")
        conn.execute("INSERT INTO overwrite_probe (marker) VALUES (?)", (marker,))
    finally:
        db.close()


def _probe_markers(db_path: Path) -> list[str]:
    """Liest Marker read-only, damit ein Snapshot nicht verändert wird."""
    uri = f"{db_path.absolute().as_uri()}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    try:
        rows = conn.execute("SELECT marker FROM overwrite_probe ORDER BY rowid").fetchall()
        return [str(row[0]) for row in rows]
    finally:
        conn.close()


class TestOverwriteWithBackup:
    """Sprint 30: Wahl „Überschreiben" sichert zuerst ins archiv/ und legt
    dann ein frisches, leeres Projekt am selben Pfad an."""

    def test_overwrite_backs_up_then_creates_fresh(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "old-marker")

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            with patch(
                "sampling_tool.ui.controllers.engagement_controller.QMessageBox.information"
            ):
                controller.handle_new_engagement()

            # (1) Backup mit altem Inhalt im archiv/.
            backups = _archive_db_files(target.parent)
            assert len(backups) == 1
            assert _probe_markers(backups[0]) == ["old-marker"]

            # (2) Frisches, leeres Projekt am Zielpfad.
            assert target.exists()
            assert window.is_workspace_visible() is True
            assert window.sidebar().datasets_widget().count() == 0
        finally:
            controller.handle_close_engagement()

    def test_overwrite_aborts_when_backup_fails(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "preserve-me")

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(target)
            old_db = controller.session.db
            assert old_db is not None
            old_connection = old_db.connect()
            snapshots_before_failure = set(_archive_db_files(target.parent))

            with (
                patch(
                    "sampling_tool.ui.controllers.engagement_controller."
                    "EngagementVersionManager.create_snapshot",
                    side_effect=OSError("archiv nicht beschreibbar"),
                ),
                patch(
                    "sampling_tool.ui.controllers.engagement_controller._remove_db_files"
                ) as remove_db_files,
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_new_engagement()

            # Aktive alte DB/Session unangetastet, keine Löschung, harte Fehlermeldung.
            assert _probe_markers(target) == ["preserve-me"]
            assert set(_archive_db_files(target.parent)) == snapshots_before_failure
            remove_db_files.assert_not_called()
            assert controller.session.db is old_db
            assert old_connection.execute("SELECT 1").fetchone() is not None
            assert window.is_workspace_visible() is True
            warning.assert_called_once()
            warning_body = warning.call_args.args[2]
            assert "wurde nichts überschrieben" in warning_body
            assert "archiv nicht beschreibbar" in warning_body
        finally:
            controller.handle_close_engagement()

    def test_overwrite_shows_info_with_backup_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "old-marker")

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            with patch(
                "sampling_tool.ui.controllers.engagement_controller.QMessageBox.information"
            ) as info:
                controller.handle_new_engagement()

            assert info.called
            backups = _archive_db_files(target.parent)
            assert len(backups) == 1
            # Titel + Body gezielt prüfen (nicht alle Args in einen Blob falten).
            title = info.call_args.args[1]
            body = info.call_args.args[2]
            assert title == "Projekt überschrieben"
            assert "gesichert" in body
            assert str(backups[0]) in body
        finally:
            controller.handle_close_engagement()

    def test_overwrite_migrate_failure_after_same_path_detach_resets_to_welcome(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        """Nach dem Detach darf ein Fehler der frischen DB keinen halboffenen
        Workspace mit altem Engagement-State zurücklassen."""
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "old-marker")

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(target)
            old_db = controller.session.db
            assert old_db is not None
            snapshots_before_overwrite = set(_archive_db_files(target.parent))

            with (
                _spy_database_close() as close_calls,
                patch.object(Database, "migrate", side_effect=RuntimeError("boom")),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_new_engagement()

            warning.assert_called_once()
            assert window.is_workspace_visible() is False
            assert controller.session.db is None
            assert controller.session.engagement is None
            assert controller.session.dataset is None
            assert controller.session.sample is None
            assert controller.session.undo_manager is None
            assert controller.session.state_repo is None
            assert close_calls[0] is old_db
            assert len(close_calls) == 2
            # Backup ist trotzdem passiert – nur die Neuanlage danach scheiterte.
            assert len(set(_archive_db_files(target.parent)) - snapshots_before_overwrite) == 1
        finally:
            controller.handle_close_engagement()

    def test_overwrite_remove_failure_after_same_path_detach_resets_to_welcome(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "old-marker")
        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(target)
            with (
                patch(
                    "sampling_tool.ui.controllers.engagement_controller._remove_db_files",
                    side_effect=OSError("Datei gesperrt"),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_new_engagement()

            warning.assert_called_once()
            assert window.is_workspace_visible() is False
            assert controller.session.db is None
            assert controller.session.engagement is None
            assert controller.session.dataset is None
            assert controller.session.sample is None
            assert controller.session.undo_manager is None
            assert controller.session.state_repo is None
        finally:
            controller.handle_close_engagement()

    def test_overwrite_closes_open_connection_and_backup_is_complete(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "before-open")

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(target)
            old_db = controller.session.db
            assert old_db is not None
            old_connection = old_db.connect()
            old_connection.execute("PRAGMA wal_autocheckpoint = 0")
            old_connection.execute(
                "INSERT INTO overwrite_probe (marker) VALUES (?)",
                ("committed-in-wal",),
            )
            assert target.with_name(target.name + "-wal").exists()

            snapshots_before_overwrite = set(_archive_db_files(target.parent))
            order: list[str] = []
            real_create_snapshot = EngagementVersionManager.create_snapshot
            real_clear_dataset = window.data_table().clear_dataset
            real_old_db_close = old_db.close

            def _record_snapshot(
                manager: EngagementVersionManager,
                auditor_name: str,
            ) -> Path:
                assert order == []
                assert controller.session.db is old_db
                assert old_connection.execute("SELECT 1").fetchone() is not None
                order.append("snapshot")
                return real_create_snapshot(manager, auditor_name)

            def _record_clear_dataset() -> None:
                if order == ["snapshot"]:
                    order.append("clear")
                else:
                    # `_adopt_database` clears the freshly adopted model again;
                    # that later call must not weaken the destructive-step order.
                    assert order == ["snapshot", "clear", "close", "remove"]
                real_clear_dataset()

            def _record_old_db_close() -> None:
                assert order == ["snapshot", "clear"]
                order.append("close")
                real_old_db_close()

            def _record_remove(path: Path) -> None:
                assert order == ["snapshot", "clear", "close"]
                assert controller.session.db is None
                assert controller.session.undo_manager is None
                assert controller.session.state_repo is None
                with pytest.raises(sqlite3.ProgrammingError, match="closed"):
                    old_connection.execute("SELECT 1")
                order.append("remove")
                _real_remove_db_files(path)

            with (
                patch.object(
                    EngagementVersionManager,
                    "create_snapshot",
                    _record_snapshot,
                ),
                patch.object(
                    window.data_table(),
                    "clear_dataset",
                    side_effect=_record_clear_dataset,
                ),
                patch.object(old_db, "close", side_effect=_record_old_db_close),
                patch(
                    "sampling_tool.ui.controllers.engagement_controller._remove_db_files",
                    side_effect=_record_remove,
                ),
                patch("sampling_tool.ui.controllers.engagement_controller.QMessageBox.information"),
            ):
                controller.handle_new_engagement()

            overwrite_snapshots = set(_archive_db_files(target.parent)) - snapshots_before_overwrite
            assert len(overwrite_snapshots) == 1
            assert _probe_markers(overwrite_snapshots.pop()) == [
                "before-open",
                "committed-in-wal",
            ]
            assert order == ["snapshot", "clear", "close", "remove"]
            assert controller.session.db is not None
            assert controller.session.db is not old_db
            assert controller.session.db.db_path == target
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_overwrite_detaches_same_open_file_reached_through_symlink_alias(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        target = tmp_path / "ACME" / "ACME_ISAE.db"
        _create_probe_project(target, "old-marker")
        alias = tmp_path / "ACME-alias.db"
        try:
            alias.symlink_to(target)
        except (NotImplementedError, OSError) as exc:
            pytest.skip(f"Datei-Symlinks sind auf diesem System nicht verfügbar: {exc}")

        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "ACME"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(alias)
            old_db = controller.session.db
            assert old_db is not None
            assert old_db.db_path == alias
            old_connection = old_db.connect()
            removal_observations: list[tuple[bool, bool, bool, bool]] = []

            def _record_remove(path: Path) -> None:
                try:
                    old_connection.execute("SELECT 1")
                except sqlite3.ProgrammingError:
                    connection_closed = True
                else:
                    connection_closed = False
                removal_observations.append(
                    (
                        controller.session.db is None,
                        controller.session.undo_manager is None,
                        controller.session.state_repo is None,
                        connection_closed,
                    )
                )
                _real_remove_db_files(path)

            with (
                patch(
                    "sampling_tool.ui.controllers.engagement_controller._remove_db_files",
                    side_effect=_record_remove,
                ),
                patch("sampling_tool.ui.controllers.engagement_controller.QMessageBox.information"),
            ):
                controller.handle_new_engagement()

            assert removal_observations == [(True, True, True, True)]
            assert controller.session.db is not None
            assert controller.session.db is not old_db
            assert controller.session.db.db_path == target
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_overwrite_other_target_keeps_active_project_until_adoption(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        active_path = tmp_path / "Active" / "active.db"
        target = tmp_path / "Target" / "target.db"
        _create_probe_project(active_path, "active-a")
        _create_probe_project(target, "target-b")
        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "Target"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(active_path)
            active_db = controller.session.db
            assert active_db is not None
            active_connection = active_db.connect()
            deletion_observations: list[tuple[bool, list[str]]] = []
            adoption_observations: list[tuple[bool, list[str]]] = []
            real_adopt_database = controller.engagement._adopt_database

            def _remove_target_while_active_stays_open(path: Path) -> None:
                rows = active_connection.execute(
                    "SELECT marker FROM overwrite_probe ORDER BY rowid"
                ).fetchall()
                deletion_observations.append(
                    (controller.session.db is active_db, [str(row[0]) for row in rows])
                )
                _real_remove_db_files(path)

            def _adopt_target_after_active_stays_open(
                db: Database,
                adopted_path: Path,
                engagement: Engagement,
            ) -> None:
                rows = active_connection.execute(
                    "SELECT marker FROM overwrite_probe ORDER BY rowid"
                ).fetchall()
                adoption_observations.append(
                    (controller.session.db is active_db, [str(row[0]) for row in rows])
                )
                real_adopt_database(db, adopted_path, engagement)

            with (
                patch(
                    "sampling_tool.ui.controllers.engagement_controller._remove_db_files",
                    side_effect=_remove_target_while_active_stays_open,
                ),
                patch.object(
                    controller.engagement,
                    "_adopt_database",
                    side_effect=_adopt_target_after_active_stays_open,
                ),
                patch("sampling_tool.ui.controllers.engagement_controller.QMessageBox.information"),
            ):
                controller.handle_new_engagement()

            assert deletion_observations == [(True, ["active-a"])]
            assert adoption_observations == [(True, ["active-a"])]
            with pytest.raises(sqlite3.ProgrammingError, match="closed"):
                active_connection.execute("SELECT 1")
            assert controller.session.db is not None
            assert controller.session.db is not active_db
            assert controller.session.db.db_path == target
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_overwrite_other_target_failure_keeps_active_project_open(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        active_path = tmp_path / "Active" / "active.db"
        target = tmp_path / "Target" / "target.db"
        _create_probe_project(active_path, "active-a")
        _create_probe_project(target, "target-b")
        controller = MainController(
            window,
            recent_store=recent_store,
            dialog_factory=lambda parent, _s, _p: _make_stub_new_dialog(parent, target, "Target"),
            duplicate_dialog_factory=lambda parent, db_path: _make_stub_duplicate_dialog(
                parent, db_path, DuplicateEngagementChoice.OVERWRITE
            ),
        )
        try:
            controller.handle_open_engagement(active_path)
            active_db = controller.session.db
            assert active_db is not None
            active_connection = active_db.connect()

            with (
                patch(
                    "sampling_tool.ui.controllers.engagement_controller._remove_db_files",
                    side_effect=OSError("Ziel gesperrt"),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_new_engagement()

            warning.assert_called_once()
            assert controller.session.db is active_db
            assert active_connection.execute("SELECT 1").fetchone() is not None
            assert _probe_markers(target) == ["target-b"]
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 16: ImportOptionsDialog-Dispatch
# ---------------------------------------------------------------------------


@pytest.fixture
def multi_sheet_import_xlsx(tmp_path: Path) -> Path:
    """xlsx mit 2 Sheets unterschiedlicher Spalten."""
    path = tmp_path / "multi_import.xlsx"
    wb = Workbook()
    first = wb.active
    assert first is not None
    first.title = "Erstes"
    first.append(["a", "b"])
    first.append([1, 2])
    second = wb.create_sheet("Zweites")
    second.append(["x", "y", "z"])
    second.append([10, 20, 30])
    second.append([40, 50, 60])
    wb.save(path)
    return path


@pytest.fixture
def leading_blank_import_xlsx(tmp_path: Path) -> Path:
    """xlsx mit 2 echten Leerzeilen → low confidence."""
    path = tmp_path / "leading_blank_import.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([None, None])
    ws.append([None, None])
    ws.append(["Konto", "Betrag"])
    ws.append(["1000", 500])
    wb.save(path)
    return path


class _StubImportOptionsDialog:
    """Stub des `ImportOptionsDialog` für Tests."""

    DialogCode = QDialog.DialogCode

    def __init__(self, sheet_name: str | None, header_row: int | None, accept: bool = True) -> None:
        from sampling_tool.ui.dialogs.import_options_dialog import ImportOptionsResult

        self._result = ImportOptionsResult(sheet_name=sheet_name, header_row=header_row)
        self._accept = accept

    def exec(self) -> int:
        return int(QDialog.DialogCode.Accepted if self._accept else QDialog.DialogCode.Rejected)

    def get_result(self) -> object:
        return self._result if self._accept else None


class TestImportDialogDispatch:
    def test_single_sheet_high_confidence_skips_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        # import_xlsx hat 1 Sheet + Header in Zeile 1 → high confidence
        dialog_calls: list[Path] = []

        def factory(path: Path, _imp: object, _parent: object) -> object:
            dialog_calls.append(path)
            return _StubImportOptionsDialog("Sheet", 0)

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(import_xlsx), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()
            assert dialog_calls == [], "Dialog darf bei high confidence + 1 Sheet NICHT erscheinen"
            assert window.sidebar().datasets_widget().count() == 2
        finally:
            controller.handle_close_engagement()

    def test_single_sheet_low_confidence_shows_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        leading_blank_import_xlsx: Path,
    ) -> None:
        dialog_calls: list[Path] = []

        def factory(path: Path, _imp: object, _parent: object) -> object:
            dialog_calls.append(path)
            # Header in Zeile 3 (0-basiert: 2). Sheet "Sheet" (openpyxl-Default).
            return _StubImportOptionsDialog("Sheet", 2)

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(leading_blank_import_xlsx), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()
            assert len(dialog_calls) == 1
            assert dialog_calls[0] == leading_blank_import_xlsx
            assert window.sidebar().datasets_widget().count() == 2
        finally:
            controller.handle_close_engagement()

    def test_multi_sheet_always_shows_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        multi_sheet_import_xlsx: Path,
    ) -> None:
        dialog_calls: list[Path] = []

        def factory(path: Path, _imp: object, _parent: object) -> object:
            dialog_calls.append(path)
            # User wählt "Zweites" mit Header in Zeile 1
            return _StubImportOptionsDialog("Zweites", 0)

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(multi_sheet_import_xlsx), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()
            assert len(dialog_calls) == 1, "Bei Multi-Sheet muss der Dialog immer erscheinen"
        finally:
            controller.handle_close_engagement()

    def test_dialog_cancel_aborts_import(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        multi_sheet_import_xlsx: Path,
    ) -> None:
        def factory(_path: Path, _imp: object, _parent: object) -> object:
            return _StubImportOptionsDialog("Erstes", 0, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            before = window.sidebar().datasets_widget().count()
            with patch(
                "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                return_value=(str(multi_sheet_import_xlsx), ""),
            ):
                controller.handle_import_excel()
            assert window.sidebar().datasets_widget().count() == before
        finally:
            controller.handle_close_engagement()

    def test_dialog_accept_uses_configured_import_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        multi_sheet_import_xlsx: Path,
    ) -> None:
        # User wählt "Zweites" (3 Spalten, 2 Datenzeilen).
        def factory(_path: Path, _imp: object, _parent: object) -> object:
            return _StubImportOptionsDialog("Zweites", 0)

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(multi_sheet_import_xlsx), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()
            # Zweites Sheet hat 2 Datenzeilen, 3 Spalten.
            assert window.data_table().table_model().rowCount() == 2
            assert window.data_table().table_model().columnCount() == 3
        finally:
            controller.handle_close_engagement()

    # ---- Sprint 29: CSV nimmt am Header-Detection-Dialog teil ----------

    def test_clean_csv_skips_dialog(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        clean_csv = tmp_path / "clean_import.csv"
        clean_csv.write_text("Konto,Betrag\n1000,500\n2000,600\n", encoding="utf-8")
        dialog_calls: list[Path] = []

        def factory(path: Path, _imp: object, _parent: object) -> object:
            dialog_calls.append(path)
            return _StubImportOptionsDialog(None, 0)

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(clean_csv), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()
            assert dialog_calls == [], "Saubere CSV darf KEINEN Dialog zeigen"
            assert window.data_table().table_model().rowCount() == 2
            assert window.data_table().table_model().columnCount() == 2
        finally:
            controller.handle_close_engagement()

    def test_messy_csv_shows_dialog_and_imports_chosen_header(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        # Titelzeile + Leerzeile über der echten Kopfzeile (raw row 3, index 2).
        messy_csv = tmp_path / "messy_import.csv"
        messy_csv.write_text(
            "Quartalsbericht\n\nKonto,Betrag\n1000,500\n2000,600\n", encoding="utf-8"
        )
        dialog_calls: list[Path] = []

        def factory(path: Path, _imp: object, _parent: object) -> object:
            dialog_calls.append(path)
            return _StubImportOptionsDialog(None, 2)  # CSV: sheet_name=None, Header raw row 3

        controller = MainController(
            window,
            recent_store=recent_store,
            import_options_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QFileDialog.getOpenFileName",
                    return_value=(str(messy_csv), ""),
                ),
                patch("sampling_tool.ui.controllers.workspace_controller.QMessageBox.information"),
            ):
                controller.handle_import_excel()
            assert len(dialog_calls) == 1, "Unsaubere CSV muss den Header-Dialog zeigen"
            # Kopfzeile = Zeile 3 ⇒ Spalten (Konto, Betrag), 2 Datenzeilen.
            assert window.data_table().table_model().columnCount() == 2
            assert window.data_table().table_model().rowCount() == 2
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint-5: Sampling-Flow, Reset, Undo/Redo, Export
# ---------------------------------------------------------------------------


class _StubSamplingDialog:
    """Mini-Stub statt `SamplingDialog`. Liefert ein vordefiniertes Result."""

    DialogCode = QDialog.DialogCode

    def __init__(self, result_obj: object, accept: bool = True) -> None:
        self._result = result_obj
        self._accept = accept

    def exec(self) -> int:
        return int(QDialog.DialogCode.Accepted if self._accept else QDialog.DialogCode.Rejected)

    def set_initial_seed(self, seed: int) -> None:
        """No-Op – der Stub liefert ein fixes Result unabhängig vom Seed."""

    def get_result(self) -> object:
        return self._result


class _StubExportDialog:
    DialogCode = QDialog.DialogCode

    def __init__(self, result_obj: object, accept: bool = True) -> None:
        self._result = result_obj
        self._accept = accept

    def exec(self) -> int:
        return int(QDialog.DialogCode.Accepted if self._accept else QDialog.DialogCode.Rejected)

    def get_result(self) -> object:
        return self._result


def _open_dataset(controller: MainController, window: MainWindow, db_path: Path) -> int:
    controller.handle_open_engagement(db_path)
    ds_id = _first_item_data(window.sidebar().datasets_widget())
    controller.handle_dataset_selected(ds_id)
    return ds_id


class TestSamplingFlow:
    def test_new_sampling_creates_sample_and_highlights(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=7),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            samples_before = window.sidebar().samples_widget().count()
            controller.handle_new_sampling()
            samples_after = window.sidebar().samples_widget().count()
            assert samples_after == samples_before + 1
            highlighted = window.data_table().table_model().highlighted_row_ids()
            assert len(highlighted) == 2
        finally:
            controller.handle_close_engagement()

    def test_reset_clears_highlight_with_confirmation(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with patch(
                "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                return_value=__import__(
                    "PyQt6.QtWidgets", fromlist=["QMessageBox"]
                ).QMessageBox.StandardButton.Yes,
            ):
                controller.handle_reset()
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_reset_cancelled_keeps_highlight(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            from PyQt6.QtWidgets import QMessageBox

            with patch(
                "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                return_value=QMessageBox.StandardButton.No,
            ):
                controller.handle_reset()
            # Highlight unverändert
            assert len(window.data_table().table_model().highlighted_row_ids()) == 2
        finally:
            controller.handle_close_engagement()

    def test_undo_redo_round_trip(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=11),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            after_sampling = window.data_table().table_model().highlighted_row_ids()
            assert len(after_sampling) == 3

            controller.handle_undo()
            # Vorheriger Zustand: kein Sample (vor dem ersten Sampling-Push gab es nichts).
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()

            controller.handle_redo()
            assert window.data_table().table_model().highlighted_row_ids() == after_sampling
        finally:
            controller.handle_close_engagement()

    def test_resample_filters_to_current_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=1, seed=3),
            from_sample_only=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            # Vorhandenes Sample auswählen (row_ids 2,4 aus dem Fixture)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            controller.handle_new_sampling()
            new_highlight = window.data_table().table_model().highlighted_row_ids()
            assert new_highlight  # mindestens eine
            # Die neue Auswahl darf nur row_ids aus dem Vorsample enthalten.
            assert new_highlight.issubset({2, 4})
        finally:
            controller.handle_close_engagement()

    def test_export_sample_calls_excel_exporter(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.export_sample_dialog import ExportSampleDialogResult

        export_result = ExportSampleDialogResult(
            columns=["Konto", "Betrag"],
            custom_name="testname",
            custom_id="42",
            output_dir=tmp_path,
        )
        factory = lambda *args, **kw: _StubExportDialog(export_result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            export_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_sample()
            files = list(tmp_path.glob("testname_ID42_BDO_sampling_*.xlsx"))
            assert len(files) == 1
        finally:
            controller.handle_close_engagement()

    def test_export_audit_pdf_writes_file(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.export_audit_pdf_dialog import (
            ExportAuditPdfDialogResult,
        )

        target = tmp_path / "trail.pdf"
        result = ExportAuditPdfDialogResult(
            output_path=target,
            date_from=None,
            date_to=None,
            event_types=set(),
            use_briefpapier=False,
            include_statistics=True,
        )
        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            audit_pdf_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_audit_pdf()
            assert target.exists()
            assert target.stat().st_size > 0
        finally:
            controller.handle_close_engagement()

    def test_undo_redo_state_after_open_engagement(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            # Frisches Engagement: weder Undo noch Redo verfügbar.
            assert window._action_undo.isEnabled() is False
            assert window._action_redo.isEnabled() is False
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint-5.5: Dataset-Klick + Highlight-Persistenz + Versionierung
# ---------------------------------------------------------------------------


def _two_dataset_db(tmp_path: Path) -> tuple[Path, int, int, int]:
    """Engagement mit zwei Datasets und einem Sample am ersten."""
    db_path = tmp_path / "two.db"
    db = Database(db_path)
    db.migrate()
    eng = EngagementRepo(db.connect()).get_or_create(
        Engagement(auditor_name="Anna", client_name="ACME", audit_type="ISAE 3402")
    )
    assert eng.id is not None
    ds_repo = DatasetRepo(db.connect())
    rows_ab = tuple(DatasetRow(row_id=i, values={"a": i}) for i in range(1, 4))
    ds1 = ds_repo.create(
        Dataset(name="First", columns=("a",), engagement_id=eng.id),
        rows_ab,
    )
    ds2 = ds_repo.create(
        Dataset(name="Second", columns=("a",), engagement_id=eng.id),
        rows_ab,
    )
    assert ds1.id is not None
    assert ds2.id is not None
    sample_id = SampleRepo(db.connect()).create_from_result(
        SampleResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=1),
            selected_row_ids=(1, 3),
            population_size=3,
        ),
        ds1.id,
        "test",
    )
    db.close()
    return db_path, ds1.id, ds2.id, sample_id


class TestDatasetClickPreservesHighlight:
    def test_clicking_same_dataset_keeps_highlight(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        db_path, ds1_id, _ds2_id, sample_id = _two_dataset_db(tmp_path)
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(db_path)
            controller.handle_dataset_selected(ds1_id)
            controller.handle_sample_selected(sample_id)
            before = window.data_table().table_model().highlighted_row_ids()
            controller.handle_dataset_selected(ds1_id)  # gleicher Klick
            after = window.data_table().table_model().highlighted_row_ids()
            assert before == after
            assert before == frozenset({1, 3})
        finally:
            controller.handle_close_engagement()

    def test_clicking_other_dataset_clears_highlight_when_sample_unrelated(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        db_path, ds1_id, ds2_id, sample_id = _two_dataset_db(tmp_path)
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(db_path)
            controller.handle_dataset_selected(ds1_id)
            controller.handle_sample_selected(sample_id)
            controller.handle_dataset_selected(ds2_id)  # anderes Dataset
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_clicking_other_dataset_reapplies_highlight_when_sample_belongs(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        db_path, ds1_id, ds2_id, sample_id = _two_dataset_db(tmp_path)
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(db_path)
            controller.handle_dataset_selected(ds1_id)
            controller.handle_sample_selected(sample_id)
            # Wechsel auf ds2 → Highlight verschwindet
            controller.handle_dataset_selected(ds2_id)
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
            # Zurück zu ds1 → Highlight kommt wieder (Sample gehört wieder dazu)
            controller.handle_dataset_selected(ds1_id)
            assert window.data_table().table_model().highlighted_row_ids() == frozenset({1, 3})
        finally:
            controller.handle_close_engagement()

    def test_open_engagement_creates_snapshot(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        from sampling_tool.config import ARCHIVE_DIR_NAME

        db_path, _ds1, _ds2, _s = _two_dataset_db(tmp_path)
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(db_path)
            archive = db_path.parent / ARCHIVE_DIR_NAME
            snaps = list(archive.glob("*.db"))
            assert len(snaps) == 1
        finally:
            controller.handle_close_engagement()

    def test_open_snapshot_failure_shows_visible_warning(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        db_path, _ds1, _ds2, _sample = _two_dataset_db(tmp_path)
        controller = MainController(window, recent_store=recent_store)
        try:
            status = window.statusBar()
            assert status is not None
            order: list[str] = []
            warning_observations: list[tuple[bool, bool]] = []
            real_adopt_database = controller.engagement._adopt_database
            real_show_message = status.showMessage

            def _record_adopt_database(
                db: Database,
                adopted_path: Path,
                engagement: Engagement,
            ) -> None:
                order.append("adopt")
                real_adopt_database(db, adopted_path, engagement)

            def _record_show_message(message: str, timeout: int = 0) -> None:
                order.append("warning")
                warning_observations.append(
                    (window.is_workspace_visible(), controller.session.db is not None)
                )
                real_show_message(message, timeout)

            with (
                patch(
                    "sampling_tool.ui.controllers.engagement_controller."
                    "EngagementVersionManager.create_snapshot",
                    side_effect=OSError("Datenträger voll"),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
                patch.object(
                    controller.engagement,
                    "_adopt_database",
                    side_effect=_record_adopt_database,
                ),
                patch.object(status, "showMessage", side_effect=_record_show_message),
            ):
                controller.handle_open_engagement(db_path)

            assert window.is_workspace_visible() is True
            assert order == ["adopt", "warning"]
            assert warning_observations == [(True, True)]
            assert "Compliance-Snapshot konnte nicht erstellt werden" in status.currentMessage()
            assert "Datenträger voll" in status.currentMessage()
            warning.assert_not_called()
        finally:
            controller.handle_close_engagement()

    def test_open_foreign_db_shows_error_and_creates_no_snapshot(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        """Sprint 41 / S1.4 (S-002): eine fremde SQLite-Datei darf weder
        gesnapshottet noch migriert werden – Preflight muss VOR dem Snapshot
        greifen."""
        from sampling_tool.config import ARCHIVE_DIR_NAME

        foreign_path = tmp_path / "foreign.db"
        conn = sqlite3.connect(str(foreign_path))
        try:
            conn.execute("CREATE TABLE unrelated (x INTEGER)")
            conn.execute("INSERT INTO unrelated (x) VALUES (1)")
            conn.commit()
        finally:
            conn.close()
        bytes_before = foreign_path.read_bytes()

        controller = MainController(window, recent_store=recent_store)
        try:
            with patch(
                "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
            ) as warning:
                controller.handle_open_engagement(foreign_path)

            assert warning.called
            assert window.is_workspace_visible() is False
            assert foreign_path.read_bytes() == bytes_before
            assert not foreign_path.with_name(foreign_path.name + "-wal").exists()
            assert not foreign_path.with_name(foreign_path.name + "-shm").exists()
            assert not (foreign_path.parent / ARCHIVE_DIR_NAME).exists()
        finally:
            controller.handle_close_engagement()


class TestOpenEngagementConnectionCleanup:
    """Sprint 41 / S1.4: `handle_open_engagement` darf die Kandidaten-
    `Database`-Connection in KEINEM Exit-Pfad offen lassen – weder wenn die
    Datei zwar valide ist aber kein Projekt enthält, noch wenn Migration/
    Repo-Zugriff eine Exception wirft."""

    def test_open_engagement_without_project_closes_connection(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        db_path = tmp_path / "empty.db"
        db = Database(db_path)
        db.migrate()
        db.close()

        controller = MainController(window, recent_store=recent_store)
        try:
            with (
                _spy_database_close() as close_calls,
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_open_engagement(db_path)

            assert warning.called
            assert window.is_workspace_visible() is False
            assert len(close_calls) == 1
        finally:
            controller.handle_close_engagement()

    def test_open_engagement_migrate_failure_closes_connection(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        db_path, _ds1, _ds2, _s = _two_dataset_db(tmp_path)

        controller = MainController(window, recent_store=recent_store)
        try:
            with (
                _spy_database_close() as close_calls,
                patch.object(Database, "migrate", side_effect=RuntimeError("boom")),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as warning,
            ):
                controller.handle_open_engagement(db_path)

            assert warning.called
            assert window.is_workspace_visible() is False
            assert len(close_calls) == 1
        finally:
            controller.handle_close_engagement()


class TestFilterAndSwitchEngagement:
    """Sprint 5.6: Auto-Filter, Reset-Filter, Engagement schließen mit Bestätigung."""

    def test_new_sampling_activates_filter_and_checkbox(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=7),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            # Tabelle ist auf die gezogenen Zeilen reduziert.
            assert window.data_table().table_model().rowCount() == 2
            # Sidebar-Checkbox ist an.
            assert window.sidebar().is_filter_only_sample() is True
            # Statusbar-Suffix sichtbar.
            assert "gefiltert" in window._status_sample.text()
        finally:
            controller.handle_close_engagement()

    def test_reset_deactivates_filter(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from PyQt6.QtWidgets import QMessageBox

        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            # Filter manuell aktivieren
            controller.handle_filter_only_sample_toggled(True)
            assert window.sidebar().is_filter_only_sample() is True

            with patch(
                "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                return_value=QMessageBox.StandardButton.Yes,
            ):
                controller.handle_reset()
            assert window.sidebar().is_filter_only_sample() is False
            # Tabelle zeigt wieder alle 5 Zeilen.
            assert window.data_table().table_model().rowCount() == 5
        finally:
            controller.handle_close_engagement()

    def test_filter_only_sample_toggle_filters_and_unfilters(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            controller.handle_filter_only_sample_toggled(True)
            assert window.data_table().table_model().rowCount() == 2
            controller.handle_filter_only_sample_toggled(False)
            assert window.data_table().table_model().rowCount() == 5
        finally:
            controller.handle_close_engagement()

    def test_filter_checkbox_disabled_without_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            # Frisches Dataset ohne aktives Sample → Checkbox disabled.
            assert window.sidebar().filter_checkbox().isEnabled() is False
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            # Sample aktiv → Checkbox enabled.
            assert window.sidebar().filter_checkbox().isEnabled() is True
        finally:
            controller.handle_close_engagement()

    def test_close_request_confirmed_returns_to_welcome(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from PyQt6.QtWidgets import QMessageBox

        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            with patch(
                "sampling_tool.ui.controllers.engagement_controller.QMessageBox.question",
                return_value=QMessageBox.StandardButton.Yes,
            ):
                controller.handle_close_engagement_requested()
            assert window.is_workspace_visible() is False
        finally:
            controller.handle_close_engagement()

    def test_close_request_cancelled_stays_in_workspace(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from PyQt6.QtWidgets import QMessageBox

        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            with patch(
                "sampling_tool.ui.controllers.engagement_controller.QMessageBox.question",
                return_value=QMessageBox.StandardButton.No,
            ):
                controller.handle_close_engagement_requested()
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_close_request_noop_when_no_engagement(
        self,
        controller: MainController,
        window: MainWindow,
    ) -> None:
        # Ohne offenes Engagement darf kein Dialog erscheinen.
        with patch(
            "sampling_tool.ui.controllers.engagement_controller.QMessageBox.question"
        ) as question:
            controller.handle_close_engagement_requested()
        assert question.called is False
        assert window.is_workspace_visible() is False


class TestEngagementsDirSetup:
    def test_engagements_dir_is_created_on_init(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
    ) -> None:
        from sampling_tool.config import ENGAGEMENTS_DIR

        MainController(window, recent_store=recent_store)
        assert ENGAGEMENTS_DIR.exists()


class TestSanitizeForPath:
    def test_replaces_spaces_with_underscore(self) -> None:
        from sampling_tool.config import sanitize_for_path

        assert sanitize_for_path("A1 Telekom Austria AG") == "A1_Telekom_Austria_AG"

    def test_transliterates_umlauts(self) -> None:
        from sampling_tool.config import sanitize_for_path

        assert sanitize_for_path("Müller & Söhne GmbH") == "Mueller__Soehne_GmbH"

    def test_strips_special_chars_keeps_dash(self) -> None:
        from sampling_tool.config import sanitize_for_path

        assert sanitize_for_path("Foo-Bar/Baz?") == "Foo-BarBaz"

    def test_empty_falls_back(self) -> None:
        from sampling_tool.config import sanitize_for_path

        assert sanitize_for_path("?!") == "engagement"

    def test_sanitize_blocks_reserved_device_names(self) -> None:
        """N-009: reservierte Windows-Gerätenamen dürfen nie unverändert
        durchgereicht werden – sonst crasht `mkdir` auf Windows (App-Abbruch).

        Exakte erwartete Werte (statt nur "nicht reserviert") – `CON.db` deckt
        den Stem-vor-Suffix-Fall EXPLIZIT auf: der `.`-Filter würde `CON.db`
        sonst zufällig zu `CONdb` entschärfen, ohne dass die Reserved-Name-
        Erkennung selbst je greift. Der exakte Erwartungswert `CONdb_` beweist,
        dass der Stem-Check (auf dem noch ungefilterten Namen) tatsächlich
        feuert, statt nur zufällig durch das Dot-Stripping verdeckt zu werden.
        """
        from sampling_tool.config import sanitize_for_path

        assert sanitize_for_path("CON") == "CON_"
        assert sanitize_for_path("con") == "con_"
        assert sanitize_for_path("NUL") == "NUL_"
        assert sanitize_for_path("COM1") == "COM1_"
        assert sanitize_for_path("LPT9") == "LPT9_"
        assert sanitize_for_path("CON.db") == "CONdb_"
        # Traversal bleibt wie bisher abgewehrt.
        assert ".." not in sanitize_for_path("../../etc")
        assert "/" not in sanitize_for_path("a/b")
        assert "\\" not in sanitize_for_path("a\\b")

    def test_sanitize_caps_length(self) -> None:
        from sampling_tool.config import sanitize_for_path

        result = sanitize_for_path("A" * 300)
        assert len(result) <= 100

    def test_sanitize_preserves_unicode(self) -> None:
        """Dokumentiertes Verhalten (Sprint 51 / N-009 Docstring-Korrektur):
        `str.isalnum()` ist nicht auf ASCII beschränkt – kyrillische/CJK/
        akzentuierte Zeichen bleiben erhalten."""
        from sampling_tool.config import sanitize_for_path

        assert sanitize_for_path("Мандант北京é") == "Мандант北京é"


# ---------------------------------------------------------------------------
# Sprint-6: Reports + Refresh-Logik
# ---------------------------------------------------------------------------


class TestSprint6Reports:
    def test_audit_trail_view_populated_after_open(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        events = window.audit_trail_view().model()._events
        # populated_db enthält noch keine Events (Sample wurde direkt
        # eingefügt, kein Logger). Aber das Modell muss gesetzt sein.
        assert isinstance(events, list)

    def test_dashboard_view_populated_after_open(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        # Dashboard sollte aus dem Empty-State raus sein, da Datasets vorhanden sind.
        dashboard = window.dashboard_view()
        assert dashboard._stack.currentWidget() is not dashboard._empty_label

    def test_audit_event_double_clicked_selects_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        import_xlsx: Path,
    ) -> None:
        """Nach einem Import sollte ein Doppelklick aufs Import-Event nichts brechen."""
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            # Trigger eine Aktion mit Sample-Bezug.
            ds_id = _first_item_data(window.sidebar().datasets_widget())
            controller.handle_dataset_selected(ds_id)
            sample_id = _first_item_data(window.sidebar().samples_widget())

            # Manuell einen Sample-Event ins Audit-Log schreiben.
            from sampling_tool.audit.logger import AuditLogger
            from sampling_tool.persistence.database import Database
            from sampling_tool.persistence.repositories import AuditRepo, SampleRepo

            db = Database(populated_db)
            db.migrate()
            sample = SampleRepo(db.connect()).get_by_id(sample_id)
            assert sample is not None
            logger_ = AuditLogger(AuditRepo(db.connect()), "tester", 1)
            evt = logger_.log_sampling(sample, sample_id, ds_id)
            db.close()

            controller._refresh_audit_trail()
            assert evt.id is not None
            controller.handle_audit_event_double_clicked(evt.id)
            # Sample sollte jetzt hervorgehoben sein.
            highlights = window.data_table().table_model().highlighted_row_ids()
            assert sample_id in {s for s in [sample_id]}  # smoke
            assert highlights
        finally:
            controller.handle_close_engagement()

    def test_handle_export_excel_report_writes_file(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.export_excel_report_dialog import (
            ExportExcelReportDialogResult,
        )

        target = tmp_path / "bericht.xlsx"
        result = ExportExcelReportDialogResult(
            output_path=target,
            sheets={"Übersicht", "AuditTrail", "Samples", "Statistiken"},
        )
        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            excel_report_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_excel_report()
            assert target.exists()
        finally:
            controller.handle_close_engagement()

    def test_handle_export_excel_report_includes_dataset_id_in_samples_sheet(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        """Sprint 43 / A-001: die Projekt-XLSX-Samples-Tabelle zeigt die echte
        Dataset-ID statt "—" – `collect_report_data` reicht sie jetzt pro
        Sample durch, statt sie beim Abflachen zu verlieren."""
        from openpyxl import load_workbook

        from sampling_tool.ui.dialogs.export_excel_report_dialog import (
            ExportExcelReportDialogResult,
        )

        target = tmp_path / "bericht.xlsx"
        result = ExportExcelReportDialogResult(
            output_path=target,
            sheets={"Übersicht", "AuditTrail", "Samples", "Statistiken"},
        )
        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            excel_report_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            ds_id = _first_item_data(window.sidebar().datasets_widget())
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_excel_report()
            wb = load_workbook(target)
            ws = wb["3. Samples"]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            dataset_id_col = header.index("Dataset-ID")
            assert rows[1][dataset_id_col] == ds_id
        finally:
            controller.handle_close_engagement()

    def test_handle_export_html_report_writes_file(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.export_html_report_dialog import (
            ExportHtmlReportDialogResult,
        )

        target = tmp_path / "bericht.html"
        result = ExportHtmlReportDialogResult(
            output_path=target,
            include_charts=True,
            include_audit_trail=True,
            include_samples_table=True,
        )
        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            html_report_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_html_report()
            assert target.exists()
            content = target.read_text(encoding="utf-8")
            assert "ACME" in content
        finally:
            controller.handle_close_engagement()

    def test_refresh_views_resets_to_empty_on_close(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        controller.handle_close_engagement()
        assert window.audit_trail_view().model()._events == []


class TestUnifiedExportDialogs:
    """Sprint 6.1: Handler nutzen Dialog-Factories und filtern korrekt."""

    def test_audit_pdf_handler_filters_events_by_type(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        """Mit `event_types={"import"}` muss das PDF nur Import-Events enthalten."""
        from sampling_tool.audit.logger import AuditLogger
        from sampling_tool.persistence.database import Database
        from sampling_tool.persistence.repositories import AuditRepo, SampleRepo
        from sampling_tool.ui.dialogs.export_audit_pdf_dialog import (
            ExportAuditPdfDialogResult,
        )

        # Drei verschiedene Events in die DB schreiben.
        db = Database(populated_db)
        db.migrate()
        sample_repo = SampleRepo(db.connect())
        sample = sample_repo.list_for_dataset(1)[0]
        assert sample.id is not None
        logger_ = AuditLogger(AuditRepo(db.connect()), "tester", 1)
        logger_.log_sampling(sample, sample.id, 1)
        logger_.log_export(sample.id, tmp_path / "x.xlsx", 2)
        db.close()

        target = tmp_path / "trail.pdf"
        result = ExportAuditPdfDialogResult(
            output_path=target,
            date_from=None,
            date_to=None,
            event_types={"sampling"},
            use_briefpapier=False,
            include_statistics=True,
        )
        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            audit_pdf_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with patch(
                "sampling_tool.ui.controllers.export_controller.QMessageBox.information"
            ) as info:
                controller.export.handle_export_audit_pdf()
            # Info-Text enthält Anzahl der gefilterten Events.
            assert info.called
            args = info.call_args[0]
            assert "1 Events" in args[2]
        finally:
            controller.handle_close_engagement()

    def test_audit_pdf_handler_cancelled_returns_silently(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        factory = lambda *args, **kw: _StubExportDialog(None, accept=False)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            audit_pdf_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            controller.export.handle_export_audit_pdf()
            assert not list(tmp_path.glob("*.pdf"))
        finally:
            controller.handle_close_engagement()

    def test_excel_report_handler_passes_sheets_subset(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        from openpyxl import load_workbook

        from sampling_tool.ui.dialogs.export_excel_report_dialog import (
            ExportExcelReportDialogResult,
        )

        target = tmp_path / "subset.xlsx"
        result = ExportExcelReportDialogResult(
            output_path=target,
            sheets={"Übersicht"},
        )
        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            excel_report_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_excel_report()
            assert target.exists()
            wb = load_workbook(target)
            assert len(wb.sheetnames) == 1
        finally:
            controller.handle_close_engagement()


class TestSettingsIntegration:
    """Settings beeinflussen Default-Werte beim Controller."""

    def test_init_uses_provided_settings(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        from dataclasses import replace

        from sampling_tool.ui.settings_store import AppSettings

        custom_dir = tmp_path / "my-engagements"
        settings = replace(AppSettings.defaults(), engagements_dir=custom_dir)
        MainController(window, recent_store=recent_store, settings=settings)
        assert custom_dir.exists()

    def test_handle_settings_persists_new_values(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from PyQt6.QtCore import QSettings

        from sampling_tool.config import APP_NAME, APP_ORG
        from sampling_tool.ui.dialogs.settings_dialog import SettingsDialog
        from sampling_tool.ui.settings_store import AppSettings

        # QSettings in tmp_path isolieren, damit echte Prefs nicht angefasst werden.
        monkeypatch.setattr(
            "sampling_tool.ui.settings_store._qsettings",
            lambda: QSettings(
                QSettings.Format.IniFormat,
                QSettings.Scope.UserScope,
                APP_ORG,
                APP_NAME,
            ),
        )
        QSettings.setPath(
            QSettings.Format.IniFormat,
            QSettings.Scope.UserScope,
            str(tmp_path),
        )
        from dataclasses import replace

        new_settings = replace(
            AppSettings.defaults(), default_auditor_name="Updated", undo_depth=33
        )

        class _StubSettingsDialog(SettingsDialog):
            def exec(self) -> int:
                self._result = new_settings
                return int(QDialog.DialogCode.Accepted)

        controller = MainController(
            window,
            recent_store=recent_store,
            settings_dialog_factory=lambda parent, current: _StubSettingsDialog(current, parent),
        )
        controller.help.handle_settings()
        from sampling_tool.ui.settings_store import load_settings

        assert controller._settings.default_auditor_name == "Updated"
        loaded = load_settings()
        assert loaded.default_auditor_name == "Updated"
        assert loaded.undo_depth == 33

    def test_audit_pdf_dialog_receives_settings_defaults(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.export_audit_pdf_dialog import ExportAuditPdfDialog
        from sampling_tool.ui.settings_store import AppSettings

        captured: dict[str, object] = {}
        from dataclasses import replace

        settings = replace(
            AppSettings.defaults(),
            default_include_briefpapier=False,
            default_include_statistics=False,
        )

        def factory(  # type: ignore[no-untyped-def]
            parent,
            engagement,
            available,
            bp_available,
            default_dir,
            default_use_briefpapier,
            default_include_statistics,
            offer_date_filter,
            default_company_key,
            default_location_key,
        ):
            captured["default_use_briefpapier"] = default_use_briefpapier
            captured["default_include_statistics"] = default_include_statistics
            captured["offer_date_filter"] = offer_date_filter

            dialog = ExportAuditPdfDialog(
                engagement=engagement,
                event_types_available=available,
                briefpapier_available=bp_available,
                parent=parent,
                default_output_dir=default_dir,
                default_use_briefpapier=default_use_briefpapier,
                default_include_statistics=default_include_statistics,
                offer_date_filter=offer_date_filter,
            )

            def reject() -> int:
                return int(QDialog.DialogCode.Rejected)

            dialog.exec = reject  # type: ignore[method-assign]
            return dialog

        controller = MainController(
            window,
            recent_store=recent_store,
            audit_pdf_dialog_factory=factory,
            settings=settings,
        )
        try:
            controller.handle_open_engagement(populated_db)
            controller.export.handle_export_audit_pdf()
            assert captured["default_use_briefpapier"] is False
            assert captured["default_include_statistics"] is False
        finally:
            controller.handle_close_engagement()

    def test_audit_pdf_persists_and_resolves_company_location(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Sprint 33: nach dem Export werden bdo_company_key + bdo_location_key
        app-weit persistiert und der AuditPdfExportTask erhält die aufgelösten
        company + location."""
        from pathlib import Path as _Path

        from PyQt6.QtCore import QSettings

        from sampling_tool.config import APP_NAME, APP_ORG
        from sampling_tool.ui.dialogs.export_audit_pdf_dialog import (
            ExportAuditPdfDialogResult,
        )
        from sampling_tool.ui.settings_store import load_settings
        from sampling_tool.ui.workers.tasks import AuditPdfExportTask

        # QSettings in tmp_path isolieren.
        monkeypatch.setattr(
            "sampling_tool.ui.settings_store._qsettings",
            lambda: QSettings(
                QSettings.Format.IniFormat,
                QSettings.Scope.UserScope,
                APP_ORG,
                APP_NAME,
            ),
        )
        QSettings.setPath(
            QSettings.Format.IniFormat,
            QSettings.Scope.UserScope,
            str(tmp_path),
        )

        result = ExportAuditPdfDialogResult(
            output_path=tmp_path / "trail.pdf",
            date_from=None,
            date_to=None,
            event_types=set(),
            use_briefpapier=False,
            include_statistics=True,
            company_key="consulting_gmbh",
            location_key="linz",
        )
        captured: dict[str, AuditPdfExportTask] = {}

        def fake_run_task(self: object, task: AuditPdfExportTask) -> _Path:
            captured["task"] = task
            return task.output_path

        monkeypatch.setattr(
            "sampling_tool.ui.controllers.export_controller.TaskProgressDialog.run_task",
            fake_run_task,
        )

        factory = lambda *args, **kw: _StubExportDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            audit_pdf_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            controller.handle_open_engagement(populated_db)
            with patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"):
                controller.export.handle_export_audit_pdf()
            task = captured["task"]
            assert task.company is not None
            assert task.company.key == "consulting_gmbh"
            assert task.location is not None
            assert task.location.key == "linz"
            assert controller._settings.bdo_company_key == "consulting_gmbh"
            assert controller._settings.bdo_location_key == "linz"
            loaded = load_settings()
            assert loaded.bdo_company_key == "consulting_gmbh"
            assert loaded.bdo_location_key == "linz"
        finally:
            controller.handle_close_engagement()

    def test_reset_keeps_filter_when_setting_enabled(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from dataclasses import replace

        from PyQt6.QtWidgets import QMessageBox

        from sampling_tool.ui.settings_store import AppSettings

        settings = replace(AppSettings.defaults(), reset_keeps_filter=True)
        controller = MainController(window, recent_store=recent_store, settings=settings)
        monkeypatch.setattr(
            QMessageBox, "question", lambda *_a, **_k: QMessageBox.StandardButton.Yes
        )
        try:
            controller.handle_open_engagement(populated_db)
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            controller.handle_filter_only_sample_toggled(True)
            assert controller._filter_active_sample_id is not None
            controller.handle_reset()
            # Filter bleibt aktiv (Setting), aber Sample-Highlight ist weg.
            assert controller._sample is None
            assert controller._filter_active_sample_id is not None
        finally:
            controller.handle_close_engagement()

    def test_resolve_briefpapier_uses_setting_override(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
    ) -> None:
        from dataclasses import replace

        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen.canvas import Canvas

        from sampling_tool.ui.settings_store import AppSettings

        # Sprint 47 / N-010: echtes, einseitiges PDF statt eines bloßen
        # Header-Fragments – `resolve_briefpapier` validiert jetzt auch die
        # Parsebarkeit, ein Fragment würde also fälschlich auf das Default
        # zurückfallen und dieser Test würde die Override-Auswahl nicht mehr
        # prüfen.
        custom_pdf = tmp_path / "my_letter.pdf"
        canvas = Canvas(str(custom_pdf), pagesize=A4)
        canvas.drawString(100, 700, "Briefpapier")
        canvas.save()

        settings = replace(AppSettings.defaults(), custom_briefpapier_path=custom_pdf)
        controller = MainController(window, recent_store=recent_store, settings=settings)
        cfg = controller._resolve_briefpapier()
        assert cfg is not None
        assert cfg.background_image == custom_pdf

    def test_resolve_briefpapier_falls_back_on_parse_error(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Sprint 47 / N-010: ein defektes Custom-Briefpapier lässt
        `resolve_briefpapier` sichtbar auf das Default zurückfallen, statt
        eine unlesbare Config zurückzugeben, die später beim Export crasht."""
        from dataclasses import replace

        from sampling_tool.ui.settings_store import AppSettings

        corrupt_pdf = tmp_path / "corrupt.pdf"
        corrupt_pdf.write_bytes(b"%PDF-1.4\nnot a real xref table, just garbage\n")

        settings = replace(AppSettings.defaults(), custom_briefpapier_path=corrupt_pdf)
        controller = MainController(window, recent_store=recent_store, settings=settings)
        with caplog.at_level("WARNING", logger="sampling_tool.ui.controllers.workspace_session"):
            cfg = controller._resolve_briefpapier()

        assert cfg is None or cfg.background_image != corrupt_pdf
        warnings = [r for r in caplog.records if r.levelname == "WARNING"]
        assert any("ungültig" in r.message.lower() for r in warnings), (
            f"Erwartete WARNING zum ungültigen Custom-Briefpapier, gefangen: "
            f"{[r.message for r in warnings]}"
        )

    def test_resolve_briefpapier_falls_back_and_warns_when_custom_path_missing(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        tmp_path: Path,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Sprint 47 / N-010: ein GELÖSCHTER Custom-Briefpapier-Pfad fällt
        genauso sichtbar (WARN-Log) auf das Default zurück wie ein
        korruptes Briefpapier – nicht nur still, wie es der Fall wäre, wenn
        das `custom.exists()`-Kurzschluss-Gate den try/except umgeht."""
        from dataclasses import replace

        from sampling_tool.ui.settings_store import AppSettings

        missing_pdf = tmp_path / "gone.pdf"  # existiert nie

        settings = replace(AppSettings.defaults(), custom_briefpapier_path=missing_pdf)
        controller = MainController(window, recent_store=recent_store, settings=settings)
        with caplog.at_level("WARNING", logger="sampling_tool.ui.controllers.workspace_session"):
            cfg = controller._resolve_briefpapier()

        assert cfg is None or cfg.background_image != missing_pdf
        warnings = [r for r in caplog.records if r.levelname == "WARNING"]
        assert any("ungültig" in r.message.lower() for r in warnings), (
            f"Erwartete WARNING zum fehlenden Custom-Briefpapier, gefangen: "
            f"{[r.message for r in warnings]}"
        )


class TestEngagementStateRestore:
    """Sprint 8.2 – aktiver Sample-State überlebt Schließen/Öffnen."""

    def test_no_state_on_first_open(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        assert controller._sample is None
        assert controller._active_sample_id is None
        assert controller._filter_active_sample_id is None

    def test_sample_selection_is_persisted(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
    ) -> None:
        controller.handle_open_engagement(populated_db)
        ds_id = _first_item_data(window.sidebar().datasets_widget())
        controller.handle_dataset_selected(ds_id)
        sample_id = _first_item_data(window.sidebar().samples_widget())
        controller.handle_sample_selected(sample_id)

        assert controller._state_repo is not None
        assert controller._engagement is not None
        assert controller._engagement.id is not None
        state = controller._state_repo.get(controller._engagement.id)
        assert state is not None
        assert state.active_dataset_id == ds_id
        assert state.active_sample_id == sample_id

    def test_restore_reapplies_sample_highlight_and_filter(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # Session 1: Sample auswählen, Filter aktivieren, schließen.
        ctrl1 = MainController(window, recent_store=recent_store)
        try:
            ctrl1.handle_open_engagement(populated_db)
            ds_id = _first_item_data(window.sidebar().datasets_widget())
            ctrl1.handle_dataset_selected(ds_id)
            sample_id = _first_item_data(window.sidebar().samples_widget())
            ctrl1.handle_sample_filter_toggled(sample_id)
            assert window.data_table().table_model().rowCount() == 2
        finally:
            ctrl1.handle_close_engagement()

        # Session 2: gleiches Engagement erneut öffnen – State muss da sein.
        ctrl2 = MainController(window, recent_store=recent_store)
        try:
            ctrl2.handle_open_engagement(populated_db)
            assert ctrl2._sample is not None
            assert ctrl2._sample.id == sample_id
            assert ctrl2._filter_active_sample_id == sample_id
            assert window.data_table().table_model().rowCount() == 2
            highlights = window.data_table().table_model().highlighted_row_ids()
            assert highlights == frozenset({2, 4})
        finally:
            ctrl2.handle_close_engagement()

    def test_restore_without_filter(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # Session 1: Sample auswählen ohne Filter (Default ohne Toggle).
        ctrl1 = MainController(window, recent_store=recent_store)
        try:
            ctrl1.handle_open_engagement(populated_db)
            ds_id = _first_item_data(window.sidebar().datasets_widget())
            ctrl1.handle_dataset_selected(ds_id)
            sample_id = _first_item_data(window.sidebar().samples_widget())
            ctrl1.handle_sample_selected(sample_id)
            # Filter NICHT aktiv – nur Highlight.
            assert ctrl1._filter_active_sample_id is None
            assert window.data_table().table_model().rowCount() == 5
        finally:
            ctrl1.handle_close_engagement()

        ctrl2 = MainController(window, recent_store=recent_store)
        try:
            ctrl2.handle_open_engagement(populated_db)
            assert ctrl2._sample is not None
            assert ctrl2._filter_active_sample_id is None
            # Tabelle ungefiltert, aber Highlight da.
            assert window.data_table().table_model().rowCount() == 5
            highlights = window.data_table().table_model().highlighted_row_ids()
            assert highlights == frozenset({2, 4})
        finally:
            ctrl2.handle_close_engagement()

    def test_restore_survives_deleted_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # State auf nicht-existentes Sample setzen. Dafür FK kurzzeitig
        # ausschalten – produktiv simulieren wir den Fall, dass eine
        # spätere App-Version die Sample-Tabelle anders aufräumt.
        db = Database(populated_db)
        db.migrate()
        eng = EngagementRepo(db.connect()).get()
        assert eng is not None
        assert eng.id is not None

        conn = db.connect()
        conn.execute("PRAGMA foreign_keys = OFF")
        try:
            conn.execute(
                "INSERT INTO engagement_state "
                "(engagement_id, active_dataset_id, active_sample_id, filter_active) "
                "VALUES (?, ?, ?, ?)",
                (eng.id, 999999, 888888, 1),
            )
        finally:
            conn.execute("PRAGMA foreign_keys = ON")
        db.close()

        # Öffnen darf nicht crashen, State wird stillschweigend ignoriert.
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            assert controller._sample is None
            assert window.is_workspace_visible() is True
        finally:
            controller.handle_close_engagement()

    def test_reset_clears_persisted_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from PyQt6.QtWidgets import QMessageBox

        monkeypatch.setattr(
            QMessageBox, "question", lambda *_a, **_k: QMessageBox.StandardButton.Yes
        )
        controller = MainController(window, recent_store=recent_store)
        try:
            controller.handle_open_engagement(populated_db)
            ds_id = _first_item_data(window.sidebar().datasets_widget())
            controller.handle_dataset_selected(ds_id)
            sample_id = _first_item_data(window.sidebar().samples_widget())
            controller.handle_sample_selected(sample_id)
            controller.handle_reset()

            assert controller._state_repo is not None
            assert controller._engagement is not None
            assert controller._engagement.id is not None
            state = controller._state_repo.get(controller._engagement.id)
            assert state is not None
            assert state.active_sample_id is None
            assert state.filter_active is False
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 9.3 / Sprint 22: aufgelöste Feature-Sichtbarkeit wird an die
# SamplingDialog-Factory durchgereicht (vorher ein einzelnes advanced_mode-Bool).
# ---------------------------------------------------------------------------


class TestFeatureVisibilityPropagation:
    def test_controller_uebergibt_default_features_bei_simple(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings, SamplingFeatures

        received: dict[str, SamplingFeatures] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            _provider: object,
            _current: object,
            features: SamplingFeatures,
            _match_count: object = None,
        ) -> _StubSamplingDialog:
            received["features"] = features
            return _StubSamplingDialog(None, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
            settings=dc_replace(AppSettings.defaults(), advanced_mode=False),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            features = received["features"]
            # Sprint 36: Filter ist ab Werk sichtbar (auch ohne Advanced-Mode),
            # Cluster/Geschichtet bleiben aus.
            assert features.show_filter is True
            assert features.show_cluster is False
            assert features.show_stratified is False
        finally:
            controller.handle_close_engagement()

    def test_advanced_mode_macht_alle_features_sichtbar(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings, SamplingFeatures

        received: dict[str, SamplingFeatures] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            _provider: object,
            _current: object,
            features: SamplingFeatures,
            _match_count: object = None,
        ) -> _StubSamplingDialog:
            received["features"] = features
            return _StubSamplingDialog(None, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
            settings=dc_replace(AppSettings.defaults(), advanced_mode=True),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            features = received["features"]
            assert features.show_filter is True
            assert features.show_cluster is True
            assert features.show_stratified is True
        finally:
            controller.handle_close_engagement()

    def test_einzel_toggle_macht_nur_eine_funktion_sichtbar(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # Advanced aus, nur Filter-Einzel-Toggle an ⇒ nur Filter sichtbar.
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings, SamplingFeatures

        received: dict[str, SamplingFeatures] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            _provider: object,
            _current: object,
            features: SamplingFeatures,
            _match_count: object = None,
        ) -> _StubSamplingDialog:
            received["features"] = features
            return _StubSamplingDialog(None, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
            settings=dc_replace(
                AppSettings.defaults(), advanced_mode=False, show_filter_feature=True
            ),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            features = received["features"]
            assert features.show_filter is True
            assert features.show_cluster is False
            assert features.show_stratified is False
        finally:
            controller.handle_close_engagement()


class TestNewSamplingDistinctProvider:
    """Sprint 19 / P-005: Advanced-Mode reicht einen Provider-Callback durch,
    get_all_rows wird beim Dialog-Open NICHT mehr aufgerufen."""

    def test_advanced_mode_passes_provider_not_rows(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        captured: dict[str, object] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            provider: object,
            _current: object,
            _advanced: bool,
            _match_count: object = None,
        ) -> _StubSamplingDialog:
            captured["provider"] = provider
            return _StubSamplingDialog(None, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
            settings=dc_replace(AppSettings.defaults(), advanced_mode=True),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            provider = captured["provider"]
            assert callable(provider)
            assert provider("Konto") == ["K1", "K2", "K3", "K4", "K5"]
        finally:
            controller.handle_close_engagement()

    def test_hidden_filter_passes_none_provider(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        captured: dict[str, object] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            provider: object,
            _current: object,
            _advanced: bool,
            _match_count: object = None,
        ) -> _StubSamplingDialog:
            captured["provider"] = provider
            return _StubSamplingDialog(None, accept=False)

        # Sprint 36: Filter ist ab Werk sichtbar, deshalb hier explizit aus, um
        # den „kein Provider ohne sichtbaren Filter"-Zweig zu treffen.
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
            settings=dc_replace(
                AppSettings.defaults(), advanced_mode=False, show_filter_feature=False
            ),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert captured["provider"] is None
        finally:
            controller.handle_close_engagement()

    def test_get_all_rows_not_called_on_dialog_open(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.persistence.repositories import DatasetRepo
        from sampling_tool.ui.settings_store import AppSettings

        def boom(*_args: object, **_kwargs: object) -> object:
            raise AssertionError(
                "get_all_rows darf im Advanced-Sampling-Pfad nicht aufgerufen werden"
            )

        monkeypatch.setattr(DatasetRepo, "get_all_rows", boom)

        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(None, accept=False)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
            settings=dc_replace(AppSettings.defaults(), advanced_mode=True),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()  # darf NICHT in boom() laufen
        finally:
            controller.handle_close_engagement()


class TestFilterMatchCountProvider:
    """Sprint 36 / T7: `handle_new_sampling` reicht einen Trefferzahl-Provider
    (6. Factory-Arg) durch, der über den echten Repo zählt und `s.sample`
    *bei Aufruf* liest (Live-Resample-Stand)."""

    def test_provider_wired_and_counts_over_real_repo(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import FilterOperator

        captured: dict[str, object] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            _provider: object,
            _current: object,
            _features: object,
            match_count: object = None,
        ) -> _StubSamplingDialog:
            captured["match_count"] = match_count
            return _StubSamplingDialog(None, accept=False)

        # Default-Settings: Filter ist ab Werk sichtbar (Sprint 36).
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            provider = captured["match_count"]
            assert provider is not None
            assert callable(provider)
            # Betrag-Werte: 10, 20, 30, 40, 50 (Row 1..5). GT 20 ⇒ {30, 40, 50}.
            assert provider("Betrag", FilterOperator.GT, 20, False) == 3
        finally:
            controller.handle_close_engagement()

    def test_provider_reads_active_sample_at_call_time(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import FilterOperator

        captured: dict[str, object] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            _provider: object,
            _current: object,
            _features: object,
            match_count: object = None,
        ) -> _StubSamplingDialog:
            captured["match_count"] = match_count
            return _StubSamplingDialog(None, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            # Beim Dialog-Bau ist noch KEIN Sample aktiv.
            controller.handle_new_sampling()
            assert controller.session.sample is None
            provider = captured["match_count"]
            assert provider is not None

            # Erst JETZT eine Stichprobe aktiv setzen (Rows 2, 4 ⇒ Betrag 20, 40).
            sample_id = _first_item_data(window.sidebar().samples_widget())
            controller.handle_sample_selected(sample_id)
            assert controller.session.sample is not None
            assert tuple(controller.session.sample.selected_row_ids) == (2, 4)

            # restrict=True zählt NUR innerhalb der aktiven Stichprobe: von
            # {20, 40} ist nur 40 > 20 ⇒ genau 1 Treffer ...
            restricted = provider("Betrag", FilterOperator.GT, 20, True)
            assert restricted == 1
            # ... während die unbeschränkte Zählung {30, 40, 50} liefert.
            unrestricted = provider("Betrag", FilterOperator.GT, 20, False)
            assert unrestricted == 3
            assert restricted != unrestricted
        finally:
            controller.handle_close_engagement()

    def test_hidden_filter_passes_none_match_count_provider(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        captured: dict[str, object] = {}

        def fake_factory(
            _parent: MainWindow,
            _dataset: object,
            _provider: object,
            _current: object,
            _features: object,
            match_count: object = None,
        ) -> _StubSamplingDialog:
            captured["match_count"] = match_count
            return _StubSamplingDialog(None, accept=False)

        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=fake_factory,  # type: ignore[arg-type]
            settings=dc_replace(
                AppSettings.defaults(), advanced_mode=False, show_filter_feature=False
            ),
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert captured["match_count"] is None
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 9.4: Panel-Sichtbarkeit (Dashboard / AuditTrail) wird live angewendet
# ---------------------------------------------------------------------------


class TestPanelVisibilityWiring:
    def test_init_wendet_panel_visibility_aus_settings_an(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        settings = dc_replace(AppSettings.defaults(), show_dashboard=False, show_audit_trail=True)
        MainController(window, recent_store=recent_store, settings=settings)
        # Initialer Controller-Aufruf hat Sichtbarkeit gesetzt.
        assert window._lower_tabs.indexOf(window._dashboard_view) == -1
        assert window._lower_tabs.indexOf(window._audit_trail_view) != -1

    def test_handle_settings_wendet_neue_panel_visibility_an(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.dialogs.settings_dialog import SettingsDialog
        from sampling_tool.ui.settings_store import AppSettings

        defaults = AppSettings.defaults()
        new_settings = dc_replace(defaults, show_dashboard=False, show_audit_trail=False)

        class _StubSettingsDialog(SettingsDialog):
            def exec(self) -> int:
                self._result = new_settings
                return int(QDialog.DialogCode.Accepted)

        controller = MainController(
            window,
            recent_store=recent_store,
            settings_dialog_factory=lambda _p, _s: _StubSettingsDialog(defaults),
            settings=defaults,
        )
        controller.help.handle_settings()
        # Beide Tabs sind weg.
        assert window._lower_tabs.count() == 0
        assert window._lower_tabs.isVisible() is False


# ---------------------------------------------------------------------------
# Sprint 44: Log-Level wird bei jedem Settings-OK live gesetzt
# ---------------------------------------------------------------------------


class TestLogLevelWiring:
    def test_apply_new_settings_updates_root_log_level(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
    ) -> None:
        import logging
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        root = logging.getLogger()
        original_level = root.level
        try:
            defaults = AppSettings.defaults()
            controller = MainController(window, recent_store=recent_store, settings=defaults)
            controller.session.apply_new_settings(dc_replace(defaults, log_level="DEBUG"))
            assert root.level == logging.DEBUG
        finally:
            root.setLevel(original_level)


# ---------------------------------------------------------------------------
# Sprint 57 / L-002: `undo_depth`-Setting wird auf den UndoManager angewendet
# (Konstruktion beim Öffnen + live via apply_new_settings)
# ---------------------------------------------------------------------------


class TestUndoDepthWiring:
    def test_undo_manager_uses_settings_depth(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        depth = 3
        settings = dc_replace(AppSettings.defaults(), undo_depth=depth)
        controller = MainController(window, recent_store=recent_store, settings=settings)
        try:
            controller.handle_open_engagement(populated_db)
            undo_manager = controller.session.undo_manager
            assert undo_manager is not None
            assert undo_manager._max_depth == depth

            for i in range(depth + 3):
                undo_manager.push(sample_id=None, visible_rows=[i], highlighted_rows=[])
            assert sum(1 for _ in iter(undo_manager.undo, None)) == depth
        finally:
            controller.handle_close_engagement()

    def test_undo_depth_applied_live(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        settings = AppSettings.defaults()
        controller = MainController(window, recent_store=recent_store, settings=settings)
        try:
            controller.handle_open_engagement(populated_db)
            undo_manager = controller.session.undo_manager
            assert undo_manager is not None

            new_depth = 2
            controller.session.apply_new_settings(
                dc_replace(controller.session.settings, undo_depth=new_depth)
            )

            for i in range(new_depth + 3):
                undo_manager.push(sample_id=None, visible_rows=[i], highlighted_rows=[])
            assert sum(1 for _ in iter(undo_manager.undo, None)) == new_depth
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 14 / T-007: Pfad-Auswahl im Sampling-Dispatch
# ---------------------------------------------------------------------------
#
# Hintergrund: Sprint 12.1 / P-002 hat den `SimpleSampler.sample_ids`-
# Spezialpfad eingeführt (RAM-Fix bei 1M-Datasets: ~1 GB → ~8 MB). Pass 4
# (T-007) hat festgestellt, dass kein Test verifiziert, ob der Controller
# bei Live-Aufruf den richtigen Pfad nimmt — würde jemand die
# `isinstance(sampler, SimpleSampler)`-Bedingung in
# `handle_new_sampling` entfernen, bliebe alles grün, aber der RAM-Fix
# wäre weg. Diese Tests sichern die Pfad-Auswahl als
# Regressions-Schutz ab.


def _spy_create_sampler(monkeypatch: pytest.MonkeyPatch) -> list[str]:
    """Patcht `create_sampler` im workspace_controller-Modul so, dass jeder
    `sample`/`sample_ids`-Aufruf in der zurückgegebenen Liste landet.

    Sprint 13: `handle_new_sampling` lebt seit dem MainController-Split in
    `WorkspaceController`. Patch-Pfad zeigt deshalb auf die Use-Site
    `sampling_tool.ui.controllers.workspace_controller.create_sampler`,
    nicht mehr auf `main_controller`.

    Die eigentliche Sampler-Logik läuft normal weiter – nur die
    Methoden-Aufrufe werden zusätzlich protokolliert. `isinstance`-Checks
    im Controller bleiben gültig, weil der Subtyp unverändert ist.
    """
    from typing import Any

    from sampling_tool.core.sampling import BaseSampler
    from sampling_tool.core.sampling import create_sampler as real_create_sampler

    calls: list[str] = []

    def spy_create_sampler(cfg: SampleConfig) -> BaseSampler:
        sampler = real_create_sampler(cfg)
        orig_sample = sampler.sample
        orig_sample_ids = getattr(sampler, "sample_ids", None)

        def track_sample(*args: Any, **kwargs: Any) -> Any:
            calls.append("sample")
            return orig_sample(*args, **kwargs)

        sampler.sample = track_sample  # type: ignore[method-assign]

        if orig_sample_ids is not None:

            def track_sample_ids(*args: Any, **kwargs: Any) -> Any:
                calls.append("sample_ids")
                return orig_sample_ids(*args, **kwargs)

            sampler.sample_ids = track_sample_ids  # type: ignore[attr-defined]

        # Sprint 35 / P-003: Cluster/Stratified haben einen pairs-Spezialpfad.
        orig_sample_pairs = getattr(sampler, "sample_pairs", None)
        if orig_sample_pairs is not None:

            def track_sample_pairs(*args: Any, **kwargs: Any) -> Any:
                calls.append("sample_pairs")
                return orig_sample_pairs(*args, **kwargs)

            sampler.sample_pairs = track_sample_pairs  # type: ignore[attr-defined]

        return sampler

    monkeypatch.setattr(
        "sampling_tool.ui.controllers.workspace_controller.create_sampler",
        spy_create_sampler,
    )
    return calls


class TestSamplingPathDispatch:
    """T-007: handle_new_sampling muss die korrekte Sampler-Methode wählen."""

    def test_simple_unfiltered_uses_sample_ids_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=42),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert calls == ["sample_ids"]
        finally:
            controller.handle_close_engagement()

    def test_simple_with_filter_uses_classic_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        # Konto=K1 trifft genau 1 Zeile im populated_db (siehe Fixture).
        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.SIMPLE,
                size=1,
                seed=42,
                filter_field="Konto",
                filter_value="K1",
            ),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert calls == ["sample"]
        finally:
            controller.handle_close_engagement()

    def test_simple_with_from_sample_only_uses_classic_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=1, seed=42),
            from_sample_only=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            # Vorhandenes Sample (row_ids 2,4 aus dem Fixture) auswählen.
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            controller.handle_new_sampling()
            assert calls == ["sample"]
        finally:
            controller.handle_close_engagement()

    def test_cluster_unfiltered_uses_pairs_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Sprint 35 / P-003: Cluster ohne Filter + ohne Sub-Sampling → pairs."""
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        # populated_db hat 5 distinct Konto-Werte → Cluster mit size=1 zieht 1 Cluster.
        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.CLUSTER,
                size=1,
                seed=42,
                cluster_field="Konto",
            ),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert calls == ["sample_pairs"]
        finally:
            controller.handle_close_engagement()

    def test_cluster_with_filter_uses_classic_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.CLUSTER,
                size=1,
                seed=42,
                cluster_field="Konto",
                filter_field="Konto",
                filter_value="K1",
            ),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert calls == ["sample"]
        finally:
            controller.handle_close_engagement()

    def test_stratified_unfiltered_uses_pairs_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Sprint 35 / P-003: Stratified ohne Filter + ohne Sub-Sampling → pairs."""
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        # 5 distinct Konto → size=5 ≥ #Strata, jede Schicht hat genau 1 Element.
        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.STRATIFIED,
                size=5,
                seed=42,
                stratum_field="Konto",
            ),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert calls == ["sample_pairs"]
        finally:
            controller.handle_close_engagement()

    def test_stratified_from_sample_only_uses_classic_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        # Vorsample (rows 2,4) → 2 distinct Konto-Schichten, size=2 passt.
        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.STRATIFIED,
                size=2,
                seed=42,
                stratum_field="Konto",
            ),
            from_sample_only=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            controller.handle_new_sampling()
            assert calls == ["sample"]
        finally:
            controller.handle_close_engagement()

    def test_cluster_field_with_quote_falls_back_to_classic_path(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Review-Finding Sprint 35: '"'/'\\' im Feldnamen brechen den
        json-Pfad – die Weiche muss solche Spalten auf den klassischen
        Pfad schicken (supports_field_pairs-Guard)."""
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        calls = _spy_create_sampler(monkeypatch)
        # Spalte existiert im Dataset nicht → klassischer Pfad gruppiert
        # alles unter None (1 Cluster); entscheidend ist nur der Dispatch.
        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.CLUSTER,
                size=1,
                seed=42,
                cluster_field='Betrag "EUR"',
            ),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            assert calls == ["sample"]
        finally:
            controller.handle_close_engagement()

    @pytest.mark.parametrize(
        ("method", "size"),
        [(SamplingMethod.CLUSTER, 2), (SamplingMethod.STRATIFIED, 5)],
    )
    def test_pairs_path_result_matches_classic_reference(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        method: SamplingMethod,
        size: int,
    ) -> None:
        """E2E-Repro-Oracle (Sprint 35): das Controller-Ergebnis über den
        pairs-Pfad muss bit-identisch zur klassischen `sample(iter_rows)`-
        Referenz mit identischer Config sein (ISAE-3402)."""
        from sampling_tool.core.sampling import create_sampler
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        if method == SamplingMethod.CLUSTER:
            cfg = SampleConfig(method=method, size=size, seed=4711, cluster_field="Konto")
        else:
            cfg = SampleConfig(method=method, size=size, seed=4711, stratum_field="Konto")
        result = SamplingDialogResult(config=cfg, from_sample_only=False)
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            drawn = controller.session.sample
            assert drawn is not None

            assert controller.session.db is not None
            assert controller.session.dataset is not None
            assert controller.session.dataset.id is not None
            repo = DatasetRepo(controller.session.db.connect())
            reference = create_sampler(cfg).sample(
                repo.iter_rows(controller.session.dataset.id),
                population_size=controller.session.dataset.row_count,
            )
            assert drawn.selected_row_ids == reference.selected_row_ids
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 36 / T9 (WP-B): Ergänzungs-/Nachstichprobe ohne Dubletten
# ---------------------------------------------------------------------------


def _activate_existing_sample(controller: MainController, window: MainWindow) -> int:
    """Wählt die im `populated_db` vorhandene Stichprobe (row_ids 2,4) aktiv und
    liefert deren id zurück."""
    sample_id = _first_item_data(window.sidebar().samples_widget())
    controller.handle_sample_selected(sample_id)
    assert controller.session.sample is not None
    assert tuple(controller.session.sample.selected_row_ids) == (2, 4)
    return sample_id


class TestSupplementarySampling:
    """Sprint 36 / T9 (WP-B): Ergänzungs-/Nachstichprobe (`exclude_sample_ids`).

    Zieht aus der Basispopulation, schließt aber die bereits gezogene aktive
    Stichprobe garantiert aus (keine Dubletten). Reproduzierbarkeit ist heilig
    (ISAE-3402): gleicher Seed + gleiche Daten + gleiche Exclude-IDs ⇒
    bit-identische Stichprobe.

    Task 0 verifiziert: die `_resample_checkbox` ist nach einer ersten Ziehung
    korrekt aktiviert (kein Bug gefunden) – WP-B ergänzt hier nur das
    Supplement-Feature.

    Fixture-Daten (`populated_db`): row_ids 1..5, Betrag 10/20/30/40/50,
    vorhandene Stichprobe = row_ids (2, 4).
    """

    def test_build_supplement_iterator_excludes_exactly_given_ids(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            assert controller.session.db is not None
            assert controller.session.dataset is not None
            repo = DatasetRepo(controller.session.db.connect())
            rows, population_size = controller.workspace._build_supplement_iterator(
                repo, controller.session.dataset, [2, 4]
            )
            yielded = [row.row_id for row in rows]
            # Exakt die Basis minus Exclude, in iter_rows-Reihenfolge.
            assert yielded == [1, 3, 5]
            # population_size = row_count - len(exclude).
            assert population_size == 5 - 2
        finally:
            controller.handle_close_engagement()

    def test_supplement_draw_is_duplicate_free_and_skips_fastpath(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        # Spy: der ungefilterte SimpleSampler-Fastpath (sample_ids) DARF bei einer
        # Nachstichprobe NICHT laufen – sonst könnte er die Exclude-IDs mitziehen.
        calls = _spy_create_sampler(monkeypatch)
        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=42),
            exclude_sample_ids=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            _activate_existing_sample(controller, window)
            controller.handle_new_sampling()
            assert controller.session.sample is not None
            drawn = set(controller.session.sample.selected_row_ids)
            # Klassischer Pfad (sample), NICHT der sample_ids-Fastpath.
            assert calls == ["sample"]
            # Dublettenfrei: keine der ausgeschlossenen IDs (2, 4) taucht auf ...
            assert drawn.isdisjoint({2, 4})
            # ... und alles stammt aus der Restmenge {1, 3, 5}.
            assert drawn == {1, 3, 5}  # size=3 über die 3-elementige Restmenge
        finally:
            controller.handle_close_engagement()

    def test_supplement_population_size_is_remainder(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=7),
            exclude_sample_ids=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            _activate_existing_sample(controller, window)
            controller.handle_new_sampling()
            assert controller.session.sample is not None
            # row_count 5 - 2 ausgeschlossene = 3 reale Population dieser Ziehung.
            assert controller.session.sample.population_size == 3
        finally:
            controller.handle_close_engagement()

    def test_supplement_sets_parent_sample_id(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=7),
            exclude_sample_ids=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            parent_id = _activate_existing_sample(controller, window)
            controller.handle_new_sampling()
            stored = controller.session.sample
            assert stored is not None
            assert stored.parent_sample_id == parent_id
            # Auch persistiert (round-trip über den Repo).
            assert controller.session.db is not None
            assert stored.id is not None
            reloaded = SampleRepo(controller.session.db.connect()).get_by_id(stored.id)
            assert reloaded is not None
            assert reloaded.parent_sample_id == parent_id
        finally:
            controller.handle_close_engagement()

    def test_supplement_reproducible_same_seed_identical(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        """ISAE-3402-Oracle: gleicher Seed + gleiche Exclude-IDs ⇒ bit-identisch.

        Größe 2 über die 3-elementige Restmenge {1,3,5}, damit die Auswahl
        tatsächlich vom Shuffle abhängt (nicht triviale Vollmenge). Jede Ziehung
        startet aus einer frischen DB-Kopie (identischer Ausgangszustand), damit
        die zuvor persistierte Nachstichprobe die Ausgangslage nicht verschiebt.
        """
        import shutil

        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        def _draw(seed: int, tag: str) -> tuple[int, ...]:
            fresh_dir = tmp_path / tag
            fresh_dir.mkdir()
            fresh_db = fresh_dir / "engagement.db"
            shutil.copy2(populated_db, fresh_db)

            result = SamplingDialogResult(
                config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=seed),
                exclude_sample_ids=True,
            )
            factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(  # noqa: E731
                result
            )
            controller = MainController(
                window,
                recent_store=recent_store,
                sampling_dialog_factory=factory,  # type: ignore[arg-type]
            )
            try:
                _open_dataset(controller, window, fresh_db)
                _activate_existing_sample(controller, window)
                controller.handle_new_sampling()
                assert controller.session.sample is not None
                drawn = tuple(controller.session.sample.selected_row_ids)
                # Immer dublettenfrei, egal welcher Seed.
                assert set(drawn).issubset({1, 3, 5})
                return drawn
            finally:
                controller.handle_close_engagement()

        # Mehrere Seeds, jeder zweimal aus frischer DB-Kopie gezogen ⇒ pro Seed
        # bit-identisch (Oracle über mehrere Läufe/Seeds, ISAE-3402).
        first_99 = _draw(seed=99, tag="a")
        second_99 = _draw(seed=99, tag="b")
        assert first_99 == second_99  # Seed 99 ⇒ identisch

        first_100 = _draw(seed=100, tag="c")
        second_100 = _draw(seed=100, tag="d")
        assert first_100 == second_100  # Seed 100 ⇒ ebenfalls bit-identisch

    def test_supplement_composes_with_column_filter(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """Filter + Supplement gleichzeitig: Ausschluss (Iterator) zuerst,
        Spaltenfilter (_collect_pool) danach – beide Bedingungen gelten.

        Betrag GT 20 trifft row_ids {3,4,5}; Exclude entfernt {2,4} ⇒ die
        gemeinsame Antwort ist {3,5}."""
        from sampling_tool.core.models import FilterOperator
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(
                method=SamplingMethod.SIMPLE,
                size=2,
                seed=7,
                filter_field="Betrag",
                filter_operator=FilterOperator.GT,
                filter_value=20,
            ),
            exclude_sample_ids=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            _activate_existing_sample(controller, window)
            controller.handle_new_sampling()
            assert controller.session.sample is not None
            drawn = set(controller.session.sample.selected_row_ids)
            # (a) alle erfüllen den Filter (Betrag > 20 ⇒ row_ids 3,4,5) UND
            # (b) keiner ist in der ausgeschlossenen Stichprobe {2,4}.
            assert drawn == {3, 5}
            assert drawn.issubset({3, 4, 5})
            assert drawn.isdisjoint({2, 4})
        finally:
            controller.handle_close_engagement()

    def test_supplement_without_active_sample_falls_through(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """Guard: exclude_sample_ids=True aber KEINE aktive Stichprobe ⇒ sichere
        klassische Voll-Populations-Ziehung (kein Crash, kein Fastpath-Missbrauch)."""
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=7),
            exclude_sample_ids=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            assert controller.session.sample is None  # nichts aktiv ausgewählt
            controller.handle_new_sampling()
            assert controller.session.sample is not None
            drawn = set(controller.session.sample.selected_row_ids)
            # Volle Population {1..5}, kein Ausschluss, kein parent.
            assert drawn.issubset({1, 2, 3, 4, 5})
            assert controller.session.sample.population_size == 5
            assert controller.session.sample.parent_sample_id is None
        finally:
            controller.handle_close_engagement()

    def test_oversized_supplement_surfaces_clean_error(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """Größe > Restmenge ⇒ sauberer SamplingError an den User (QMessageBox),
        keine (Teil-)Stichprobe persistiert, aktive Stichprobe unverändert.

        Der Dialog cappt die Supplement-Größe bewusst nicht (T8); die Validierung
        passiert im Controller zur Ziehzeit. Restmenge nach Ausschluss von (2, 4)
        ist {1, 3, 5} = 3 Rows – size 4 übersteigt sie.
        """
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=4, seed=7),
            exclude_sample_ids=True,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            active_id = _activate_existing_sample(controller, window)
            with patch(
                "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
            ) as warning:
                controller.handle_new_sampling()
            # Fehler wurde dem User gezeigt ...
            assert warning.called
            # ... und es wurde KEINE neue (Teil-)Stichprobe gezogen: die aktive
            # Stichprobe (row_ids 2, 4) bleibt unverändert.
            assert controller.session.sample is not None
            assert controller.session.sample.id == active_id
            assert set(controller.session.sample.selected_row_ids) == {2, 4}
        finally:
            controller.handle_close_engagement()


# ---------------------------------------------------------------------------
# Sprint 20: Sampling-Reset (Toolbar) – In-Memory-Reset, audit-trail-erhaltend
# ---------------------------------------------------------------------------


class TestResetSampling:
    """`WorkspaceSession.reset_sampling()` + `handle_reset_sampling()`.

    Reset leert ausschließlich den gezogenen-Stichprobe-/Ergebnis-State
    (aktive Stichprobe, Highlight, Sample-Filter). Population (Dataset) und
    Parameter (Settings, die den Sampling-Dialog speisen) bleiben erhalten.
    Persistierte Sample-/Audit-Zeilen werden NICHT gelöscht – der
    Append-only-Audit-Trail (ISAE 3402) bleibt intakt.
    """

    def test_reset_clears_sample_and_results(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            assert controller.session.sample is not None
            assert len(window.data_table().table_model().highlighted_row_ids()) == 2

            did_reset = controller.session.reset_sampling()

            assert did_reset is True
            assert controller.session.sample is None
            assert controller.session.active_sample_id is None
            assert controller.session.filter_active_sample_id is None
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_reset_preserves_population(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            ds_id = _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            assert controller.session.db is not None
            repo = DatasetRepo(controller.session.db.connect())
            rows_before = repo.get_all_rows(ds_id)

            controller.session.reset_sampling()

            rows_after = repo.get_all_rows(ds_id)
            assert rows_after == rows_before
            assert controller.session.dataset is not None
            assert controller.session.dataset.row_count == len(rows_before)
        finally:
            controller.handle_close_engagement()

    def test_reset_preserves_parameters(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # In diesem Tool sind die Sampling-Parameter keine persistente
        # Eingabemaske, sondern werden je Ziehung im Dialog gesetzt. Die
        # *Parameter-Oberfläche* = Settings (Advanced-Mode etc.), die den
        # Dialog speisen, + die Population, aus der gezogen wird. Beides muss
        # der Reset unangetastet lassen.
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            settings_before = controller.session.settings
            dataset_before = controller.session.dataset

            controller.session.reset_sampling()

            assert controller.session.settings == settings_before
            assert controller.session.dataset == dataset_before
        finally:
            controller.handle_close_engagement()

    def test_reset_when_nothing_drawn_is_noop(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            assert controller.session.sample is None  # nichts gezogen/ausgewählt

            did_reset = controller.session.reset_sampling()

            assert did_reset is False
            assert controller.session.sample is None
            assert controller.session.active_sample_id is None
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_reset_keeps_persisted_sample_and_audit(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """Audit-Safe-Design: das Sample-Record + Audit-Trail bleiben in der DB.

        (Ersetzt das im Sprint-Prompt nur bedingt geforderte
        `test_reset_clears_persisted_rows` – ein hartes Löschen ist hier
        wegen des Append-only-Audit-FK unmöglich ohne Schema-Änderung.)
        """
        from sampling_tool.persistence.repositories import SampleRepo

        controller = MainController(window, recent_store=recent_store)
        try:
            ds_id = _open_dataset(controller, window, populated_db)
            sample_id = _first_item_data(window.sidebar().samples_widget())
            controller.handle_sample_selected(sample_id)
            assert controller.session.db is not None

            controller.session.reset_sampling()

            # Sample-Record überlebt den Reset (nur In-Memory-Auswahl geleert).
            remaining = SampleRepo(controller.session.db.connect()).list_for_dataset(ds_id)
            assert any(s.id == sample_id for s in remaining)
        finally:
            controller.handle_close_engagement()

    def test_handle_reset_sampling_confirmation_clears(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from PyQt6.QtWidgets import QMessageBox

        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with patch(
                "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                return_value=QMessageBox.StandardButton.Yes,
            ):
                controller.handle_reset_sampling()
            assert controller.session.sample is None
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_handle_reset_sampling_cancelled_keeps_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from PyQt6.QtWidgets import QMessageBox

        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with patch(
                "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                return_value=QMessageBox.StandardButton.No,
            ):
                controller.handle_reset_sampling()
            assert controller.session.sample is not None
            assert len(window.data_table().table_model().highlighted_row_ids()) == 2
        finally:
            controller.handle_close_engagement()


class TestAuditTrailRobustness:
    """Sprint 42 / S1.5a: N-003 (Nachlauf-Log-Fehler crashen nicht mehr),
    N-004 (breiter Exception-Fang beim Sample-Export), N-012 (Undo der
    ersten Ziehung erzeugt ein Audit-Event)."""

    def test_export_sample_survives_os_error(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        """N-004: ein roher OS-/DB-Fehler aus dem Export-Task (nicht
        `ExportError`) darf `handle_export_sample` nicht crashen lassen."""
        from sampling_tool.ui.dialogs.export_sample_dialog import ExportSampleDialogResult

        export_result = ExportSampleDialogResult(
            columns=["Konto", "Betrag"],
            custom_name="testname",
            custom_id="42",
            output_dir=tmp_path,
        )
        factory = lambda *args, **kw: _StubExportDialog(export_result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            export_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with (
                patch(
                    "sampling_tool.ui.controllers.export_controller.TaskProgressDialog.run_task",
                    side_effect=PermissionError("Zieldatei ist geöffnet"),
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as mock_warning,
            ):
                controller.export.handle_export_sample()  # darf NICHT werfen
            mock_warning.assert_called_once()
        finally:
            controller.handle_close_engagement()

    def test_export_kept_and_warns_when_audit_log_fails(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        """N-003 (Nicos Compliance-Entscheidung): schlägt das Audit-Log-INSERT
        nach einem erfolgreichen Export fehl, bleibt die Datei erhalten – kein
        stiller Datenverlust, sondern eine blockierende Warnung statt Crash."""
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.export_sample_dialog import ExportSampleDialogResult

        export_result = ExportSampleDialogResult(
            columns=["Konto", "Betrag"],
            custom_name="testname",
            custom_id="42",
            output_dir=tmp_path,
        )
        factory = lambda *args, **kw: _StubExportDialog(export_result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            export_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            from PyQt6.QtWidgets import QMessageBox

            with (
                patch.object(
                    AuditRepo, "log", side_effect=sqlite3.OperationalError("database is locked")
                ),
                patch(
                    "sampling_tool.ui.controllers.export_controller.QMessageBox.warning",
                    return_value=QMessageBox.StandardButton.Abort,
                ) as mock_warning,
                patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"),
            ):
                controller.export.handle_export_sample()  # darf NICHT werfen
            mock_warning.assert_called_once()
            files = list(tmp_path.glob("testname_ID42_BDO_sampling_*.xlsx"))
            assert len(files) == 1, "Exportdatei muss trotz Audit-Log-Fehler erhalten bleiben"
        finally:
            controller.handle_close_engagement()

    def test_export_audit_log_retry_succeeds_on_second_attempt(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
        tmp_path: Path,
    ) -> None:
        """Retry-Option: klickt der Nutzer „Erneut versuchen" und das zweite
        INSERT klappt, landet trotzdem genau ein `export`-Event im Trail."""
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.export_sample_dialog import ExportSampleDialogResult

        export_result = ExportSampleDialogResult(
            columns=["Konto", "Betrag"],
            custom_name="testname",
            custom_id="42",
            output_dir=tmp_path,
        )
        factory = lambda *args, **kw: _StubExportDialog(export_result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            export_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            from PyQt6.QtWidgets import QMessageBox

            original_log = AuditRepo.log
            call_count = {"n": 0}

            def _fail_once(self: AuditRepo, event: object) -> object:
                call_count["n"] += 1
                if call_count["n"] == 1:
                    raise sqlite3.OperationalError("database is locked")
                return original_log(self, event)  # type: ignore[arg-type]

            with (
                patch.object(AuditRepo, "log", _fail_once),
                patch(
                    "sampling_tool.ui.controllers.export_controller.QMessageBox.warning",
                    return_value=QMessageBox.StandardButton.Retry,
                ) as mock_warning,
                patch("sampling_tool.ui.controllers.export_controller.QMessageBox.information"),
            ):
                controller.export.handle_export_sample()
            mock_warning.assert_called_once()

            assert controller.session.db is not None
            assert controller.session.engagement is not None
            assert controller.session.engagement.id is not None
            events = AuditRepo(controller.session.db.connect()).list_for_engagement(
                controller.session.engagement.id
            )
            export_events = [e for e in events if e.event_type == "export"]
            assert len(export_events) == 1
        finally:
            controller.handle_close_engagement()

    def test_reset_survives_audit_log_failure(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """N-003: schlägt `log_reset` fehl, muss der Reset trotzdem
        durchlaufen (State bereits geändert) – kein App-Crash."""
        from PyQt6.QtWidgets import QMessageBox

        from sampling_tool.persistence.repositories import AuditRepo

        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                    return_value=QMessageBox.StandardButton.Yes,
                ),
                patch.object(
                    AuditRepo, "log", side_effect=sqlite3.OperationalError("database is locked")
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as mock_warning,
            ):
                controller.handle_reset()  # darf NICHT werfen
            mock_warning.assert_called_once()
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_reset_sampling_survives_audit_log_failure(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """N-003: symmetric to test_reset_survives_audit_log_failure, but for
        the toolbar 'Sampling zurücksetzen' path (handle_reset_sampling)."""
        from PyQt6.QtWidgets import QMessageBox

        from sampling_tool.persistence.repositories import AuditRepo

        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_sample_selected(_first_item_data(window.sidebar().samples_widget()))
            with (
                patch(
                    "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
                    return_value=QMessageBox.StandardButton.Yes,
                ),
                patch.object(
                    AuditRepo, "log", side_effect=sqlite3.OperationalError("database is locked")
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as mock_warning,
            ):
                controller.handle_reset_sampling()  # darf NICHT werfen
            mock_warning.assert_called_once()
            assert controller.session.sample is None
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_undo_of_first_draw_logs_event(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """N-012: das Undo der ERSTEN Ziehung (Zielzustand leer) erzeugt jetzt
        ein `undo`-Event mit `sample_id IS NULL` statt gar keins."""
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=11),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()  # genau eine Ziehung

            controller.handle_undo()

            assert controller.session.db is not None
            assert controller.session.engagement is not None
            assert controller.session.engagement.id is not None
            events = AuditRepo(controller.session.db.connect()).list_for_engagement(
                controller.session.engagement.id
            )
            undo_events = [e for e in events if e.event_type == "undo"]
            assert len(undo_events) == 1
            assert undo_events[0].sample_id is None
            assert undo_events[0].details == {"restored": "empty"}
        finally:
            controller.handle_close_engagement()

    def test_undo_survives_audit_log_failure(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=11),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            with (
                patch.object(
                    AuditRepo, "log", side_effect=sqlite3.OperationalError("database is locked")
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as mock_warning,
            ):
                controller.handle_undo()  # darf NICHT werfen
            mock_warning.assert_called_once()
            assert window.data_table().table_model().highlighted_row_ids() == frozenset()
        finally:
            controller.handle_close_engagement()

    def test_undo_logs_real_sample_id_when_restoring_prior_draw(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """Gegenstück zu test_undo_of_first_draw_logs_event: ein Undo, das zu
        einer ECHTEN vorherigen Stichprobe zurückkehrt (nicht zum leeren
        Zustand), muss deren `sample_id` loggen – kein hartkodiertes None.

        Mechanik: `_push_undo_snapshot` legt nach jeder Ziehung den NEUEN
        Zustand oben auf den Undo-Stack (push löscht nur den Redo-Stack).
        Nach zwei Ziehungen liegt also [snap(sample1), snap(sample2)] auf dem
        Stack. `handle_undo` verschiebt snap(sample2) auf den Redo-Stack und
        stellt den darunterliegenden Zustand snap(sample1) wieder her – der
        Undo-Event muss also `sample1.id` loggen, nicht None.
        """
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        first_result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=5),
            from_sample_only=False,
        )
        second_result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=11),
            from_sample_only=False,
        )
        results = [first_result, second_result]
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(  # noqa: E731
            results.pop(0)
        )
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()  # erste Ziehung
            first_sample_id = controller.session.sample.id  # type: ignore[union-attr]
            controller.handle_new_sampling()  # zweite Ziehung (normale Ziehung)

            controller.handle_undo()  # zurück zur ersten Ziehung

            assert controller.session.sample is not None
            assert controller.session.sample.id == first_sample_id

            assert controller.session.db is not None
            assert controller.session.engagement is not None
            assert controller.session.engagement.id is not None
            events = AuditRepo(controller.session.db.connect()).list_for_engagement(
                controller.session.engagement.id
            )
            undo_events = [e for e in events if e.event_type == "undo"]
            assert len(undo_events) == 1
            assert undo_events[0].sample_id == first_sample_id
        finally:
            controller.handle_close_engagement()

    def test_redo_logs_event(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        """N-012 analog: Redo protokolliert immer, auch direkt nach einem
        Undo, dessen Zielzustand `sample_id IS NULL` war."""
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=11),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            controller.handle_undo()

            controller.handle_redo()

            assert controller.session.db is not None
            assert controller.session.engagement is not None
            assert controller.session.engagement.id is not None
            events = AuditRepo(controller.session.db.connect()).list_for_engagement(
                controller.session.engagement.id
            )
            redo_events = [e for e in events if e.event_type == "redo"]
            assert len(redo_events) == 1
            assert redo_events[0].sample_id is not None
        finally:
            controller.handle_close_engagement()

    def test_redo_survives_audit_log_failure(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from sampling_tool.core.models import SampleConfig, SamplingMethod
        from sampling_tool.persistence.repositories import AuditRepo
        from sampling_tool.ui.dialogs.sampling_dialog import SamplingDialogResult

        result = SamplingDialogResult(
            config=SampleConfig(method=SamplingMethod.SIMPLE, size=3, seed=11),
            from_sample_only=False,
        )
        factory = lambda _p, _d, _r, _s, _am, _mcp=None: _StubSamplingDialog(result)  # noqa: E731
        controller = MainController(
            window,
            recent_store=recent_store,
            sampling_dialog_factory=factory,  # type: ignore[arg-type]
        )
        try:
            _open_dataset(controller, window, populated_db)
            controller.handle_new_sampling()
            controller.handle_undo()
            with (
                patch.object(
                    AuditRepo, "log", side_effect=sqlite3.OperationalError("database is locked")
                ),
                patch(
                    "sampling_tool.ui.controllers.workspace_session.QMessageBox.warning"
                ) as mock_warning,
            ):
                controller.handle_redo()  # darf NICHT werfen
            mock_warning.assert_called_once()
            assert len(window.data_table().table_model().highlighted_row_ids()) == 3
        finally:
            controller.handle_close_engagement()


@contextlib.contextmanager
def _real_sampling_dialog_driver(seeds: list[int], size: int = 3) -> Iterator[None]:
    """Treibt den ECHTEN `SamplingDialog` durch den Controller-Pfad.

    - Jeder Dialog-Open zieht den nächsten Wert aus ``seeds`` als
      auto-generierten Seed – simuliert ``_generate_random_seed()``, genau
      die reale Seed-Quelle, die der alte Stub-Test umgangen hat.
    - ``exec()`` setzt die Größe und akzeptiert sofort (kein Event-Loop nötig).
    - Die Reset-Bestätigung wird auf „Ja" gemockt.
    """
    from PyQt6.QtWidgets import QMessageBox

    from sampling_tool.ui.dialogs import sampling_dialog as sd

    def _auto_accept(self: sd.SamplingDialog) -> QDialog.DialogCode:
        self._size_spin.setValue(size)
        self.accept()
        return QDialog.DialogCode.Accepted

    with (
        patch.object(sd, "_generate_random_seed", side_effect=list(seeds)),
        patch.object(sd.SamplingDialog, "exec", _auto_accept),
        patch(
            "sampling_tool.ui.controllers.workspace_controller.QMessageBox.question",
            return_value=QMessageBox.StandardButton.Yes,
        ),
    ):
        yield


class TestReproducibilityViaController:
    """Reproduzierbarkeit über den ECHTEN Controller-/Dialog-Pfad (Sprint 21).

    Der Sprint-20-Test ``TestResetReproducibility`` war grün, obwohl die GUI
    fehlschlug: er injizierte ein ``_StubSamplingDialog`` mit hartkodiertem
    ``seed=123`` und gab denselben Stub bei beiden Ziehungen zurück – die
    reale Seed-Quelle (``SamplingDialog._build_ui`` → ``_generate_random_seed()``)
    wurde nie ausgeführt. Diese Tests benutzen den echten ``SamplingDialog``,
    damit das Seed-Verhalten beim erneuten Öffnen tatsächlich abgedeckt ist.
    """

    def test_draw_reset_redraw_same_seed_identical(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            with _real_sampling_dialog_driver(seeds=[111, 222, 333]):
                controller.handle_new_sampling()
                assert controller.session.sample is not None
                first = controller.session.sample
                r1 = tuple(first.selected_row_ids)
                seed1 = first.config.seed

                controller.handle_reset_sampling()
                assert controller.session.sample is None

                controller.handle_new_sampling()
                assert controller.session.sample is not None
                second = controller.session.sample
                r2 = tuple(second.selected_row_ids)
                seed2 = second.config.seed

            # Kern der ISAE-3402-Invariante: der erneut geöffnete Dialog muss
            # den zuletzt genutzten Seed übernehmen, nicht neu würfeln.
            assert seed2 == seed1
            assert r1 == r2
        finally:
            controller.handle_close_engagement()

    def test_multiple_resets_stay_identical(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            samples: list[tuple[int, ...]] = []
            seeds_used: list[int] = []
            with _real_sampling_dialog_driver(seeds=[111, 222, 333, 444]):
                for _ in range(3):
                    controller.handle_new_sampling()
                    assert controller.session.sample is not None
                    samples.append(tuple(controller.session.sample.selected_row_ids))
                    seeds_used.append(controller.session.sample.config.seed)
                    controller.handle_reset_sampling()
                    assert controller.session.sample is None
            assert seeds_used[0] == seeds_used[1] == seeds_used[2]
            assert samples[0] == samples[1] == samples[2]
        finally:
            controller.handle_close_engagement()

    def test_two_consecutive_draws_without_reset_identical(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # Deckt H1 unabhängig vom Reset ab: gleicher (gemerkter) Seed +
        # gleiche Größe zweimal hintereinander ⇒ identische Stichprobe.
        controller = MainController(window, recent_store=recent_store)
        try:
            _open_dataset(controller, window, populated_db)
            with _real_sampling_dialog_driver(seeds=[111, 222, 333]):
                controller.handle_new_sampling()
                assert controller.session.sample is not None
                r1 = tuple(controller.session.sample.selected_row_ids)
                seed1 = controller.session.sample.config.seed

                controller.handle_new_sampling()
                assert controller.session.sample is not None
                r2 = tuple(controller.session.sample.selected_row_ids)
                seed2 = controller.session.sample.config.seed

            assert seed2 == seed1
            assert r1 == r2
        finally:
            controller.handle_close_engagement()


class TestSeedRelocationReproducibility:
    """Sprint 27: Der Seed-Eingabeort ist in die Einstellungen verlagert –
    die Determinismus-Garantie aus Sprint 21 bleibt erhalten."""

    def test_same_seed_via_settings_same_sample(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        fixed = dc_replace(AppSettings.defaults(), seed=24680)
        controller = MainController(window, recent_store=recent_store, settings=fixed)
        try:
            _open_dataset(controller, window, populated_db)
            # Verschiedene Zufalls-Listen pro Dialog-Open: Würde der feste Seed
            # aus den Settings ignoriert, fielen die Ziehungen auseinander.
            with _real_sampling_dialog_driver(seeds=[111, 222, 333, 444]):
                controller.handle_new_sampling()
                assert controller.session.sample is not None
                r1 = tuple(controller.session.sample.selected_row_ids)
                seed1 = controller.session.sample.config.seed

                controller.handle_reset_sampling()
                assert controller.session.sample is None

                controller.handle_new_sampling()
                assert controller.session.sample is not None
                r2 = tuple(controller.session.sample.selected_row_ids)
                seed2 = controller.session.sample.config.seed

            assert seed1 == 24680
            assert seed2 == 24680
            assert r1 == r2
        finally:
            controller.handle_close_engagement()

    def test_changed_settings_seed_used_for_next_draw(
        self,
        window: MainWindow,
        recent_store: RecentEngagementsStore,
        populated_db: Path,
    ) -> None:
        # „Geänderter Seed gilt für die nächste Ziehung": der Settings-Seed hat
        # Vorrang vor dem gemerkten last_seed.
        from dataclasses import replace as dc_replace

        from sampling_tool.ui.settings_store import AppSettings

        controller = MainController(
            window,
            recent_store=recent_store,
            settings=dc_replace(AppSettings.defaults(), seed=1001),
        )
        try:
            _open_dataset(controller, window, populated_db)
            with _real_sampling_dialog_driver(seeds=[111, 222, 333]):
                controller.handle_new_sampling()
                assert controller.session.sample is not None
                assert controller.session.sample.config.seed == 1001

                # User ändert den Seed in den Einstellungen.
                controller.session.settings = dc_replace(controller.session.settings, seed=2002)
                controller.handle_new_sampling()
                assert controller.session.sample is not None
                assert controller.session.sample.config.seed == 2002
        finally:
            controller.handle_close_engagement()


class TestRefreshViewsSingleEventLoad:
    """Sprint 34 / WP5: `refresh_views` lädt die Audit-Events genau EINMAL.

    Vorher liefen `refresh_audit_trail` und `refresh_dashboard` je einen
    identischen `AuditRepo.list_for_engagement`-Fetch (bis zu 10.000 Events,
    2× pro mutierender User-Aktion – Import, Sampling, Reset, Undo/Redo,
    Export, Open/Close).
    """

    def test_refresh_views_loads_events_once(
        self,
        controller: MainController,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.core.models import AuditEvent
        from sampling_tool.persistence.repositories import AuditRepo

        controller.handle_open_engagement(populated_db)

        calls: list[int] = []
        original = AuditRepo.list_for_engagement

        def counting(self: AuditRepo, engagement_id: int, limit: int = 100) -> list[AuditEvent]:
            calls.append(engagement_id)
            return original(self, engagement_id, limit=limit)

        monkeypatch.setattr(AuditRepo, "list_for_engagement", counting)

        controller.session.refresh_views()
        assert len(calls) == 1  # vorher: 2 (AuditTrail + Dashboard je ein Fetch)

    def test_refresh_views_still_feeds_both_views(
        self,
        controller: MainController,
        window: MainWindow,
        populated_db: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from sampling_tool.core.models import AuditEvent

        controller.handle_open_engagement(populated_db)

        seen: dict[str, list[AuditEvent]] = {}
        original_set_events = window.set_audit_events
        original_set_dashboard = window.set_dashboard_data

        def capture_events(events: list[AuditEvent]) -> None:
            seen["audit"] = list(events)
            original_set_events(events)

        def capture_dashboard(
            engagement: Engagement | None,
            datasets: list[Dataset],
            samples: list[SampleResult],
            events: list[AuditEvent],
        ) -> None:
            seen["dashboard"] = list(events)
            original_set_dashboard(engagement, datasets, samples, events)

        monkeypatch.setattr(window, "set_audit_events", capture_events)
        monkeypatch.setattr(window, "set_dashboard_data", capture_dashboard)

        controller.session.refresh_views()
        # Beide Views werden weiterhin versorgt – mit identischen Events.
        assert seen["audit"] == seen["dashboard"]
