"""ImportOptionsDialog – Sheet-Auswahl, Preview, Header-Detection."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from PyQt6.QtWidgets import QDialog, QDialogButtonBox
from pytestqt.qtbot import QtBot

from sampling_tool.io.importer import ExcelImporter
from sampling_tool.ui.dialogs.import_options_dialog import (
    ImportOptionsDialog,
    ImportOptionsResult,
)

pytestmark = pytest.mark.ui


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def importer() -> ExcelImporter:
    return ExcelImporter()


@pytest.fixture
def simple_path(tmp_path: Path) -> Path:
    path = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Daten"
    ws.append(["Konto", "Bezeichnung", "Saldo"])
    ws.append(["1000", "Kasse", 500.50])
    ws.append(["2000", "Bank", 1234.0])
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_path(tmp_path: Path) -> Path:
    path = tmp_path / "multi.xlsx"
    wb = Workbook()
    first = wb.active
    assert first is not None
    first.title = "Erstes"
    first.append(["a", "b"])
    first.append([1, 2])
    second = wb.create_sheet("Zweites")
    second.append(["x", "y", "z"])
    second.append([10, 20, 30])
    third = wb.create_sheet("Drittes")
    third.append(["nur_spalte"])
    third.append([42])
    wb.save(path)
    return path


@pytest.fixture
def leading_blank_path(tmp_path: Path) -> Path:
    path = tmp_path / "leading_blank.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([None, None, None])
    ws.append([None, None, None])
    ws.append(["Konto", "Bezeichnung", "Saldo"])
    ws.append([1000, "Kasse", 500.50])
    wb.save(path)
    return path


@pytest.fixture
def ambiguous_path(tmp_path: Path) -> Path:
    path = tmp_path / "ambiguous.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([100, 200, 300])
    ws.append([400, 500, 600])
    wb.save(path)
    return path


def _ok_enabled(dialog: ImportOptionsDialog) -> bool:
    box = dialog.findChild(QDialogButtonBox)
    assert box is not None
    btn = box.button(QDialogButtonBox.StandardButton.Ok)
    assert btn is not None
    return bool(btn.isEnabled())


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestImportOptionsDialog:
    def test_dialog_lists_all_sheets_in_dropdown(
        self, qtbot: QtBot, multi_sheet_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(multi_sheet_path, importer)
        qtbot.addWidget(dialog)
        # UserData = Sheet-Name; Display-Text enthält zusätzlich Dimensionen.
        combo = dialog._sheet_combo
        assert combo is not None
        names = [combo.itemData(i) for i in range(combo.count())]
        assert names == ["Erstes", "Zweites", "Drittes"]

    def test_dialog_loads_preview_on_sheet_change(
        self, qtbot: QtBot, multi_sheet_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(multi_sheet_path, importer)
        qtbot.addWidget(dialog)
        # Initial: erstes Sheet → 2 Spalten
        assert dialog._preview_table.columnCount() == 2
        combo = dialog._sheet_combo
        assert combo is not None
        combo.setCurrentIndex(combo.findData("Zweites"))
        assert dialog._preview_table.columnCount() == 3

    def test_dialog_preselects_detected_header_row_high(
        self, qtbot: QtBot, simple_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(simple_path, importer)
        qtbot.addWidget(dialog)
        # High Confidence: Header in Zeile 1 (UI: 1-basiert, intern 0)
        assert dialog._header_spin.value() == 1
        assert dialog.get_result_header_row() == 0

    def test_dialog_preselects_detected_header_row_low(
        self, qtbot: QtBot, leading_blank_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(leading_blank_path, importer)
        qtbot.addWidget(dialog)
        # Low Confidence: Header in Zeile 3 (UI: 1-basiert)
        assert dialog._header_spin.value() == 3
        assert dialog.get_result_header_row() == 2

    def test_dialog_shows_red_warning_when_ambiguous(
        self, qtbot: QtBot, ambiguous_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(ambiguous_path, importer)
        qtbot.addWidget(dialog)
        # Confidence-Label muss BDO-Rot enthalten oder einen klaren Hinweis-Text.
        text = dialog._confidence_label.text().lower()
        assert "nicht eindeutig" in text or "manuell" in text

    def test_dialog_disables_import_button_when_header_row_invalid(
        self, qtbot: QtBot, simple_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(simple_path, importer)
        qtbot.addWidget(dialog)
        # simple_path hat 3 Zeilen. Header in Zeile 3 → keine Daten dahinter.
        dialog._header_spin.setValue(3)
        assert _ok_enabled(dialog) is False
        # Zeile 2 ist OK (1 Datenzeile danach).
        dialog._header_spin.setValue(2)
        assert _ok_enabled(dialog) is True

    def test_dialog_result_returns_sheet_and_header(
        self, qtbot: QtBot, multi_sheet_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(multi_sheet_path, importer)
        qtbot.addWidget(dialog)
        combo = dialog._sheet_combo
        assert combo is not None
        combo.setCurrentIndex(combo.findData("Zweites"))
        dialog._header_spin.setValue(1)
        dialog._on_accept()
        result = dialog.get_result()
        assert isinstance(result, ImportOptionsResult)
        assert result.sheet_name == "Zweites"
        assert result.header_row == 0

    def test_dialog_cancel_returns_rejected(
        self, qtbot: QtBot, simple_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(simple_path, importer)
        qtbot.addWidget(dialog)
        dialog.reject()
        assert dialog.result() == int(QDialog.DialogCode.Rejected)
        assert dialog.get_result() is None


# ---------------------------------------------------------------------------
# Sprint 29: „keine Kopfzeile" + CSV
# ---------------------------------------------------------------------------


@pytest.fixture
def header_csv_path(tmp_path: Path) -> Path:
    """CSV mit einer Titelzeile + Leerzeile, Kopfzeile in Zeile 3 (raw index 2)."""
    path = tmp_path / "header.csv"
    path.write_text(
        "Quartalsbericht\n\nKonto,Bezeichnung,Saldo\n1000,Kasse,500.5\n2000,Bank,1234\n",
        encoding="utf-8",
    )
    return path


class TestNoHeaderOption:
    def test_no_header_checkbox_disables_header_spin(
        self, qtbot: QtBot, simple_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(simple_path, importer)
        qtbot.addWidget(dialog)
        assert dialog._header_spin.isEnabled() is True
        dialog._no_header_check.setChecked(True)
        assert dialog._header_spin.isEnabled() is False

    def test_no_header_result_header_row_is_none(
        self, qtbot: QtBot, simple_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(simple_path, importer)
        qtbot.addWidget(dialog)
        dialog._no_header_check.setChecked(True)
        dialog._on_accept()
        result = dialog.get_result()
        assert isinstance(result, ImportOptionsResult)
        assert result.header_row is None
        assert result.sheet_name == "Daten"

    def test_no_header_keeps_ok_enabled_even_at_last_row(
        self, qtbot: QtBot, simple_path: Path, importer: ExcelImporter
    ) -> None:
        # Header-Spin auf letzte Zeile = ungültig für „mit Kopfzeile", aber
        # „keine Kopfzeile" ist trotzdem gültig (alle Zeilen sind Daten).
        dialog = ImportOptionsDialog(simple_path, importer)
        qtbot.addWidget(dialog)
        dialog._header_spin.setValue(3)
        assert _ok_enabled(dialog) is False
        dialog._no_header_check.setChecked(True)
        assert _ok_enabled(dialog) is True


class TestCsvImportOptions:
    def test_csv_has_no_sheet_combo(
        self, qtbot: QtBot, header_csv_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(header_csv_path, importer)
        qtbot.addWidget(dialog)
        assert dialog._sheet_combo is None

    def test_csv_preview_renders_and_preselects_header(
        self, qtbot: QtBot, header_csv_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(header_csv_path, importer)
        qtbot.addWidget(dialog)
        assert dialog._preview_table.rowCount() == 5
        # Kopfzeile in raw row 3 (1-basiert) erkannt → Spin-Default 3.
        assert dialog._header_spin.value() == 3

    def test_csv_result_sheet_name_none(
        self, qtbot: QtBot, header_csv_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(header_csv_path, importer)
        qtbot.addWidget(dialog)
        dialog._header_spin.setValue(3)
        dialog._on_accept()
        result = dialog.get_result()
        assert isinstance(result, ImportOptionsResult)
        assert result.sheet_name is None
        assert result.header_row == 2

    def test_csv_no_header_result(
        self, qtbot: QtBot, header_csv_path: Path, importer: ExcelImporter
    ) -> None:
        dialog = ImportOptionsDialog(header_csv_path, importer)
        qtbot.addWidget(dialog)
        dialog._no_header_check.setChecked(True)
        dialog._on_accept()
        result = dialog.get_result()
        assert isinstance(result, ImportOptionsResult)
        assert result.sheet_name is None
        assert result.header_row is None
