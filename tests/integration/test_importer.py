"""Integration: ExcelImporter – Excel/CSV → Dataset, Header-Detection, Encoding."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from sampling_tool.core.cancellation import CancellationToken, OperationCancelled
from sampling_tool.io.importer import (
    DataImportError,
    ExcelImporter,
    ImportResult,
    SheetInfo,
    SheetPreview,
)


@pytest.fixture
def importer() -> ExcelImporter:
    return ExcelImporter()


class TestImportXlsx:
    def test_importiert_einfache_xlsx_mit_header(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        result = importer.import_file(simple_xlsx)
        rows = list(result.rows)
        assert isinstance(result, ImportResult)
        ds = result.dataset
        assert ds.columns == ("Name", "Betrag", "Quote", "Buchungsdatum")
        assert len(rows) == 10
        assert rows[0].row_id == 1
        assert rows[0].values["Name"] == "Posten 1"
        assert rows[0].values["Betrag"] == 101
        assert isinstance(rows[0].values["Buchungsdatum"], datetime)
        assert ds.source_file == str(simple_xlsx)
        assert result.stats.skipped_rows == 0

    def test_importiert_xlsm_und_ignoriert_makros(
        self, importer: ExcelImporter, xlsm_macro: Path
    ) -> None:
        result = importer.import_file(xlsm_macro)
        rows = list(result.rows)
        assert result.dataset.columns == ("A", "B")
        assert len(rows) == 2
        assert rows[0].values == {"A": 1, "B": 2}

    def test_sheet_auswahl_bei_multi_sheet(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        result = importer.import_file(multi_sheet_xlsx, sheet_name="Stammdaten")
        rows = list(result.rows)
        assert result.dataset.columns == ("KundenID", "Land")
        assert len(rows) == 3
        assert rows[0].values["Land"] == "AUT"

    def test_default_sheet_ohne_argument(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        result = importer.import_file(multi_sheet_xlsx)
        list(result.rows)
        # Erstes/aktives Sheet ist "Buchungen"
        assert result.dataset.columns == ("BuchungsID", "Betrag")

    def test_unbekanntes_sheet_wirft_klare_meldung(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        with pytest.raises(DataImportError, match="existiert nicht"):
            importer.import_file(multi_sheet_xlsx, sheet_name="GibtEsNicht")

    def test_header_detection_mit_leerzeilen(
        self, importer: ExcelImporter, leading_blank_xlsx: Path
    ) -> None:
        result = importer.import_file(leading_blank_xlsx)
        rows = list(result.rows)
        assert result.dataset.columns == ("Konto", "Bezeichnung", "Saldo")
        assert len(rows) == 2
        # 3 leere Vorzeilen wurden geskipped
        assert result.stats.skipped_rows == 3
        assert rows[0].values["Saldo"] == 500.50

    def test_duplikat_spaltennamen_bekommen_suffix(
        self, importer: ExcelImporter, duplicate_columns_xlsx: Path
    ) -> None:
        result = importer.import_file(duplicate_columns_xlsx)
        list(result.rows)
        assert result.dataset.columns == ("Betrag", "Betrag_2", "Betrag_3")
        assert any("Doppelter Spaltenname" in w for w in result.stats.warnings)

    def test_leere_xlsx_wirft_fachlichen_fehler(
        self, importer: ExcelImporter, empty_xlsx: Path
    ) -> None:
        with pytest.raises(DataImportError, match="Keine Spaltenüberschriften"):
            importer.import_file(empty_xlsx)

    def test_unbekannter_dateityp(self, importer: ExcelImporter, tmp_path: Path) -> None:
        bogus = tmp_path / "data.xyz"
        bogus.write_text("nope")
        with pytest.raises(DataImportError, match="wird nicht unterstützt"):
            importer.import_file(bogus)

    def test_nicht_existente_datei(self, importer: ExcelImporter, tmp_path: Path) -> None:
        with pytest.raises(DataImportError, match="nicht gefunden"):
            importer.import_file(tmp_path / "fehlt.xlsx")

    def test_datentyp_konvertierung_zahl_string_datum(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        result = importer.import_file(simple_xlsx)
        rows = list(result.rows)
        first = rows[0].values
        assert isinstance(first["Name"], str)
        assert isinstance(first["Betrag"], int)
        assert isinstance(first["Quote"], float)
        assert isinstance(first["Buchungsdatum"], datetime)


class TestImportCsv:
    def test_csv_utf8(self, importer: ExcelImporter, utf8_csv: Path) -> None:
        result = importer.import_file(utf8_csv)
        rows = list(result.rows)
        assert result.dataset.columns == ("Name", "Stadt")
        assert rows[0].values == {"Name": "Müller", "Stadt": "Wien"}
        # utf-8 ohne BOM erzeugt keine Encoding-Warnung
        assert not any("Encoding" in w for w in result.stats.warnings)

    def test_csv_utf8_bom(self, importer: ExcelImporter, utf8_bom_csv: Path) -> None:
        result = importer.import_file(utf8_bom_csv)
        list(result.rows)
        assert result.dataset.columns == ("Name", "Stadt")
        # utf-8-sig wurde gewählt → Warnung
        assert any("utf-8-sig" in w for w in result.stats.warnings)

    def test_csv_cp1252(self, importer: ExcelImporter, cp1252_csv: Path) -> None:
        # latin-1 nimmt jedes Byte und liest in dem Zeichensatz – das ist für
        # die hier verwendeten Umlaute kompatibel mit cp1252.
        result = importer.import_file(cp1252_csv)
        rows = list(result.rows)
        assert result.dataset.columns == ("Name", "Stadt")
        assert rows[0].values["Name"] == "Müller"
        assert any("Encoding" in w for w in result.stats.warnings)

    def test_csv_mit_semikolon_separator(self, importer: ExcelImporter, tmp_path: Path) -> None:
        path = tmp_path / "semi.csv"
        path.write_text("a;b;c\n1;2;3\n4;5;6\n", encoding="utf-8")
        result = importer.import_file(path)
        rows = list(result.rows)
        assert result.dataset.columns == ("a", "b", "c")
        assert rows[0].values == {"a": 1, "b": 2, "c": 3}

    def test_leere_csv(self, importer: ExcelImporter, tmp_path: Path) -> None:
        path = tmp_path / "empty.csv"
        path.write_text("", encoding="utf-8")
        with pytest.raises(DataImportError, match="keine Daten"):
            importer.import_file(path)


class TestPreviewUndDetectSheets:
    def test_detect_sheets_listet_alle(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        sheets = importer.detect_sheets(multi_sheet_xlsx)
        assert sheets == ["Buchungen", "Stammdaten"]

    def test_detect_sheets_nur_excel(self, importer: ExcelImporter, utf8_csv: Path) -> None:
        with pytest.raises(DataImportError, match="nur für Excel"):
            importer.detect_sheets(utf8_csv)

    def test_preview_liefert_n_rows_xlsx(self, importer: ExcelImporter, simple_xlsx: Path) -> None:
        cols, rows = importer.preview(simple_xlsx, n_rows=3)
        assert cols == ["Name", "Betrag", "Quote", "Buchungsdatum"]
        assert len(rows) == 3
        assert rows[0]["Name"] == "Posten 1"

    def test_preview_n_rows_groesser_als_datei(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        _cols, rows = importer.preview(simple_xlsx, n_rows=999)
        assert len(rows) == 10

    def test_preview_csv(self, importer: ExcelImporter, utf8_csv: Path) -> None:
        cols, rows = importer.preview(utf8_csv, n_rows=1)
        assert cols == ["Name", "Stadt"]
        assert len(rows) == 1
        assert rows[0] == {"Name": "Müller", "Stadt": "Wien"}


class TestCalamineIntegration:
    """Sprint 10.2: ExcelImporter nutzt python-calamine als Excel-Engine."""

    def test_importer_quelle_referenziert_calamine_nicht_openpyxl(self) -> None:
        """openpyxl darf im Excel-Import-Pfad nicht mehr auftauchen."""
        import inspect

        from sampling_tool.io import importer

        src = inspect.getsource(importer)
        # Calamine ist da
        assert "CalamineWorkbook" in src
        # openpyxl darf nicht mehr importiert werden (Module-Level)
        assert "from openpyxl" not in src
        assert "import openpyxl" not in src

    def test_native_python_types_aus_excel(self, simple_xlsx: Path) -> None:
        """Zahlen, Datums-Werte und Strings kommen als Python-Native-Typen."""
        result = ExcelImporter().import_file(simple_xlsx)
        rows = list(result.rows)
        first = rows[0].values
        # Ganzzahliger Excel-Wert → int (auch wenn Calamine intern float liefert)
        assert isinstance(first["Betrag"], int)
        assert first["Betrag"] == 101
        # Echter Float bleibt float
        assert isinstance(first["Quote"], float)
        # datetime statt date (Calamine liefert date für 00:00:00-Werte)
        assert isinstance(first["Buchungsdatum"], datetime)

    def test_leere_zellen_als_none(self, tmp_path: Path) -> None:
        """Calamine liefert leere Zellen als ``""`` – muss zu None normalisiert werden."""
        path = tmp_path / "with_blanks.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["A", "B", "C"])
        ws.append(["x", None, None])
        ws.append([None, 5, None])
        wb.save(path)

        result = ExcelImporter().import_file(path)
        rows = list(result.rows)
        first = rows[0].values
        second = rows[1].values
        assert first["A"] == "x"
        assert first["B"] is None
        assert first["C"] is None
        assert second["A"] is None
        assert second["B"] == 5


class TestProgressCallback:
    """Sprint 11.3 – Streaming: Progress feuert erst beim Generator-Verbrauch."""

    def test_callback_finaler_tick_nach_voller_iteration(self, simple_xlsx: Path) -> None:
        events: list[tuple[int, int]] = []
        importer = ExcelImporter(progress=lambda c, t: events.append((c, t)))
        result = importer.import_file(simple_xlsx)
        # Vor dem Konsumieren ist der Generator nicht gelaufen → keine Events.
        assert events == []
        list(result.rows)
        # Nach voller Iteration mindestens der finale Tick.
        assert events[-1] == (10, 10)

    def test_callback_bei_leeren_daten(self, tmp_path: Path) -> None:
        # Datei mit Header aber ohne Datenzeilen → 0 Events vor Konsum,
        # ein finaler (0, 0)-Tick nach Konsum.
        path = tmp_path / "header_only.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["A", "B"])
        wb.save(path)

        events: list[tuple[int, int]] = []
        result = ExcelImporter(progress=lambda c, t: events.append((c, t))).import_file(path)
        list(result.rows)
        assert events == [(0, 0)]


class TestStreamingImport:
    """Sprint 11.3: ImportResult.rows ist ein einmalig konsumierbarer Iterator."""

    def test_rows_ist_kein_tuple_oder_list(self, simple_xlsx: Path) -> None:
        result = ExcelImporter().import_file(simple_xlsx)
        # Kein materialisierter Container.
        assert not isinstance(result.rows, list | tuple)
        # Aber iterierbar – Generator hat __next__.
        assert hasattr(result.rows, "__next__")

    def test_rows_nur_einmal_konsumierbar(self, simple_xlsx: Path) -> None:
        result = ExcelImporter().import_file(simple_xlsx)
        first_pass = list(result.rows)
        assert len(first_pass) == 10
        second_pass = list(result.rows)
        assert second_pass == []

    def test_stats_vor_iteration_leer(self, simple_xlsx: Path) -> None:
        result = ExcelImporter().import_file(simple_xlsx)
        assert result.stats.processed_count == 0
        assert result.stats.skipped_rows == 0

    def test_stats_nach_iteration_gefuellt(self, leading_blank_xlsx: Path) -> None:
        result = ExcelImporter().import_file(leading_blank_xlsx)
        list(result.rows)
        # 3 Leerzeilen vorm Header werden beim Header-Pass schon erkannt.
        assert result.stats.skipped_rows == 3
        assert result.stats.processed_count == 2


class TestRepoStreamingRowCount:
    """Sprint 11.3: DatasetRepo.create korrigiert row_count basierend auf
    der echten Anzahl persistierter Rows (wichtig bei Skipped-Rows im
    Streaming-Import)."""

    def test_repo_create_uebernimmt_actual_count(self, simple_xlsx: Path, tmp_path: Path) -> None:
        from dataclasses import replace as dc_replace

        from sampling_tool.core.models import Engagement
        from sampling_tool.persistence.database import Database
        from sampling_tool.persistence.repositories import DatasetRepo, EngagementRepo

        db = Database(tmp_path / "streaming.db")
        db.migrate()
        eng = EngagementRepo(db.connect()).get_or_create(
            Engagement(
                auditor_name="A", client_name="C", auditor_position="S", audit_type="ISAE 3402"
            )
        )
        assert eng.id is not None

        result = ExcelImporter().import_file(simple_xlsx)
        dataset = dc_replace(result.dataset, engagement_id=eng.id)
        stored = DatasetRepo(db.connect()).create(dataset, result.rows)

        assert stored.row_count == 10
        assert stored.id is not None
        # Echte Persistenz prüfen.
        all_rows = DatasetRepo(db.connect()).get_all_rows(stored.id)
        assert len(all_rows) == 10
        db.close()

    def test_repo_create_korrigiert_bei_skipped(
        self, leading_blank_xlsx: Path, tmp_path: Path
    ) -> None:
        # leading_blank_xlsx hat 3 Leerzeilen vor dem Header, 2 Daten-Rows.
        # `sheet.total_height` enthält die Leerzeilen, der Importer
        # schätzt also evtl. zu hoch – Repo korrigiert auf 2.
        from dataclasses import replace as dc_replace

        from sampling_tool.core.models import Engagement
        from sampling_tool.persistence.database import Database
        from sampling_tool.persistence.repositories import DatasetRepo, EngagementRepo

        db = Database(tmp_path / "skipped.db")
        db.migrate()
        eng = EngagementRepo(db.connect()).get_or_create(
            Engagement(
                auditor_name="A", client_name="C", auditor_position="S", audit_type="ISAE 3402"
            )
        )
        assert eng.id is not None

        result = ExcelImporter().import_file(leading_blank_xlsx)
        dataset = dc_replace(result.dataset, engagement_id=eng.id)
        stored = DatasetRepo(db.connect()).create(dataset, result.rows)

        assert stored.row_count == 2
        db.close()


# ---------------------------------------------------------------------------
# Sprint 16: Multi-Sheet + Header-Detection-Dialog-API
# ---------------------------------------------------------------------------


@pytest.fixture
def header_in_row_3_xlsx(tmp_path: Path) -> Path:
    """xlsx mit 2 Metadaten-Zeilen über dem Header in Zeile 3 (0-basiert: index 2)."""
    path = tmp_path / "header_row_3.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Audit-Report Q1 2026", None, None])
    ws.append(["Erstellt am 2026-04-12", None, None])
    ws.append(["Konto", "Bezeichnung", "Saldo"])
    ws.append(["1000", "Kasse", 500.50])
    ws.append(["2000", "Bank", 1234.00])
    ws.append(["3000", "Forderungen", 750.25])
    wb.save(path)
    return path


@pytest.fixture
def ambiguous_xlsx(tmp_path: Path) -> Path:
    """xlsx ohne klare Header-Zeile: erste non-blank Zeile sieht aus wie Daten (Zahlen)."""
    path = tmp_path / "ambiguous.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([100, 200, 300])
    ws.append([400, 500, 600])
    ws.append([700, 800, 900])
    wb.save(path)
    return path


@pytest.fixture
def three_sheet_xlsx(tmp_path: Path) -> Path:
    """xlsx mit drei Sheets, jeder mit Header in Zeile 1."""
    path = tmp_path / "three_sheet.xlsx"
    wb = Workbook()
    first = wb.active
    assert first is not None
    first.title = "Erstes"
    first.append(["a", "b"])
    first.append([1, 2])

    second = wb.create_sheet("Zweites")
    second.append(["x", "y", "z"])
    second.append(["A", "B", "C"])
    second.append(["D", "E", "F"])

    third = wb.create_sheet("Drittes")
    third.append(["nur_eine_spalte"])
    third.append([42])
    wb.save(path)
    return path


class TestListSheets:
    def test_single_sheet_workbook(self, importer: ExcelImporter, simple_xlsx: Path) -> None:
        sheets = importer.list_sheets(simple_xlsx)
        assert len(sheets) == 1
        assert isinstance(sheets[0], SheetInfo)
        assert sheets[0].name == "Daten"

    def test_multi_sheet_workbook_returns_all(
        self, importer: ExcelImporter, three_sheet_xlsx: Path
    ) -> None:
        sheets = importer.list_sheets(three_sheet_xlsx)
        assert [s.name for s in sheets] == ["Erstes", "Zweites", "Drittes"]

    def test_sheet_info_enthaelt_row_und_column_count(
        self, importer: ExcelImporter, three_sheet_xlsx: Path
    ) -> None:
        sheets = importer.list_sheets(three_sheet_xlsx)
        # "Zweites" hat 3 Zeilen × 3 Spalten
        zweites = next(s for s in sheets if s.name == "Zweites")
        assert zweites.row_count >= 3
        assert zweites.column_count == 3
        drittes = next(s for s in sheets if s.name == "Drittes")
        assert drittes.column_count == 1

    def test_list_sheets_csv_wirft_fehler(self, importer: ExcelImporter, utf8_csv: Path) -> None:
        with pytest.raises(DataImportError, match="nur für Excel"):
            importer.list_sheets(utf8_csv)


class TestPreviewSheet:
    def test_preview_returns_raw_2d_rows(self, importer: ExcelImporter, simple_xlsx: Path) -> None:
        preview = importer.preview_sheet(simple_xlsx, "Daten", max_rows=3)
        assert isinstance(preview, SheetPreview)
        assert preview.sheet_name == "Daten"
        # 3 Zeilen: Header + 2 Daten
        assert len(preview.rows) == 3
        # Rohzellen, kein Dict-Mapping – Header steht IN den rows.
        assert preview.rows[0] == ("Name", "Betrag", "Quote", "Buchungsdatum")

    def test_preview_confidence_high_when_header_in_row_1(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        preview = importer.preview_sheet(simple_xlsx, "Daten")
        assert preview.confidence == "high"
        assert preview.detected_header_row == 0

    def test_preview_confidence_low_when_header_in_row_4(
        self, importer: ExcelImporter, leading_blank_xlsx: Path
    ) -> None:
        # leading_blank_xlsx hat 3 echte Leerzeilen vor dem Header in Zeile 4.
        preview = importer.preview_sheet(leading_blank_xlsx, "Sheet")
        assert preview.confidence == "low"
        # 0-basiert: Header in Zeile 4 → index 3
        assert preview.detected_header_row == 3

    def test_preview_confidence_ambiguous_when_no_clear_header(
        self, importer: ExcelImporter, ambiguous_xlsx: Path
    ) -> None:
        preview = importer.preview_sheet(ambiguous_xlsx, "Sheet")
        assert preview.confidence == "ambiguous"

    def test_preview_max_rows_respected(self, importer: ExcelImporter, simple_xlsx: Path) -> None:
        preview = importer.preview_sheet(simple_xlsx, "Daten", max_rows=5)
        assert len(preview.rows) == 5

    def test_preview_max_rows_groesser_als_datei(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        # Datei hat 11 Zeilen (1 Header + 10 Daten); max_rows=999 → alles
        preview = importer.preview_sheet(simple_xlsx, "Daten", max_rows=999)
        assert len(preview.rows) == 11

    def test_preview_unbekanntes_sheet_wirft(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        with pytest.raises(DataImportError, match="existiert nicht"):
            importer.preview_sheet(multi_sheet_xlsx, "GibtEsNicht")

    def test_preview_csv_wirft(self, importer: ExcelImporter, utf8_csv: Path) -> None:
        with pytest.raises(DataImportError, match="nur für Excel"):
            importer.preview_sheet(utf8_csv, "Sheet")


class TestImportFileConfigured:
    def test_import_mit_explizitem_header_row(
        self, importer: ExcelImporter, header_in_row_3_xlsx: Path
    ) -> None:
        # Header steht in Zeile 3 (0-basiert: 2). Explizit übergeben.
        result = importer.import_file_configured(
            header_in_row_3_xlsx, sheet_name="Sheet", header_row=2
        )
        rows = list(result.rows)
        assert result.dataset.columns == ("Konto", "Bezeichnung", "Saldo")
        assert len(rows) == 3
        assert rows[0].values["Konto"] == 1000

    def test_import_mit_explizitem_sheet_name(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        result = importer.import_file_configured(
            multi_sheet_xlsx, sheet_name="Stammdaten", header_row=0
        )
        rows = list(result.rows)
        assert result.dataset.columns == ("KundenID", "Land")
        assert len(rows) == 3
        assert rows[0].values["Land"] == "AUT"

    def test_configured_gleich_auto_bei_perfektem_header(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        # Wenn der Auto-Header korrekt war, müssen configured und auto
        # dasselbe Resultat liefern.
        auto = importer.import_file(simple_xlsx)
        auto_rows = list(auto.rows)
        configured = importer.import_file_configured(simple_xlsx, sheet_name="Daten", header_row=0)
        configured_rows = list(configured.rows)
        assert auto.dataset.columns == configured.dataset.columns
        assert [r.values for r in auto_rows] == [r.values for r in configured_rows]

    def test_configured_invalid_sheet_name_wirft(
        self, importer: ExcelImporter, multi_sheet_xlsx: Path
    ) -> None:
        with pytest.raises(DataImportError, match="existiert nicht"):
            importer.import_file_configured(multi_sheet_xlsx, sheet_name="NopeNope", header_row=0)

    def test_configured_header_row_jenseits_der_daten_wirft(
        self, importer: ExcelImporter, simple_xlsx: Path
    ) -> None:
        # simple_xlsx hat 11 Zeilen. header_row=50 ist out of bounds.
        with pytest.raises(DataImportError, match="Header"):
            importer.import_file_configured(simple_xlsx, sheet_name="Daten", header_row=50)

    def test_configured_uebersprungene_metadaten_landen_in_stats(
        self, importer: ExcelImporter, header_in_row_3_xlsx: Path
    ) -> None:
        result = importer.import_file_configured(
            header_in_row_3_xlsx, sheet_name="Sheet", header_row=2
        )
        list(result.rows)
        # 2 Metadaten-Zeilen über dem Header wurden geskipped.
        assert result.stats.skipped_rows == 2

    def test_configured_reproducible(
        self, importer: ExcelImporter, header_in_row_3_xlsx: Path
    ) -> None:
        """Gleiche Datei + gleiche Sheet/Header-Wahl → identische DatasetRows."""
        result1 = importer.import_file_configured(
            header_in_row_3_xlsx, sheet_name="Sheet", header_row=2
        )
        result2 = importer.import_file_configured(
            header_in_row_3_xlsx, sheet_name="Sheet", header_row=2
        )
        assert [r.values for r in result1.rows] == [r.values for r in result2.rows]

    def test_configured_csv_wirft(self, importer: ExcelImporter, utf8_csv: Path) -> None:
        # CSVs haben keine Sheets – import_file_configured ist Excel-only.
        with pytest.raises(DataImportError, match="nur für Excel"):
            importer.import_file_configured(utf8_csv, sheet_name="Sheet", header_row=0)


# ---------------------------------------------------------------------------
# Sprint 17: Cancellation-Support
# ---------------------------------------------------------------------------


class TestImporterCancellation:
    """ExcelImporter mit CancellationToken – Streaming-Pfad muss
    periodisch prüfen und OperationCancelled werfen."""

    def test_import_respects_cancellation_token_set_before_start(self, simple_xlsx: Path) -> None:
        token = CancellationToken()
        token.set()
        importer = ExcelImporter(cancellation=token)
        result = importer.import_file(simple_xlsx)
        with pytest.raises(OperationCancelled):
            list(result.rows)

    def test_import_without_cancellation_runs_to_completion(self, simple_xlsx: Path) -> None:
        # Sanity: kein Token → läuft wie bisher.
        importer = ExcelImporter()
        result = importer.import_file(simple_xlsx)
        rows = list(result.rows)
        assert len(rows) == 10

    def test_import_with_unset_token_runs_to_completion(self, simple_xlsx: Path) -> None:
        token = CancellationToken()  # nicht gesetzt
        importer = ExcelImporter(cancellation=token)
        result = importer.import_file(simple_xlsx)
        rows = list(result.rows)
        assert len(rows) == 10

    def test_import_checks_cancellation_during_iteration(self, tmp_path: Path) -> None:
        """Token wird während der Iteration gesetzt → bricht beim nächsten
        Progress-Tick (alle ``_PROGRESS_INTERVAL``=1000 Rows) ab."""
        path = tmp_path / "many_rows.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["a", "b"])
        for i in range(1, 3001):
            ws.append([i, f"r{i}"])
        wb.save(path)

        from sampling_tool.core.models import DatasetRow

        token = CancellationToken()
        importer = ExcelImporter(cancellation=token)
        result = importer.import_file(path)
        consumed: list[DatasetRow] = []

        def _consume_until_cancel() -> None:
            for i, row in enumerate(result.rows):
                consumed.append(row)
                # Nach 500 Rows abbrechen – nächster Tick (bei 1000) bricht ab.
                if i == 500:
                    token.set()

        with pytest.raises(OperationCancelled):
            _consume_until_cancel()
        # Hat gestoppt bevor die 3000 fertig waren.
        assert 500 < len(consumed) < 3000

    def test_configured_import_respects_cancellation(self, simple_xlsx: Path) -> None:
        token = CancellationToken()
        token.set()
        importer = ExcelImporter(cancellation=token)
        result = importer.import_file_configured(simple_xlsx, "Daten", 0)
        with pytest.raises(OperationCancelled):
            list(result.rows)

    def test_csv_import_respects_cancellation(self, utf8_csv: Path) -> None:
        token = CancellationToken()
        token.set()
        importer = ExcelImporter(cancellation=token)
        result = importer.import_file(utf8_csv)
        with pytest.raises(OperationCancelled):
            list(result.rows)
