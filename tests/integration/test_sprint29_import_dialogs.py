"""Sprint 29 – Import-Dialoge: Multi-Sheet-Auswahl + Header-Detection.

Akzeptanz-Tests für die zwei Komfort-Features, additiv auf der
Sprint-16-Infrastruktur erweitert:

- **Multi-Sheet-Auswahl** (Sprint 16) – hier als Regressions-Guard gepinnt:
  >1 Blatt ⇒ Auswahl nötig, genau 1 Blatt ⇒ direkter Import.
- **Header-Detection** inkl. **„keine Kopfzeile"** (neu) und **CSV** (neu):
  Kopfzeile in Zeile N überspringt die Zeilen darüber; „keine Kopfzeile"
  vergibt generische Spaltennamen; gilt für Excel UND CSV.
- **Sauberer Default-Pfad bleibt unverändert**: ein Blatt, Kopfzeile Zeile 1,
  Default akzeptiert ⇒ Ergebnis identisch zum bisherigen Auto-Import
  (gleiche Spalten, Zeilen, Werte, Typen).

Die Coercion/Werte-Logik (Sprint 26, `_coerce_*`) wird NICHT angefasst –
Header-/Blatt-Wahl ändert nur die Auswahl der Rohzeilen, nicht deren
Kodierung.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from sampling_tool.core.cancellation import CancellationToken, OperationCancelled
from sampling_tool.io.importer import DataImportError, ExcelImporter

pytestmark = pytest.mark.integration


@pytest.fixture
def importer() -> ExcelImporter:
    return ExcelImporter()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def clean_xlsx(tmp_path: Path) -> Path:
    """Ein Blatt, Kopfzeile in Zeile 1, drei Datenzeilen – der „saubere" Fall."""
    path = tmp_path / "clean.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Daten"
    ws.append(["Konto", "Bezeichnung", "Saldo"])
    ws.append([1000, "Kasse", 500.50])
    ws.append([2000, "Bank", 1234.00])
    ws.append([3000, "Forderungen", 750.25])
    wb.save(path)
    return path


@pytest.fixture
def three_sheet_xlsx(tmp_path: Path) -> Path:
    """Drei Blätter; „Zweites" hat drei Spalten + zwei Datenzeilen."""
    path = tmp_path / "three.xlsx"
    wb = Workbook()
    first = wb.active
    assert first is not None
    first.title = "Erstes"
    first.append(["a", "b"])
    first.append([1, 2])
    second = wb.create_sheet("Zweites")
    second.append(["x", "y", "z"])
    second.append([10, 20, 30])
    second.append([40, 50, 60])
    third = wb.create_sheet("Drittes")
    third.append(["nur_spalte"])
    third.append([42])
    wb.save(path)
    return path


@pytest.fixture
def header_in_row_5_xlsx(tmp_path: Path) -> Path:
    """Vier (teils leere) Titelzeilen, Kopfzeile in Zeile 5 (0-basiert: 4)."""
    path = tmp_path / "header_row_5.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Monatsabschluss April 2026", None, None])
    ws.append(["Mandant: ACME GmbH", None, None])
    ws.append([None, None, None])
    ws.append(["Stand: 2026-04-30", None, None])
    ws.append(["Konto", "Bezeichnung", "Saldo"])
    ws.append([1000, "Kasse", 500.50])
    ws.append([2000, "Bank", 1234.00])
    wb.save(path)
    return path


@pytest.fixture
def no_header_xlsx(tmp_path: Path) -> Path:
    """Reine Datenzeilen ohne Kopfzeile (alles Zahlen)."""
    path = tmp_path / "no_header.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([100, 200, 300])
    ws.append([400, 500, 600])
    ws.append([700, 800, 900])
    wb.save(path)
    return path


@pytest.fixture
def clean_csv(tmp_path: Path) -> Path:
    """CSV mit Kopfzeile in Zeile 1 – der „saubere" CSV-Fall."""
    path = tmp_path / "clean.csv"
    path.write_text(
        "Konto,Bezeichnung,Saldo\n1000,Kasse,500.5\n2000,Bank,1234\n",
        encoding="utf-8",
    )
    return path


@pytest.fixture
def header_in_row_3_csv(tmp_path: Path) -> Path:
    """CSV mit zwei Titel-/Leerzeilen, Kopfzeile in Zeile 3 (0-basiert: 2)."""
    path = tmp_path / "header_row_3.csv"
    path.write_text(
        "Quartalsbericht\n\nKonto,Bezeichnung,Saldo\n1000,Kasse,500.5\n2000,Bank,1234\n",
        encoding="utf-8",
    )
    return path


@pytest.fixture
def no_header_csv(tmp_path: Path) -> Path:
    """CSV ohne Kopfzeile (alles Zahlen)."""
    path = tmp_path / "no_header.csv"
    path.write_text("100,200,300\n400,500,600\n", encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# Multi-Sheet-Auswahl
# ---------------------------------------------------------------------------


class TestSheetSelection:
    def test_multi_sheet_requires_choice(
        self, importer: ExcelImporter, three_sheet_xlsx: Path
    ) -> None:
        # Drei Blätter ⇒ Auswahl nötig (Dialog).
        assert importer.requires_options_dialog(three_sheet_xlsx) is True
        # Das gewählte Blatt wird importiert (Zeilenzahl passt zum Blatt).
        result = importer.import_file_configured(
            three_sheet_xlsx, sheet_name="Zweites", header_row=0
        )
        rows = list(result.rows)
        assert result.dataset.columns == ("x", "y", "z")
        assert len(rows) == 2
        assert rows[0].values == {"x": 10, "y": 20, "z": 30}

    def test_single_sheet_no_dialog(self, importer: ExcelImporter, clean_xlsx: Path) -> None:
        # Genau ein Blatt + Kopfzeile in Zeile 1 ⇒ kein Dialog, direkter Import.
        assert importer.requires_options_dialog(clean_xlsx) is False
        rows = list(importer.import_file(clean_xlsx).rows)
        assert len(rows) == 3


# ---------------------------------------------------------------------------
# Header-Detection (inkl. „keine Kopfzeile" und CSV)
# ---------------------------------------------------------------------------


class TestHeaderDetection:
    def test_header_row_skips_rows_above(
        self, importer: ExcelImporter, header_in_row_5_xlsx: Path
    ) -> None:
        # Kopfzeile = Zeile 5 (0-basiert: 4) ⇒ Zeilen 1–4 übersprungen,
        # Zeile 5 liefert die Spaltennamen, Daten ab Zeile 6.
        result = importer.import_file_configured(
            header_in_row_5_xlsx, sheet_name=None, header_row=4
        )
        rows = list(result.rows)
        assert result.dataset.columns == ("Konto", "Bezeichnung", "Saldo")
        assert result.stats.skipped_rows == 4
        assert len(rows) == 2
        assert rows[0].values["Konto"] == 1000

    def test_no_header_generates_generic_names(
        self, importer: ExcelImporter, no_header_xlsx: Path
    ) -> None:
        # „keine Kopfzeile" ⇒ generische Spaltennamen, ALLE Zeilen sind Daten.
        result = importer.import_file_configured(no_header_xlsx, sheet_name=None, header_row=None)
        rows = list(result.rows)
        assert result.dataset.columns == ("Spalte 1", "Spalte 2", "Spalte 3")
        assert len(rows) == 3
        assert rows[0].values == {"Spalte 1": 100, "Spalte 2": 200, "Spalte 3": 300}

    def test_autodetect_picks_header_row(
        self, importer: ExcelImporter, header_in_row_5_xlsx: Path
    ) -> None:
        # Bei Titelzeilen über der Tabelle erkennt der Default die richtige Kopfzeile.
        preview = importer.preview_sheet(header_in_row_5_xlsx, "Sheet")
        assert preview.detected_header_row == 4  # 0-basiert
        assert preview.confidence == "low"

    # ---- CSV (gilt gleichermaßen) -------------------------------------

    def test_csv_header_row_skips_rows_above(
        self, importer: ExcelImporter, header_in_row_3_csv: Path
    ) -> None:
        result = importer.import_file_configured(header_in_row_3_csv, sheet_name=None, header_row=2)
        rows = list(result.rows)
        assert result.dataset.columns == ("Konto", "Bezeichnung", "Saldo")
        assert result.stats.skipped_rows == 2
        assert len(rows) == 2
        assert rows[0].values["Konto"] == 1000

    def test_csv_no_header_generates_generic_names(
        self, importer: ExcelImporter, no_header_csv: Path
    ) -> None:
        result = importer.import_file_configured(no_header_csv, sheet_name=None, header_row=None)
        rows = list(result.rows)
        assert result.dataset.columns == ("Spalte 1", "Spalte 2", "Spalte 3")
        assert len(rows) == 2
        assert rows[0].values == {"Spalte 1": 100, "Spalte 2": 200, "Spalte 3": 300}

    def test_csv_autodetect_picks_header_row(
        self, importer: ExcelImporter, header_in_row_3_csv: Path
    ) -> None:
        preview = importer.preview_csv(header_in_row_3_csv)
        assert preview.detected_header_row == 2
        assert preview.confidence == "low"


# ---------------------------------------------------------------------------
# Sauberer Default-Pfad bleibt unverändert (Oracle gegen den Auto-Import)
# ---------------------------------------------------------------------------


class TestImportUnchangedForCleanFiles:
    def test_clean_xlsx_default_identical_to_baseline(
        self, importer: ExcelImporter, clean_xlsx: Path
    ) -> None:
        # Auto-Import (= bisheriges Verhalten) als Baseline.
        baseline = list(ExcelImporter().import_file(clean_xlsx).rows)
        baseline_cols = ExcelImporter().import_file(clean_xlsx).dataset.columns

        # Sprint-29-Pfad mit den Default-Werten: erstes Blatt, Kopfzeile Zeile 1.
        configured_result = importer.import_file_configured(
            clean_xlsx, sheet_name="Daten", header_row=0
        )
        configured = list(configured_result.rows)

        assert configured_result.dataset.columns == baseline_cols
        assert [r.row_id for r in configured] == [r.row_id for r in baseline]
        assert [r.values for r in configured] == [r.values for r in baseline]
        # Typ-Identität (datetime/int/float/str) – kein Drift durch die Auswahl.
        for got, want in zip(configured, baseline, strict=True):
            for col, want_val in want.values.items():
                assert type(got.values[col]) is type(want_val), f"Typ-Drift bei '{col}'"

    def test_clean_csv_default_identical_to_baseline(
        self, importer: ExcelImporter, clean_csv: Path
    ) -> None:
        baseline_result = ExcelImporter().import_file(clean_csv)
        baseline = list(baseline_result.rows)
        baseline_cols = baseline_result.dataset.columns

        configured_result = importer.import_file_configured(
            clean_csv, sheet_name=None, header_row=0
        )
        configured = list(configured_result.rows)

        assert configured_result.dataset.columns == baseline_cols
        assert [r.row_id for r in configured] == [r.row_id for r in baseline]
        assert [r.values for r in configured] == [r.values for r in baseline]
        for got, want in zip(configured, baseline, strict=True):
            for col, want_val in want.values.items():
                assert type(got.values[col]) is type(want_val), f"Typ-Drift bei '{col}'"

    def test_id_column_choice_does_not_change_import(
        self,
        importer: ExcelImporter,
        clean_xlsx: Path,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Sprint 31: die Wahl einer ID-Spalte (QSettings) ändert `dataset_rows` NICHT.

        Die ID-Spalte ist eine reine Anzeige-Hilfe (Sidebar) und lebt in
        QSettings – der Importer liest sie nie. Gleiche Datei → byte-identisches
        Importergebnis, unabhängig davon, ob eine ID-Spalte gewählt wurde.
        """
        from PyQt6.QtCore import QSettings

        from sampling_tool.config import APP_NAME, APP_ORG
        from sampling_tool.ui.dataset_id_store import DatasetIdColumnStore

        # QSettings isolieren, damit der Store-Schreibvorgang echte Prefs nicht anfasst.
        QSettings.setPath(QSettings.Format.IniFormat, QSettings.Scope.UserScope, str(tmp_path))
        QSettings.setDefaultFormat(QSettings.Format.IniFormat)
        monkeypatch.setattr(
            "sampling_tool.ui.settings_store._qsettings",
            lambda: QSettings(
                QSettings.Format.IniFormat, QSettings.Scope.UserScope, APP_ORG, APP_NAME
            ),
        )

        # Referenz ohne ID-Wahl.
        before = list(ExcelImporter().import_file(clean_xlsx).rows)
        before_cols = ExcelImporter().import_file(clean_xlsx).dataset.columns

        # ID-Spalte wählen = reine QSettings-Schreibung, völlig getrennt vom Import.
        DatasetIdColumnStore().set("ACME_ISAE", 1, before_cols[0])

        after_result = ExcelImporter().import_file(clean_xlsx)
        after = list(after_result.rows)
        assert after_result.dataset.columns == before_cols
        assert [r.row_id for r in after] == [r.row_id for r in before]
        assert [r.values for r in after] == [r.values for r in before]
        for got, want in zip(after, before, strict=True):
            for col, want_val in want.values.items():
                assert type(got.values[col]) is type(want_val), f"Typ-Drift bei '{col}'"


# ---------------------------------------------------------------------------
# Cancellation (Sprint-17-Verhalten bleibt erhalten)
# ---------------------------------------------------------------------------


def test_cancellation_still_works(header_in_row_5_xlsx: Path, no_header_csv: Path) -> None:
    # Vorab gesetztes Token: der konfigurierte Excel-Import bricht beim Konsum ab.
    token = CancellationToken()
    token.set()
    importer = ExcelImporter(cancellation=token)
    with pytest.raises(OperationCancelled):
        list(
            importer.import_file_configured(
                header_in_row_5_xlsx, sheet_name=None, header_row=4
            ).rows
        )

    # Auch der konfigurierte CSV-Import respektiert das Token.
    token2 = CancellationToken()
    token2.set()
    importer2 = ExcelImporter(cancellation=token2)
    with pytest.raises(OperationCancelled):
        list(importer2.import_file_configured(no_header_csv, sheet_name=None, header_row=None).rows)


def test_configured_unsupported_filetype_wirft(importer: ExcelImporter, tmp_path: Path) -> None:
    weird = tmp_path / "data.json"
    weird.write_text("{}", encoding="utf-8")
    with pytest.raises(DataImportError):
        importer.import_file_configured(weird, sheet_name=None, header_row=0)


# ---------------------------------------------------------------------------
# Heuristik-Robustheit: ein breiterer Datensatz unter der Kopfzeile darf die
# Kopfzeile in Zeile 1 NICHT vetoen (Review-Findings #1/#4/#7). Saubere
# Dateien sollen weiter lautlos (ohne Dialog) importieren.
# ---------------------------------------------------------------------------


class TestHeaderHeuristicRobustness:
    def test_xlsx_header_with_unlabeled_column_stays_high(
        self, importer: ExcelImporter, tmp_path: Path
    ) -> None:
        # Erste Spalte ohne Überschrift (Index-Spalte), Daten füllen sie aber.
        # Die Kopfzeile (Zeile 1) ist schmaler als die Datenbreite – trotzdem
        # „high" ⇒ kein erzwungener Dialog, lautloser Import wie bisher.
        path = tmp_path / "unlabeled_col.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append([None, "Konto", "Betrag"])
        ws.append([1, 1000, 500])
        ws.append([2, 2000, 600])
        wb.save(path)
        preview = importer.preview_sheet(path, "Sheet")
        assert preview.detected_header_row == 0
        assert preview.confidence == "high"
        assert importer.requires_options_dialog(path) is False

    def test_xlsx_wide_footer_does_not_veto_header(
        self, importer: ExcelImporter, tmp_path: Path
    ) -> None:
        # Eine breitere Fuß-/Summenzeile unter der Tabelle darf die Kopfzeile
        # in Zeile 1 nicht entwerten.
        path = tmp_path / "wide_footer.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Konto", "Betrag"])
        ws.append([1000, 500])
        ws.append([2000, 600])
        ws.append(["Summe", 1100, "geprüft", "Anmerkung"])  # breiter als Header
        wb.save(path)
        preview = importer.preview_sheet(path, "Sheet")
        assert preview.detected_header_row == 0
        assert preview.confidence == "high"

    def test_csv_trailing_comma_header_not_forced_into_dialog(
        self, importer: ExcelImporter, tmp_path: Path
    ) -> None:
        # Kopfzeile mit Trailing-Komma (eine Leer-Spalte), Daten füllen die
        # dritte Spalte. Vor Sprint 29 importierte das lautlos – das muss so
        # bleiben (kein erzwungener Dialog, kein Datenverlust).
        path = tmp_path / "trailing_comma.csv"
        path.write_text("Name,Betrag,\nAlice,100,X\nBob,200,Y\n", encoding="utf-8")
        assert importer.requires_options_dialog(path) is False
        rows = list(importer.import_file(path).rows)
        assert importer.import_file(path).dataset.columns == ("Name", "Betrag", "Spalte_3")
        assert len(rows) == 2

    def test_csv_no_header_ignores_blank_but_wide_row(
        self, importer: ExcelImporter, tmp_path: Path
    ) -> None:
        # Eine leere Zeile mit Trenner (',,' → drei leere Felder) darf die
        # generische Spaltenzahl NICHT aufblähen (Finding #3).
        path = tmp_path / "blank_wide.csv"
        path.write_text("1,2\n,,\n3,4\n", encoding="utf-8")
        result = importer.import_file_configured(path, sheet_name=None, header_row=None)
        rows = list(result.rows)
        assert result.dataset.columns == ("Spalte 1", "Spalte 2")
        assert len(rows) == 2  # die ',,'-Zeile ist leer → übersprungen

    def test_csv_configured_trailing_blanks_match_auto_skipped(
        self, importer: ExcelImporter, tmp_path: Path
    ) -> None:
        # Trailing-Leerzeilen dürfen die „übersprungen"-Bilanz nicht aufblähen
        # (konfigurierter Pfad konsistent mit Auto-Pfad – Findings #2/#9).
        path = tmp_path / "trailing_blanks.csv"
        path.write_text("Konto,Betrag\n1000,500\n2000,600\n\n\n", encoding="utf-8")
        auto = importer.import_file(path)
        list(auto.rows)
        configured = importer.import_file_configured(path, sheet_name=None, header_row=0)
        rows = list(configured.rows)
        assert len(rows) == 2
        assert configured.stats.skipped_rows == auto.stats.skipped_rows
