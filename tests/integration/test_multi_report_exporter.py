"""Tests für `MultiSheetReportExporter` – alle 4 Sheets + Chart-Bild."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from sampling_tool.core.models import (
    AuditEvent,
    Dataset,
    Engagement,
    FilterOperator,
    SampleConfig,
    SampleResult,
    SamplingMethod,
)
from sampling_tool.io.exporter import ExportError
from sampling_tool.io.multi_report_exporter import MultiSheetReportExporter

pytestmark = pytest.mark.integration


@pytest.fixture
def engagement() -> Engagement:
    return Engagement(
        auditor_name="Anna",
        client_name="ACME GmbH",
        auditor_position="Senior",
        audit_type="ISAE 3402",
        id=1,
    )


@pytest.fixture
def datasets() -> list[Dataset]:
    return [
        Dataset(
            name="Buchungen",
            columns=("Konto", "Betrag"),
            row_count=1,
            engagement_id=1,
            id=1,
        ),
    ]


@pytest.fixture
def samples() -> list[SampleResult]:
    cfg1 = SampleConfig(method=SamplingMethod.SIMPLE, size=5, seed=42)
    cfg2 = SampleConfig(method=SamplingMethod.STRATIFIED, size=3, seed=7, stratum_field="Land")
    return [
        SampleResult(
            config=cfg1,
            selected_row_ids=(1, 2, 3, 4, 5),
            population_size=10,
            drawn_at=datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
            created_by="anna",
            id=1,
        ),
        SampleResult(
            config=cfg2,
            selected_row_ids=(7, 8, 9),
            population_size=10,
            drawn_at=datetime(2026, 5, 2, 11, 0, tzinfo=UTC),
            created_by="bob",
            id=2,
        ),
    ]


@pytest.fixture
def audit_events() -> list[AuditEvent]:
    return [
        AuditEvent(
            event_type="sampling",
            engagement_id=1,
            user_name="anna",
            sample_id=1,
            sample_size=5,
            sample_percent=50.0,
            seed=42,
            timestamp=datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
            id=1,
        ),
        AuditEvent(
            event_type="export",
            engagement_id=1,
            user_name="anna",
            sample_id=1,
            export_file="/exports/sample.xlsx",
            timestamp=datetime(2026, 5, 1, 11, 0, tzinfo=UTC),
            id=2,
        ),
    ]


class TestMultiSheetReportExporter:
    def test_creates_all_four_sheets(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "bericht.xlsx"
        result = MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        assert result.exists()
        wb = load_workbook(result)
        names = wb.sheetnames
        assert any("Übersicht" in n for n in names)
        assert any("AuditTrail" in n for n in names)
        assert any("Samples" in n for n in names)
        assert any("Statistiken" in n for n in names)

    def test_uebersicht_contains_engagement_info(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        wb = load_workbook(out)
        ws = wb["1. Übersicht"]
        flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        assert "ACME GmbH" in flat
        assert "Anna" in flat
        assert any(v == "2" for v in flat)  # samples-count

    def test_samples_sheet_has_method_rows(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        wb = load_workbook(out)
        ws = wb["3. Samples"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "ID"
        methods = {row[1] for row in rows[1:] if row[1] is not None}
        assert "simple" in methods
        assert "stratified" in methods

    def test_audit_trail_in_chronological_order(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        wb = load_workbook(out)
        ws = wb["2. AuditTrail"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 events
        assert len(rows) == 3
        first_ts = str(rows[1][0])
        second_ts = str(rows[2][0])
        assert first_ts < second_ts

    def test_statistik_sheet_includes_chart_image(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        wb = load_workbook(out)
        ws = wb["4. Statistiken"]
        # Mindestens ein eingebettetes Bild im Sheet.
        assert len(ws._images) >= 1

    def test_atomic_write_no_tmp_left(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        # Kein .tmp-Rest, nur die finale .xlsx.
        leftovers = list(tmp_path.glob("*.tmp"))
        assert leftovers == []
        assert out.exists()

    def test_multi_report_atomic_no_partial_on_replace_error(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        """Schließt die A-004-Lücke: vorher lag `os.replace` außerhalb jedes
        try/except/finally – ein Fehlschlag (z. B. Ziel in Excel geöffnet)
        ließ die Tmp-Datei liegen und warf einen rohen `OSError`. Jetzt:
        aufgeräumt + als `ExportError` (Parität zu exporter.py)."""
        out = tmp_path / "bericht.xlsx"
        with (
            patch(
                "sampling_tool.io._atomic.os.replace",
                side_effect=PermissionError("target locked"),
            ),
            pytest.raises(ExportError, match="Excel geöffnet"),
        ):
            MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        assert not out.exists()
        leftovers = list(tmp_path.glob("*.tmp"))
        assert leftovers == [], f"Kein .tmp-Rest erwartet, gefunden: {leftovers}"

    def test_only_uebersicht_when_sheets_filtered(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "subset.xlsx"
        MultiSheetReportExporter().export(
            engagement,
            datasets,
            samples,
            audit_events,
            out,
            sheets={"Übersicht"},
        )
        wb = load_workbook(out)
        names = wb.sheetnames
        assert len(names) == 1
        assert "Übersicht" in names[0]

    def test_formula_injection_neutralized_all_sheets(
        self,
        tmp_path: Path,
        datasets: list[Dataset],
    ) -> None:
        """S-001: bösartige Werte in Engagement-/Auditor-/Event-/Dateinamen-
        und Sample-Feldern dürfen nach Reopen in KEINEM Sheet als Formel
        gespeichert sein (`data_type == "f"`) – geprüft über ALLE Zellen
        aller Sheets, nicht nur die gezielt injizierten.

        Echtes End-to-End-Regressionssignal liefert hier nur `formula_payload`
        (`=1+1`): openpyxl setzt `data_type == "f"` ausschließlich bei einem
        führenden `=` (verifiziert gegen openpyxl 3.1.5), `plus_payload` wäre
        schon ohne jede Neutralisierung `data_type == "s"`. Der Mechanismus,
        der `+`/`-`/`@`/Tab-Präfixe neutralisiert, ist isoliert in
        `tests/unit/test_xlsx_safe.py::TestSafeRow` regressionsgetestet;
        `plus_payload` bleibt hier trotzdem drin, um die byte-identische
        Werterhaltung über den gesamten Multi-Sheet-Export-Pfad zu sichern."""
        formula_payload = "=1+1"
        plus_payload = '+HYPERLINK("http://evil")'

        malicious_engagement = Engagement(
            auditor_name=formula_payload,
            client_name=plus_payload,
            auditor_position="Senior",
            audit_type="ISAE 3402",
            id=1,
        )
        malicious_cfg = SampleConfig(
            method=SamplingMethod.SIMPLE,
            size=1,
            seed=1,
            filter_field="Konto",
            filter_value=formula_payload,
            cluster_field=plus_payload,
            stratum_field=formula_payload,
        )
        malicious_samples = [
            SampleResult(
                config=malicious_cfg,
                selected_row_ids=(1,),
                population_size=1,
                drawn_at=datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
                created_by=plus_payload,
                id=1,
            ),
        ]
        malicious_events = [
            AuditEvent(
                event_type="export",
                engagement_id=1,
                user_name=formula_payload,
                sample_id=1,
                export_file=f"{plus_payload}.xlsx",
                timestamp=datetime(2026, 5, 1, 11, 0, tzinfo=UTC),
                id=1,
            ),
        ]

        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(
            malicious_engagement, datasets, malicious_samples, malicious_events, out
        )

        wb = load_workbook(out, data_only=False)
        assert len(wb.worksheets) >= 1
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    assert cell.data_type != "f", (
                        f"{ws.title}!{cell.coordinate} wurde als Formel gespeichert: {cell.value!r}"
                    )

    def test_subset_sheets_writes_exact_selection(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        out = tmp_path / "subset.xlsx"
        MultiSheetReportExporter().export(
            engagement,
            datasets,
            samples,
            audit_events,
            out,
            sheets={"AuditTrail", "Samples"},
        )
        wb = load_workbook(out)
        names = wb.sheetnames
        assert len(names) == 2
        assert any("AuditTrail" in n for n in names)
        assert any("Samples" in n for n in names)
        assert not any("Übersicht" in n for n in names)
        assert not any("Statistiken" in n for n in names)

    def test_samples_sheet_zeigt_volle_provenienz(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        audit_events: list[AuditEvent],
    ) -> None:
        """A-001 Contract-Test: Operator, Parent, Algorithmus-Version,
        angeforderte UND tatsächliche Größe, Dataset-ID sind sichtbar."""
        cfg = SampleConfig(
            method=SamplingMethod.CLUSTER,
            size=5,
            seed=99,
            cluster_field="Land",
            filter_field="Betrag",
            filter_value=100,
            filter_operator=FilterOperator.GTE,
        )
        samples = [
            SampleResult(
                config=cfg,
                selected_row_ids=(1, 2, 3, 4, 5, 6, 7),
                population_size=10,
                parent_sample_id=17,
                created_by="anna",
                id=3,
            )
        ]
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(
            engagement,
            datasets,
            samples,
            audit_events,
            out,
            dataset_ids_by_sample={3: 1},
        )
        wb = load_workbook(out)
        ws = wb["3. Samples"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        row = dict(zip(header, rows[1], strict=True))
        assert row["ID"] == 3
        assert row["Methode"] == "cluster"
        assert row["Angeforderte Größe"] == 5
        assert row["Tatsächliche Größe"] == 7
        assert row["Filter-Operator"] == "≥"
        assert row["Parent-Sample-ID"] == 17
        assert row["Algorithmus-Version"] == "bdo-v1"
        assert row["Dataset-ID"] == 1

    def test_audit_trail_details_spalte_zeigt_details_json(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
    ) -> None:
        events = [
            AuditEvent(
                event_type="sampling",
                engagement_id=1,
                user_name="anna",
                sample_id=1,
                details={"filter_operator": "gte", "algorithm_version": "bdo-v1"},
                timestamp=datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
                id=1,
            )
        ]
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, events, out)
        wb = load_workbook(out)
        ws = wb["2. AuditTrail"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][-1] == "Details"
        assert "filter_operator" in rows[1][-1]
        assert "gte" in rows[1][-1]

    def test_audit_trail_details_spalte_zeigt_dash_fuer_leere_details(
        self,
        tmp_path: Path,
        engagement: Engagement,
        datasets: list[Dataset],
        samples: list[SampleResult],
        audit_events: list[AuditEvent],
    ) -> None:
        """Alt-Events ohne `details` (alle Nicht-Sampling-Events, sowie jedes
        Event vor diesem Sprint) zeigen `"—"`, nicht leer/None."""
        out = tmp_path / "bericht.xlsx"
        MultiSheetReportExporter().export(engagement, datasets, samples, audit_events, out)
        wb = load_workbook(out)
        ws = wb["2. AuditTrail"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][-1] == "—"
