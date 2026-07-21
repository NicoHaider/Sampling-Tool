"""Integration: ExcelExporter – Sample → .xlsx mit Metadaten-Sheet."""

from __future__ import annotations

from collections.abc import Iterator
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from sampling_tool import __version__
from sampling_tool.core.models import (
    Dataset,
    DatasetRow,
    Engagement,
    FilterOperator,
    SampleConfig,
    SampleResult,
    SamplingMethod,
)
from sampling_tool.io.exporter import ExcelExporter, ExportError
from sampling_tool.persistence.database import Database
from sampling_tool.persistence.repositories import (
    DatasetRepo,
    EngagementRepo,
)


@pytest.fixture
def rows() -> tuple[DatasetRow, ...]:
    return tuple(
        DatasetRow(
            row_id=i,
            values={
                "Name": f"Posten {i}",
                "Betrag": 100 + i,
                "Land": "AUT" if i % 2 == 0 else "DEU",
                "Datum": datetime(2026, 1, i, 9, 0, 0),
            },
        )
        for i in range(1, 11)
    )


@pytest.fixture
def db() -> Iterator[Database]:
    database = Database(Path(":memory:"))
    database.migrate()
    yield database
    database.close()


@pytest.fixture
def dataset_with_repo(db: Database, rows: tuple[DatasetRow, ...]) -> tuple[Dataset, DatasetRepo]:
    """Persistiert das Test-Dataset in einer In-Memory-DB und liefert
    (Dataset, DatasetRepo) – die neue Sprint-11.4-API."""
    eng = EngagementRepo(db.connect()).get_or_create(
        Engagement(
            auditor_name="Anna Auditorin",
            client_name="ACME GmbH",
            auditor_position="Senior Auditor",
            audit_type="ISAE 3402 Typ II",
        )
    )
    assert eng.id is not None
    repo = DatasetRepo(db.connect())
    stored = repo.create(
        Dataset(
            name="TestData",
            columns=("Name", "Betrag", "Land", "Datum"),
            row_count=len(rows),
            source_file="/tmp/source.xlsx",
            engagement_id=eng.id,
        ),
        rows,
    )
    return stored, repo


@pytest.fixture
def dataset(dataset_with_repo: tuple[Dataset, DatasetRepo]) -> Dataset:
    return dataset_with_repo[0]


@pytest.fixture
def dataset_repo(dataset_with_repo: tuple[Dataset, DatasetRepo]) -> DatasetRepo:
    return dataset_with_repo[1]


@pytest.fixture
def sample(dataset: Dataset) -> SampleResult:
    cfg = SampleConfig(
        method=SamplingMethod.SIMPLE,
        size=4,
        seed=42,
    )
    return SampleResult(
        config=cfg,
        selected_row_ids=(1, 3, 5, 7),
        population_size=len(dataset),
    )


@pytest.fixture
def engagement() -> Engagement:
    return Engagement(
        auditor_name="Anna Auditorin",
        client_name="ACME GmbH",
        auditor_position="Senior Auditor",
        audit_type="ISAE 3402 Typ II",
        id=1,
    )


@pytest.fixture
def exporter() -> ExcelExporter:
    return ExcelExporter()


class TestExportSample:
    def test_dateiname_folgt_vba_schema(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name", "Betrag"],
            output_dir=tmp_path,
            custom_name="NewHires_Q2_2026",
            custom_id="001",
        )
        assert out.exists()
        assert out.name.startswith("NewHires_Q2_2026_ID001_BDO_sampling_")
        assert out.name.endswith(".xlsx")
        # Datum im Namen plausibel (8 Ziffern)
        date_token = out.name.split("_BDO_sampling_")[1].replace(".xlsx", "")
        assert len(date_token) == 8
        assert date_token.isdigit()

    def test_sample_sheet_enthaelt_genau_die_gewaehlten_spalten(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Land", "Name"],  # bewusst andere Reihenfolge
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
        )
        wb = load_workbook(out)
        ws = wb["Sample"]
        header = [c.value for c in ws[1]]
        assert header == ["Land", "Name"]
        # Zeilenanzahl = sample.actual_size (4) + 1 Header
        assert ws.max_row == 5
        # Werte aus row_id=1: Land="DEU", Name="Posten 1"
        first_data = [c.value for c in ws[2]]
        assert first_data == ["DEU", "Posten 1"]

    def test_metadaten_sheet_enthaelt_seed_und_engagement(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        engagement: Engagement,
        tmp_path: Path,
    ) -> None:
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name"],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
            engagement=engagement,
        )
        wb = load_workbook(out)
        ws = wb["Metadaten"]
        meta = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
        assert meta["Seed"] == "42"
        assert meta["Angeforderte Größe"] == "4"
        assert meta["Tatsächliche Größe"] == "4"
        assert meta["Population (Zeilen)"] == "10"
        assert meta["Sampling-Methode"] == "simple"
        assert meta["Auditor"] == "Anna Auditorin"
        assert meta["Mandant"] == "ACME GmbH"

    def test_metadaten_sheet_zeigt_volle_provenienz(
        self,
        exporter: ExcelExporter,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        """A-001 Contract-Test: Operator, Parent, Algorithmus-/App-Version,
        angeforderte UND tatsächliche Größe sind sichtbar."""
        cfg = SampleConfig(
            method=SamplingMethod.CLUSTER,
            size=5,
            seed=99,
            cluster_field="Land",
            filter_field="Betrag",
            filter_value=100,
            filter_operator=FilterOperator.GTE,
        )
        sample = SampleResult(
            config=cfg,
            selected_row_ids=(1, 2, 3, 4, 5, 6, 7),
            population_size=10,
            parent_sample_id=17,
            created_by="anna",
        )
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name"],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
        )
        wb = load_workbook(out)
        ws = wb["Metadaten"]
        meta = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
        assert meta["Angeforderte Größe"] == "5"
        assert meta["Tatsächliche Größe"] == "7"
        assert meta["Filter-Operator"] == "≥"
        assert meta["Parent-Sample-ID"] == "17"
        assert meta["Algorithmus-Version"] == "bdo-v1"
        assert meta["App-Version"] == __version__
        assert meta["Erstellt von"] == "anna"

    def test_atomic_write_kein_halbes_file_bei_exception(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        # openpyxl-Save soll fehlschlagen → die Tmp-Datei muss verschwinden,
        # die Ziel-Datei darf gar nicht erst entstehen.
        with (
            patch("sampling_tool.io.exporter.Workbook.save", side_effect=OSError("disk full")),
            pytest.raises(OSError, match="disk full"),
        ):
            exporter.export_sample(
                sample=sample,
                dataset=dataset,
                dataset_repo=dataset_repo,
                columns=["Name"],
                output_dir=tmp_path,
                custom_name="X",
                custom_id="1",
            )
        leftover = list(tmp_path.iterdir())
        assert leftover == [], f"Es sollten keine Dateien übrig bleiben, gefunden: {leftover}"

    def test_exporter_wraps_os_replace_error(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        """N-004: schlägt der finale `os.replace` fehl (z. B. Zieldatei in
        Excel geöffnet, Windows-`PermissionError`), muss das ein deutscher,
        fachlicher `ExportError` sein statt eines rohen `OSError` – und die
        Tmp-Datei darf nicht liegen bleiben."""
        with (
            patch(
                "sampling_tool.io._atomic.os.replace",
                side_effect=PermissionError("target locked"),
            ),
            pytest.raises(ExportError, match="Excel geöffnet"),
        ):
            exporter.export_sample(
                sample=sample,
                dataset=dataset,
                dataset_repo=dataset_repo,
                columns=["Name"],
                output_dir=tmp_path,
                custom_name="X",
                custom_id="1",
            )
        leftover = list(tmp_path.iterdir())
        assert leftover == [], f"Tmp-Datei sollte aufgeräumt sein, gefunden: {leftover}"

    def test_spaltenbreiten_sind_gesetzt(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name", "Betrag"],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
        )
        wb = load_workbook(out)
        ws = wb["Sample"]
        assert ws.column_dimensions["A"].width is not None
        assert ws.column_dimensions["A"].width >= 8
        # Spalte mit "Posten 10" (9 Zeichen) → mindestens 9 + 2
        assert ws.column_dimensions["A"].width >= 9

    def test_header_ist_gefettet_und_gefaerbt(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name"],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
        )
        wb = load_workbook(out)
        ws = wb["Sample"]
        cell = ws["A1"]
        assert cell.font.bold is True
        # BDO_RED = #E81A3B → ARGB FFE81A3B
        assert cell.fill.start_color.rgb == "FFE81A3B"
        # Weiße Schrift
        assert cell.font.color.rgb == "FFFFFFFF"

    def test_umlaute_im_pfad_funktionieren(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        umlaut_dir = tmp_path / "Prüfung_Müller_2026"
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name"],
            output_dir=umlaut_dir,
            custom_name="Stichprobe_März",
            custom_id="042",
        )
        assert out.exists()
        assert "Prüfung_Müller_2026" in str(out)
        assert "Stichprobe_März" in out.name

    def test_progress_callback(
        self,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        events: list[tuple[int, int]] = []
        exp = ExcelExporter(progress=lambda c, t: events.append((c, t)))
        exp.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name"],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
        )
        # 4 Zeilen → 4 Ticks
        assert len(events) == 4
        assert events[-1] == (4, 4)

    def test_unbekannte_spalte_wirft_export_error(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        with pytest.raises(ExportError, match="existieren nicht"):
            exporter.export_sample(
                sample=sample,
                dataset=dataset,
                dataset_repo=dataset_repo,
                columns=["GibtsNicht"],
                output_dir=tmp_path,
                custom_name="X",
                custom_id="1",
            )

    @pytest.mark.parametrize(
        "payload",
        ["=1+1", "+1+1", "-1+1", "@SUM(A1)", "\t=1+1", "=cmd|' /C calc'!A0"],
    )
    def test_formula_injection_neutralized(
        self,
        db: Database,
        tmp_path: Path,
        payload: str,
    ) -> None:
        """S-001: ein bösartiger Wert gleichzeitig als Datenwert, Spaltenname,
        UND Auditor-/Mandantenname darf nach Reopen NIRGENDS als
        Formel gespeichert sein (`data_type == "f"`), und der Wert muss exakt
        erhalten bleiben (keine stille Mutation, z. B. kein Apostroph-Prefix).

        Echtes End-to-End-Regressionssignal für den `data_type`-Assert liefern
        NUR die beiden `=`-Payloads: openpyxls `Cell._bind_value` setzt
        `data_type == "f"` ausschließlich bei einem führenden `=` (verifiziert
        gegen openpyxl 3.1.5) – für `+`/`-`/`@`/Tab-Präfixe ist `data_type`
        bereits ohne jede Neutralisierung `"s"`, der Assert wäre für diese
        Payloads auch ohne den Fix grün. Der Mechanismus, der diese
        defense-in-depth-Präfixe neutralisiert (`is_dangerous`-Gate in
        `safe_row`), ist stattdessen isoliert in
        `tests/unit/test_xlsx_safe.py::TestSafeRow` regressionsgetestet. Hier
        bleiben alle Payloads trotzdem im Battery: sie sichern zusätzlich die
        byte-identische Werterhaltung über den vollen Export-Pfad ab (Spalten-
        name, Datenwert, Auditor/Mandant – `payload_cells_seen`)."""
        malicious_column = payload
        eng = EngagementRepo(db.connect()).get_or_create(
            Engagement(
                auditor_name=payload,
                client_name=payload,
                auditor_position="Senior Auditor",
                audit_type="ISAE 3402 Typ II",
            )
        )
        assert eng.id is not None
        repo = DatasetRepo(db.connect())
        rows = (
            DatasetRow(row_id=1, values={"Name": "Posten 1", malicious_column: payload}),
            DatasetRow(row_id=2, values={"Name": "Posten 2", malicious_column: "harmlos"}),
        )
        dataset = repo.create(
            Dataset(
                name="MaliciousData",
                columns=("Name", malicious_column),
                row_count=len(rows),
                source_file="/tmp/source.xlsx",
                engagement_id=eng.id,
            ),
            rows,
        )
        cfg = SampleConfig(
            method=SamplingMethod.SIMPLE,
            size=2,
            seed=42,
        )
        sample = SampleResult(
            config=cfg,
            selected_row_ids=(1, 2),
            population_size=len(dataset),
        )
        engagement = Engagement(
            auditor_name=payload,
            client_name=payload,
            auditor_position="Senior Auditor",
            audit_type="ISAE 3402 Typ II",
            id=eng.id,
        )
        exporter = ExcelExporter()
        out = exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=repo,
            columns=["Name", malicious_column],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
            engagement=engagement,
        )

        wb = load_workbook(out, data_only=False)
        payload_cells_seen = 0
        for sheet_name in ("Sample", "Metadaten"):
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == payload:
                        payload_cells_seen += 1
                        assert cell.data_type != "f", (
                            f"{sheet_name}!{cell.coordinate} wurde als Formel gespeichert: "
                            f"{cell.value!r}"
                        )

        # Payload muss an mehreren Stellen (Spaltenname, Datenwert, Auditor,
        # Mandant) exakt und byte-identisch wiedergefunden werden.
        assert payload_cells_seen >= 4

    def test_leere_spaltenliste_wirft_export_error(
        self,
        exporter: ExcelExporter,
        sample: SampleResult,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        with pytest.raises(ExportError, match="Mindestens eine"):
            exporter.export_sample(
                sample=sample,
                dataset=dataset,
                dataset_repo=dataset_repo,
                columns=[],
                output_dir=tmp_path,
                custom_name="X",
                custom_id="1",
            )

    def test_streaming_loads_only_sample_rows_not_all(
        self,
        exporter: ExcelExporter,
        dataset: Dataset,
        dataset_repo: DatasetRepo,
        tmp_path: Path,
    ) -> None:
        """Sprint 11.4: Exporter darf NUR get_rows_by_ids aufrufen,
        nicht get_all_rows (= keinen voll-materialisierten Load)."""
        cfg = SampleConfig(method=SamplingMethod.SIMPLE, size=2, seed=1)
        sample = SampleResult(
            config=cfg,
            selected_row_ids=(2, 4),
            population_size=10,
        )

        get_all_calls: list[int] = []
        get_by_ids_calls: list[list[int]] = []
        original_get_all = dataset_repo.get_all_rows
        original_get_by_ids = dataset_repo.get_rows_by_ids

        def track_get_all(ds_id: int) -> tuple[DatasetRow, ...]:
            get_all_calls.append(ds_id)
            return original_get_all(ds_id)

        def track_get_by_ids(ds_id: int, ids: list[int]) -> list[DatasetRow]:
            get_by_ids_calls.append(list(ids))
            return original_get_by_ids(ds_id, ids)

        dataset_repo.get_all_rows = track_get_all  # type: ignore[assignment]
        dataset_repo.get_rows_by_ids = track_get_by_ids  # type: ignore[assignment]

        exporter.export_sample(
            sample=sample,
            dataset=dataset,
            dataset_repo=dataset_repo,
            columns=["Name"],
            output_dir=tmp_path,
            custom_name="X",
            custom_id="1",
        )

        assert get_all_calls == [], "Exporter sollte NICHT get_all_rows aufrufen"
        assert get_by_ids_calls == [[2, 4]], (
            "Exporter sollte get_rows_by_ids genau mit den Sample-IDs aufrufen"
        )
