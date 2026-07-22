"""Tests für den gemeinsamen xlsx-Spaltenbreiten-Helfer (Sprint 60 / Q-003)."""

from __future__ import annotations

from datetime import date, datetime

import pytest
from openpyxl import Workbook

from sampling_tool.io._xlsx_util import autosize_columns, display_string

pytestmark = pytest.mark.unit


class TestDisplayString:
    def test_none_is_empty(self) -> None:
        assert display_string(None) == ""

    def test_datetime_uses_seconds_precision(self) -> None:
        assert display_string(datetime(2026, 1, 2, 9, 30, 15)) == "2026-01-02 09:30:15"

    def test_date_uses_isoformat(self) -> None:
        assert display_string(date(2026, 1, 2)) == "2026-01-02"

    def test_other_values_use_str(self) -> None:
        assert display_string(42) == "42"
        assert display_string("Foo") == "Foo"


class TestAutosizeColumns:
    def test_width_grows_with_longest_cell(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["kurz"])
        ws.append(["ein deutlich längerer Zellinhalt"])
        autosize_columns(ws, 1, min_width=8)
        assert ws.column_dimensions["A"].width == len("ein deutlich längerer Zellinhalt") + 2

    def test_min_width_is_respected(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["x"])
        autosize_columns(ws, 1, min_width=12)
        assert ws.column_dimensions["A"].width == 12

    def test_width_is_capped_at_fifty(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["x" * 200])
        autosize_columns(ws, 1, min_width=8)
        assert ws.column_dimensions["A"].width == 52  # 50 + 2


class TestAutosizeSharedHelper:
    """SSOT-Nachweis: `exporter.py` und `multi_report_exporter.py` nutzen
    dieselbe Funktion (Identity-Check), keine eigene Kopie mehr."""

    def test_exporter_and_multi_report_share_autosize_columns(self) -> None:
        import sampling_tool.io.exporter as exporter_mod
        import sampling_tool.io.multi_report_exporter as multi_mod

        assert exporter_mod.autosize_columns is autosize_columns
        assert multi_mod.autosize_columns is autosize_columns
        assert not hasattr(exporter_mod, "_autosize")
        assert not hasattr(multi_mod, "_autosize")
