"""Unit: `io._xlsx_safe` – Formel-Injection-Policy (S-001, Sprint 40 / S1.3)."""

from __future__ import annotations

from datetime import date, datetime
from unittest.mock import call, patch

import pytest
from openpyxl import Workbook
from openpyxl.cell.cell import TYPE_STRING

from sampling_tool.io._xlsx_safe import is_dangerous, safe_cell_value, safe_row


class TestIsDangerous:
    @pytest.mark.parametrize(
        "value",
        [
            "=1+1",
            "+1+1",
            "-1+1",
            "@SUM(A1)",
            "\t=1+1",
            "\rgefahr",
            "\ngefahr",
            "=cmd|' /C calc'!A0",
            "Zeile1\nZeile2",
            "Zeile1\rZeile2",
        ],
    )
    def test_gefaehrliche_strings_erkannt(self, value: str) -> None:
        assert is_dangerous(value) is True

    @pytest.mark.parametrize(
        "value",
        [
            "",
            "Posten 1",
            "a=b",
            "AUT",
            "5 - 3 = 2",
        ],
    )
    def test_harmlose_strings_nicht_erkannt(self, value: str) -> None:
        assert is_dangerous(value) is False

    def test_nicht_strings_sind_nie_gefaehrlich(self) -> None:
        assert is_dangerous(1) is False
        assert is_dangerous(1.5) is False
        assert is_dangerous(True) is False
        assert is_dangerous(False) is False
        assert is_dangerous(None) is False
        assert is_dangerous(datetime(2026, 1, 1)) is False
        assert is_dangerous(date(2026, 1, 1)) is False


class TestSafeCellValue:
    def test_passthrough_typen_unveraendert(self) -> None:
        dt = datetime(2026, 1, 1, 9, 0, 0)
        d = date(2026, 1, 1)
        assert safe_cell_value(dt) is dt
        assert safe_cell_value(d) is d
        assert safe_cell_value(True) is True
        assert safe_cell_value(42) == 42
        assert safe_cell_value(3.14) == 3.14
        assert safe_cell_value("Posten 1") == "Posten 1"
        assert safe_cell_value(None) is None

    def test_gefaehrlicher_string_bleibt_beim_normalisieren_unveraendert(self) -> None:
        """Neutralisierung passiert erst beim Zell-Schreiben (safe_row), nicht hier."""
        assert safe_cell_value("=1+1") == "=1+1"

    def test_unbekannter_typ_wird_str(self) -> None:
        class Custom:
            def __str__(self) -> str:
                return "custom-repr"

        assert safe_cell_value(Custom()) == "custom-repr"


class TestSafeRow:
    """Mechanismus-Tests für `safe_row`, entkoppelt von openpyxls eigener
    Formel-Erkennung.

    Reopen-Tests auf Export-Ebene (`tests/integration/test_exporter.py`,
    `test_multi_report_exporter.py`) können die Neutralisierung nur für ein
    führendes `=` beobachten: openpyxls `Cell._bind_value` setzt
    `data_type == "f"` ausschließlich dafür (verifiziert gegen openpyxl
    3.1.5). Für die defense-in-depth-Präfixe `+`/`-`/`@`/Tab/CR/LF ist
    `data_type` nach einem `ws.append(...)` bereits ohne jede Neutralisierung
    `"s"` – ein reopen-Assert `data_type != "f"` liefert für diese Werte also
    kein Regressions-Signal (weder mit noch ohne den `is_dangerous`-Gate in
    `safe_row`). Die folgenden Tests prüfen daher direkt den Mechanismus:
    ruft `safe_row` `is_dangerous` für jeden Wert auf, und reagiert es
    korrekt (nur dann) mit einer expliziten `data_type = TYPE_STRING`-
    Zuweisung? Das ist unabhängig davon, was openpyxl von sich aus für den
    jeweiligen Wert gewählt hätte, und deckt damit auch die Präfixe ab, die
    über das hinausgehen, was openpyxl selbst als Formel erkennt.
    """

    def test_ruft_is_dangerous_fuer_jeden_wert_der_zeile_auf(self) -> None:
        ws = Workbook().active
        assert ws is not None
        with patch("sampling_tool.io._xlsx_safe.is_dangerous", wraps=is_dangerous) as spy:
            safe_row(ws, ["Posten 1", "=1+1", 42, None])
        assert spy.call_args_list == [
            call("Posten 1"),
            call("=1+1"),
            call(42),
            call(None),
        ]

    def test_erzwingt_type_string_wenn_is_dangerous_true_liefert(self) -> None:
        """Sonde mit einem für openpyxl völlig harmlosen Wert: `is_dangerous`
        wird auf `True` gemockt, obwohl openpyxl diesen Wert nie als Formel
        interpretieren würde. Trotzdem muss `safe_row` `data_type` explizit
        auf `TYPE_STRING` erzwingen – das beweist, dass die Zwangs-Zuweisung
        tatsächlich am `is_dangerous`-Ergebnis hängt, nicht an openpyxls
        eigener (Formel-)Erkennung."""
        ws = Workbook().active
        assert ws is not None
        with patch("sampling_tool.io._xlsx_safe.is_dangerous", return_value=True):
            safe_row(ws, ["Posten 1"])
        assert ws.cell(row=1, column=1).data_type == TYPE_STRING

    def test_erzwingt_type_string_nicht_wenn_is_dangerous_false_liefert(self) -> None:
        """Gegenprobe zum vorigen Test: `is_dangerous` liefert `False` für
        einen Wert, den openpyxl selbst als Formel erkennt (`=1+1` →
        `data_type == "f"` durch openpyxl selbst). Bleibt `data_type == "f"`
        erhalten, beweist das, dass `safe_row` NICHT blanket auf
        `TYPE_STRING` zwingt, sondern exakt dem `is_dangerous`-Gate folgt
        (kein toter/immer-True-Branch)."""
        ws = Workbook().active
        assert ws is not None
        with patch("sampling_tool.io._xlsx_safe.is_dangerous", return_value=False):
            safe_row(ws, ["=1+1"])
        assert ws.cell(row=1, column=1).data_type == "f"

    @pytest.mark.parametrize(
        "payload",
        ["=1+1", "+1+1", "-1+1", "@SUM(A1)", "\t=1+1", "\rgefahr", "\ngefahr"],
    )
    def test_end_to_end_fuer_alle_gefaehrlichen_praefixe(self, payload: str) -> None:
        """Kombiniert mit den beiden Tests oben (Gate wird für jeden Wert
        befragt + reagiert exakt auf dessen Ergebnis) UND den `is_dangerous`-
        Korrektheits-Tests in `TestIsDangerous` ist das hier die
        End-to-End-Komposition: reales `is_dangerous`, reales `safe_row`,
        alle sieben gefährlichen Präfixe landen als `TYPE_STRING`."""
        ws = Workbook().active
        assert ws is not None
        safe_row(ws, [payload])
        assert ws.cell(row=1, column=1).data_type == TYPE_STRING
        assert ws.cell(row=1, column=1).value == payload
